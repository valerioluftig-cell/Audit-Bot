"""
Investigate why LaSalle, St. Martin, Avoyelles-REISSUE have 0 comparison flags.
For each: show what the pipeline extracted vs what's in the coded Excel.
"""
import json, sys, pathlib
sys.path.insert(0, r'C:\Users\valer\Downloads\audit_project\pipeline')
from coded_compare import parse_cbs_sheet, parse_sona_sheet, parse_ca_sheet, parse_soa_sheet, _num
import openpyxl

coded_dir = pathlib.Path(r'C:\Users\valer\Downloads\audit_project\coded\2013')
wb_cbs  = openpyxl.load_workbook(str(coded_dir / 'Louisiana 2013 CBS.xlsx'),  data_only=True)
wb_sona = openpyxl.load_workbook(str(coded_dir / 'Louisiana 2013 SONA.xlsx'), data_only=True)
wb_ca   = openpyxl.load_workbook(str(coded_dir / 'Louisiana 2013 Capital Assets.xlsx'), data_only=True)
wb_soa  = openpyxl.load_workbook(str(coded_dir / 'Louisiana 2013 SOA.xlsx'),  data_only=True)

cache_dir = pathlib.Path(r'C:\Users\valer\Downloads\audit_project\runs\e44ac3c9\output\cache')

ZERO_PARISHES = ['LaSalle', 'St. Martin', 'Avoyelles - REISSUE']

for parish in ZERO_PARISHES:
    jf = cache_dir / f'{parish}_2013.json'
    if not jf.exists():
        print(f'\n{parish}: NO CACHE FILE'); continue

    cache = json.loads(jf.read_text())
    print(f'\n{"="*60}')
    print(f'PARISH: {parish}')
    print(f'  Statements in cache: { {k: (type(v).__name__ if v else "None") for k,v in cache.items() if k in ["cbs","sona","ca","soa"]} }')

    # CBS
    cbs_ws = wb_cbs[parish] if parish in wb_cbs.sheetnames else None
    if cbs_ws and cache.get('cbs'):
        coded = parse_cbs_sheet(cbs_ws)
        pipe = cache['cbs']
        pa = (pipe.get('assets') or {})
        pl = (pipe.get('liabilities') or {})
        pfb = (pipe.get('fund_balances') or {})
        pipe_ta = ((pa.get('total_assets') or {}).get('Total Governmental Funds') or pa.get('total_assets_total'))
        pipe_tl = ((pl.get('total_liabilities') or {}).get('Total Governmental Funds') or pl.get('total_liabilities_total'))
        fb_raw = pfb.get('total_fund_balances') or pfb.get('total_fund_balance')
        pipe_fb = fb_raw.get('Total Governmental Funds') if isinstance(fb_raw, dict) else fb_raw
        print(f'  CBS coded  assets={coded["total_assets"]}  liab={coded["total_liabilities"]}  fb={coded["total_fund_balance"]}')
        print(f'  CBS pipe   assets={pipe_ta}  liab={pipe_tl}  fb={pipe_fb}')
    elif not cbs_ws:
        print(f'  CBS: no coded sheet')

    # SONA
    sona_ws = wb_sona[parish] if parish in wb_sona.sheetnames else None
    if sona_ws and cache.get('sona'):
        coded = parse_sona_sheet(sona_ws)
        pipe_ga = (cache['sona'].get('governmental_activities') or {})
        print(f'  SONA coded assets={coded["total_assets"]}  liab={coded["total_liabilities"]}  net={coded["total_net_assets"]}')
        print(f'  SONA pipe  assets={pipe_ga.get("total_assets")}  liab={pipe_ga.get("total_liabilities")}  net={pipe_ga.get("total_net_position") or pipe_ga.get("net_position")}')
    elif not sona_ws:
        print(f'  SONA: no coded sheet')

    # CA
    ca_ws = wb_ca[parish] if parish in wb_ca.sheetnames else None
    if ca_ws and cache.get('ca'):
        coded = parse_ca_sheet(ca_ws)
        pipe_ga = (cache['ca'].get('governmental_activities') or {})
        pipe_net = pipe_ga.get('total_capital_assets_net') or {}
        print(f'  CA coded   beg={coded["total_capital_assets_net_beginning"]}  end={coded["total_capital_assets_net_ending"]}')
        print(f'  CA pipe    beg={pipe_net.get("beginning") if isinstance(pipe_net,dict) else pipe_net}  end={pipe_net.get("ending") if isinstance(pipe_net,dict) else None}')

    # SOA
    soa_ws = wb_soa[parish] if parish in wb_soa.sheetnames else None
    if soa_ws and cache.get('soa'):
        coded = parse_soa_sheet(soa_ws)
        pipe = cache['soa']
        gen_rev = (pipe.get('general_revenues') or {}).get('total_general_revenues')
        change  = (pipe.get('governmental_activities') or {}).get('change_in_net_assets') or pipe.get('change_in_net_assets')
        print(f'  SOA coded  gen_rev={coded["total_general_revenues"]}  change={coded["change_in_net_assets"]}')
        print(f'  SOA pipe   gen_rev={gen_rev}  change={change}')
