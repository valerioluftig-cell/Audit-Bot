"""
Financial validation engine for the Louisiana Parish Audit Pipeline.

Runs accounting identity checks on Claude-extracted statement data.
Flag-and-write: never blocks output, always records status for human review.

Statuses:
  PASS             — all checks pass within rounding tolerance
  WARNING          — minor rounding differences or optional fields missing
  REVIEW REQUIRED  — significant imbalance (> $1,000) or key fields absent
  FAILED           — multiple critical checks fail / critical fields all missing
"""

import csv
import os
from dataclasses import dataclass, field
from typing import Optional

# ── Tolerance ──────────────────────────────────────────────────────────────────
# Government statements are exact-dollar; tolerate up to $1 for rounding.
# Differences > WARN_THRESHOLD flag as WARNING; > FAIL_THRESHOLD → REVIEW REQUIRED.
ROUND_TOLERANCE  = 1       # $1  → PASS (pure rounding)
WARN_THRESHOLD   = 1_000   # $1k → WARNING
FAIL_THRESHOLD   = 10_000  # $10k → REVIEW REQUIRED


# ── Data structures ────────────────────────────────────────────────────────────

@dataclass
class CheckResult:
    name: str
    passed: bool            # True = within round tolerance
    lhs_label: str
    rhs_label: str
    lhs_value: Optional[int]
    rhs_value: Optional[int]
    difference: Optional[int]   # lhs - rhs; None if either side is None
    severity: str               # PASS / WARNING / REVIEW REQUIRED
    note: str = ""


@dataclass
class StatementValidation:
    stmt_type: str          # cbs / soa / sona / ca
    checks: list            # list[CheckResult]
    missing_key_fields: list
    status: str             # PASS / WARNING / REVIEW REQUIRED / FAILED
    confidence: str         # HIGH / MEDIUM / LOW
    confidence_pct: float   # 0.0 – 1.0


@dataclass
class ParishValidation:
    parish: str
    year: int
    statements: dict        # stmt_type → StatementValidation
    overall_status: str     # worst status across all statements


# ── Helpers ────────────────────────────────────────────────────────────────────

def _safe(d, *keys):
    """Safe nested dict access; returns None if any key is missing."""
    for k in keys:
        if not isinstance(d, dict):
            return None
        d = d.get(k)
    return d


def _coerce(v) -> Optional[int]:
    """Return int if numeric, else None."""
    if isinstance(v, (int, float)) and v is not None:
        return int(v)
    return None


def _diff_severity(diff: Optional[int]) -> str:
    if diff is None:
        return "REVIEW REQUIRED"
    abs_diff = abs(diff)
    if abs_diff <= ROUND_TOLERANCE:
        return "PASS"
    if abs_diff <= WARN_THRESHOLD:
        return "WARNING"
    return "REVIEW REQUIRED"


def _make_check(name: str, lhs_label: str, lhs: Optional[int],
                rhs_label: str, rhs: Optional[int], note: str = "") -> CheckResult:
    if lhs is None or rhs is None:
        return CheckResult(
            name=name, passed=False,
            lhs_label=lhs_label, rhs_label=rhs_label,
            lhs_value=lhs, rhs_value=rhs,
            difference=None,
            severity="REVIEW REQUIRED",
            note=note or "One or more values missing — cannot evaluate",
        )
    diff = lhs - rhs
    severity = _diff_severity(diff)
    return CheckResult(
        name=name, passed=(severity == "PASS"),
        lhs_label=lhs_label, rhs_label=rhs_label,
        lhs_value=lhs, rhs_value=rhs,
        difference=diff,
        severity=severity,
        note=note or (f"Difference: ${diff:+,}" if severity != "PASS" else ""),
    )


def _worst_status(statuses: list[str]) -> str:
    order = ["PASS", "WARNING", "REVIEW REQUIRED", "FAILED"]
    worst = 0
    for s in statuses:
        if s in order:
            worst = max(worst, order.index(s))
    return order[worst]


def _checks_to_status(checks: list[CheckResult], missing: list[str]) -> str:
    severities = [c.severity for c in checks]
    if not checks and not missing:
        return "WARNING"  # no checks could run
    if all(s == "PASS" for s in severities) and not missing:
        return "PASS"
    if all(s in ("PASS", "WARNING") for s in severities):
        return "WARNING"
    # Count REVIEW REQUIRED
    review_count = sum(1 for s in severities if s == "REVIEW REQUIRED")
    if review_count >= 2 or (review_count >= 1 and missing):
        return "FAILED"
    return "REVIEW REQUIRED"


# ── Statement-level validators ─────────────────────────────────────────────────

def _validate_sona(data: dict) -> StatementValidation:
    """
    Statement of Net Position checks:
    1. Assets + Deferred Outflows = Liabilities + Deferred Inflows + Net Position
    2. Net Position components sum to total net position
    """
    ga = data.get("governmental_activities", {}) or {}
    np_data = ga.get("net_position", {}) or {}
    do_data = ga.get("deferred_outflows", {}) or {}
    di_data = ga.get("deferred_inflows", {}) or {}

    total_assets   = _coerce(ga.get("total_assets"))
    total_liab     = _coerce(ga.get("total_liabilities"))
    total_np       = _coerce(np_data.get("total_net_position"))
    deferred_out   = _coerce(do_data.get("total")) or 0
    deferred_in    = _coerce(di_data.get("total")) or 0

    net_invest     = _coerce(np_data.get("net_investment_in_capital_assets"))
    restricted     = _coerce(np_data.get("restricted"))
    unrestricted   = _coerce(np_data.get("unrestricted"))

    checks = []

    # Check 1: Balance sheet identity
    lhs = (total_assets or 0) + deferred_out if total_assets is not None else None
    rhs = (total_liab or 0) + deferred_in + (total_np or 0) \
          if (total_liab is not None and total_np is not None) else None
    checks.append(_make_check(
        "SONA Balance Sheet Identity",
        "Total Assets + Deferred Outflows",
        lhs,
        "Total Liabilities + Deferred Inflows + Net Position",
        rhs,
    ))

    # Check 2: Net position components
    if net_invest is not None or restricted is not None or unrestricted is not None:
        components_sum = (net_invest or 0) + (restricted or 0) + (unrestricted or 0)
        checks.append(_make_check(
            "SONA Net Position Components",
            "Net Investment + Restricted + Unrestricted",
            components_sum,
            "Total Net Position",
            total_np,
        ))

    # Missing key fields
    key_fields = {
        "total_assets": total_assets,
        "total_liabilities": total_liab,
        "net_position.total_net_position": total_np,
    }
    missing = [k for k, v in key_fields.items() if v is None]

    # Confidence: how many of the main asset fields are populated?
    current_assets = ga.get("current_assets", {}) or {}
    expected_fields = [
        "cash_and_deposits", "total_assets", "total_liabilities",
        "net_position.total_net_position",
    ]
    filled = sum(1 for f in expected_fields
                 if (current_assets.get(f) or ga.get(f) or np_data.get(f.split(".")[-1])) is not None)
    conf_pct = filled / len(expected_fields)
    confidence = "HIGH" if conf_pct >= 0.75 else ("MEDIUM" if conf_pct >= 0.5 else "LOW")

    status = _checks_to_status(checks, missing)
    return StatementValidation("sona", checks, missing, status, confidence, conf_pct)


def _validate_soa(data: dict) -> StatementValidation:
    """
    Statement of Activities checks:
    1. Net Position Beginning + Change = Net Position Ending
    2. Total General Revenues + Net Expense Revenue = Change in Net Position
    """
    change   = _coerce(data.get("change_in_net_position"))
    beg      = _coerce(data.get("net_position_beginning"))
    ending   = _coerce(data.get("net_position_ending"))
    total_gr = _coerce(_safe(data, "general_revenues", "total_general_revenues"))
    ga_total = data.get("governmental_activities", {}) or {}
    total_ner = _coerce(_safe(ga_total, "total", "net_expense_revenue"))

    checks = []

    # Check 1: Roll-forward
    lhs = (beg or 0) + (change or 0) if (beg is not None and change is not None) else None
    checks.append(_make_check(
        "SOA Net Position Roll-Forward",
        "Beginning Balance + Change",
        lhs,
        "Ending Net Position",
        ending,
    ))

    # Check 2: Revenue - Expense = Change
    if total_gr is not None and total_ner is not None:
        implied_change = total_gr + total_ner   # total_ner is net expense (negative for net cost)
        checks.append(_make_check(
            "SOA Revenue-Expense = Change",
            "Total General Revenues + Net Program Revenue/(Expense)",
            implied_change,
            "Change in Net Position",
            change,
        ))

    key_fields = {
        "change_in_net_position": change,
        "net_position_beginning": beg,
        "net_position_ending": ending,
        "general_revenues.total_general_revenues": total_gr,
    }
    missing = [k for k, v in key_fields.items() if v is None]

    filled = sum(1 for v in key_fields.values() if v is not None)
    conf_pct = filled / len(key_fields)
    confidence = "HIGH" if conf_pct >= 0.75 else ("MEDIUM" if conf_pct >= 0.5 else "LOW")

    status = _checks_to_status(checks, missing)
    return StatementValidation("soa", checks, missing, status, confidence, conf_pct)


def _validate_cbs(data: dict) -> StatementValidation:
    """
    Combined Balance Sheet checks (using Total Governmental Funds column):
    1. Total Assets = Total Liabilities & Fund Balances
    2. Total Liabilities + Deferred Inflows + Total Fund Balances = Total L&FB
    """
    TGF = "Total Governmental Funds"

    def _tgf(path_parts):
        d = data
        for k in path_parts:
            if not isinstance(d, dict):
                return None
            d = d.get(k)
        if isinstance(d, dict):
            return _coerce(d.get(TGF))
        return None

    total_assets  = _tgf(["assets", "total_assets"])
    total_liab    = _tgf(["liabilities", "total_liabilities"])
    total_fb      = _tgf(["fund_balances", "total_fund_balances"])
    total_l_fb    = _coerce((data.get("total_liabilities_and_fund_balances") or {}).get(TGF))
    di_total      = _coerce((_safe(data, "deferred_inflows", "total") or {}).get(TGF)) or 0

    checks = []

    # Check 1: Assets = Total L&FB
    checks.append(_make_check(
        "CBS Assets = Liabilities & Fund Balances",
        "Total Assets",
        total_assets,
        "Total Liabilities & Fund Balances",
        total_l_fb,
    ))

    # Check 2: Liabilities + Deferred Inflows + Fund Balances = Total L&FB
    if total_liab is not None and total_fb is not None:
        lhs = total_liab + di_total + total_fb
        checks.append(_make_check(
            "CBS L + Deferred Inflows + FB = Total L&FB",
            "Total Liabilities + Deferred Inflows + Fund Balances",
            lhs,
            "Total Liabilities & Fund Balances",
            total_l_fb,
        ))

    key_fields = {
        f"assets.total_assets[{TGF}]": total_assets,
        f"liabilities.total_liabilities[{TGF}]": total_liab,
        f"fund_balances.total_fund_balances[{TGF}]": total_fb,
        f"total_liabilities_and_fund_balances[{TGF}]": total_l_fb,
    }
    missing = [k for k, v in key_fields.items() if v is None]

    filled = sum(1 for v in key_fields.values() if v is not None)
    conf_pct = filled / len(key_fields)
    confidence = "HIGH" if conf_pct >= 0.75 else ("MEDIUM" if conf_pct >= 0.5 else "LOW")

    status = _checks_to_status(checks, missing)
    return StatementValidation("cbs", checks, missing, status, confidence, conf_pct)


def _validate_ca(data: dict) -> StatementValidation:
    """
    Capital Assets checks:
    1. Beginning + Increases + Decreases = Ending (total_capital_assets_net)
       Note: Decreases are typically negative, so this is arithmetic addition.
    """
    ga = data.get("governmental_activities", {}) or {}
    net = ga.get("total_capital_assets_net", {}) or {}

    beg  = _coerce(net.get("beginning"))
    inc  = _coerce(net.get("increases"))
    dec  = _coerce(net.get("decreases"))
    end  = _coerce(net.get("ending"))

    checks = []

    if beg is not None and end is not None:
        inc_val = inc or 0
        dec_val = dec or 0
        implied_end = beg + inc_val + dec_val
        ca_note = "Decreases should be negative values; verify sign convention if flagged"

        # PDFs often show decreases as positive numbers. If the roll-forward fails
        # with raw values but succeeds when decreases are negated, auto-correct.
        if abs(implied_end - end) > ROUND_TOLERANCE and dec_val > 0:
            implied_end_neg = beg + inc_val - dec_val
            if abs(implied_end_neg - end) <= ROUND_TOLERANCE:
                implied_end = implied_end_neg
                ca_note = "Sign auto-corrected: decreases negated (PDF presented as positive)"

        checks.append(_make_check(
            "CA Roll-Forward (Beginning + Changes = Ending)",
            "Beginning + Increases + Decreases",
            implied_end,
            "Ending Balance",
            end,
            note=ca_note,
        ))

    key_fields = {
        "total_capital_assets_net.beginning": beg,
        "total_capital_assets_net.ending": end,
    }
    missing = [k for k, v in key_fields.items() if v is None]

    # Confidence: count non-null items across not_depreciated + depreciable sections
    non_dep = ga.get("not_depreciated", {}) or {}
    dep     = ga.get("depreciable", {}) or {}
    items_with_data = 0
    for section in [non_dep, dep]:
        for k, v in section.items():
            if isinstance(v, dict) and any(x is not None for x in v.values()):
                items_with_data += 1
    conf_pct = min(1.0, items_with_data / 4)  # expect at least 4 non-null items
    confidence = "HIGH" if conf_pct >= 0.75 else ("MEDIUM" if conf_pct >= 0.5 else "LOW")

    status = _checks_to_status(checks, missing)
    return StatementValidation("ca", checks, missing, status, confidence, conf_pct)


_VALIDATORS = {
    "sona": _validate_sona,
    "soa":  _validate_soa,
    "cbs":  _validate_cbs,
    "ca":   _validate_ca,
}

# ── Parish-level entry point ───────────────────────────────────────────────────

def validate_parish(parish: str, year: int, results: dict) -> ParishValidation:
    """
    Run all applicable validators for one parish.

    Args:
        parish:  Parish name string.
        year:    Fiscal year integer.
        results: Dict keyed by stmt_type ("cbs", "soa", "sona", "ca") with
                 extracted data dicts (may be None if extraction failed).

    Returns:
        ParishValidation with per-statement results and overall status.
    """
    statements = {}
    for stmt_type, validator in _VALIDATORS.items():
        data = results.get(stmt_type)
        if data:
            try:
                statements[stmt_type] = validator(data)
            except Exception as exc:
                # Never let validation crash the pipeline
                statements[stmt_type] = StatementValidation(
                    stmt_type=stmt_type,
                    checks=[],
                    missing_key_fields=[f"Validator error: {exc}"],
                    status="REVIEW REQUIRED",
                    confidence="LOW",
                    confidence_pct=0.0,
                )
        # If data is None → statement was not extracted; omit from validation

    statuses = [sv.status for sv in statements.values()] or ["WARNING"]
    overall = _worst_status(statuses)
    return ParishValidation(parish, year, statements, overall)


# ── Quality report writer ──────────────────────────────────────────────────────

_STATUS_ORDER = ["PASS", "WARNING", "REVIEW REQUIRED", "FAILED"]

QUALITY_REPORT_COLUMNS = [
    "Parish",
    "Statement_Type",
    "Detection_Method",
    "Checks_Run",
    "Checks_Passed",
    "Checks_Failed",
    "Missing_Key_Fields",
    "Validation_Errors",
    "Extraction_Confidence",
    "Confidence_Pct",
    "Status",
]


def _stmt_label(stmt_type: str) -> str:
    return {
        "cbs":  "Combined Balance Sheet",
        "soa":  "Statement of Activities",
        "sona": "Statement of Net Position",
        "ca":   "Capital Assets",
    }.get(stmt_type, stmt_type.upper())


def _validation_error_summary(checks: list) -> str:
    """Return semicolon-separated list of failed check names with differences."""
    errors = []
    for c in checks:
        if not c.passed:
            if c.difference is not None:
                errors.append(f"{c.name} (diff: ${c.difference:+,})")
            else:
                errors.append(f"{c.name} (missing data)")
    return "; ".join(errors) if errors else ""


def write_quality_report_csv(
    parish_validations: list,    # list[ParishValidation]
    output_dir: str,
    manual_pages: dict,          # parish → override dict; used for detection method
    year: int,
) -> str:
    """
    Write quality_report.csv to output_dir.
    Returns the full file path.
    """
    path = os.path.join(output_dir, f"quality_report_{year}.csv")
    rows = []

    for pv in parish_validations:
        for stmt_type in ["cbs", "soa", "sona", "ca"]:
            sv = pv.statements.get(stmt_type)
            if sv is None:
                # Statement not extracted — report as missing
                rows.append({
                    "Parish": pv.parish,
                    "Statement_Type": _stmt_label(stmt_type),
                    "Detection_Method": "manual" if pv.parish in manual_pages else "auto",
                    "Checks_Run": 0,
                    "Checks_Passed": 0,
                    "Checks_Failed": 0,
                    "Missing_Key_Fields": "Statement not extracted",
                    "Validation_Errors": "",
                    "Extraction_Confidence": "NONE",
                    "Confidence_Pct": "0%",
                    "Status": "FAILED",
                })
                continue

            checks_run    = len(sv.checks)
            checks_passed = sum(1 for c in sv.checks if c.passed)
            checks_failed = checks_run - checks_passed

            rows.append({
                "Parish": pv.parish,
                "Statement_Type": _stmt_label(stmt_type),
                "Detection_Method": "manual" if pv.parish in manual_pages else "auto",
                "Checks_Run": checks_run,
                "Checks_Passed": checks_passed,
                "Checks_Failed": checks_failed,
                "Missing_Key_Fields": "; ".join(sv.missing_key_fields) if sv.missing_key_fields else "",
                "Validation_Errors": _validation_error_summary(sv.checks),
                "Extraction_Confidence": sv.confidence,
                "Confidence_Pct": f"{sv.confidence_pct:.0%}",
                "Status": sv.status,
            })

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=QUALITY_REPORT_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)

    return path


def write_quality_report_excel_tab(wb, parish_validations: list, manual_pages: dict, year: int):
    """
    Add a 'Quality Report' sheet to an open openpyxl Workbook.
    Call this before saving the workbook.
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    sheet_name = "Quality Report"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # Palette
    _GREEN  = "FF92D050"   # light green
    _YELLOW = "FFFFFF00"   # yellow
    _ORANGE = "FFFFC000"   # amber
    _RED    = "FFFF0000"   # red
    _GRAY   = "FFD9D9D9"   # light gray for header

    STATUS_FILL = {
        "PASS":             PatternFill("solid", start_color=_GREEN),
        "WARNING":          PatternFill("solid", start_color=_YELLOW),
        "REVIEW REQUIRED":  PatternFill("solid", start_color=_ORANGE),
        "FAILED":           PatternFill("solid", start_color=_RED),
    }

    # Title
    title_cell = ws.cell(row=1, column=1, value=f"Quality Report — {year} Parish Audit Extraction")
    title_cell.font = Font(bold=True, name="Arial", size=12)
    ws.merge_cells("A1:K1")
    title_cell.alignment = Alignment(horizontal="center")

    # Header row
    headers = QUALITY_REPORT_COLUMNS
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=j, value=h.replace("_", " "))
        c.font = Font(bold=True, name="Arial", size=10)
        c.fill = PatternFill("solid", start_color=_GRAY)
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data rows
    row = 3
    for pv in parish_validations:
        for stmt_type in ["cbs", "soa", "sona", "ca"]:
            sv = pv.statements.get(stmt_type)

            if sv is None:
                status = "FAILED"
                row_data = [
                    pv.parish,
                    _stmt_label(stmt_type),
                    "manual" if pv.parish in manual_pages else "auto",
                    0, 0, 0,
                    "Statement not extracted",
                    "",
                    "NONE", "0%", status,
                ]
            else:
                checks_run    = len(sv.checks)
                checks_passed = sum(1 for c in sv.checks if c.passed)
                status = sv.status
                row_data = [
                    pv.parish,
                    _stmt_label(stmt_type),
                    "manual" if pv.parish in manual_pages else "auto",
                    checks_run,
                    checks_passed,
                    checks_run - checks_passed,
                    "; ".join(sv.missing_key_fields) if sv.missing_key_fields else "",
                    _validation_error_summary(sv.checks),
                    sv.confidence,
                    f"{sv.confidence_pct:.0%}",
                    status,
                ]

            for j, val in enumerate(row_data, 1):
                c = ws.cell(row=row, column=j, value=val)
                c.font = Font(name="Arial", size=10)
                c.alignment = Alignment(wrap_text=(j in (7, 8)))

            # Color the Status cell
            status_cell = ws.cell(row=row, column=11)
            if status in STATUS_FILL:
                status_cell.fill = STATUS_FILL[status]
                status_cell.font = Font(name="Arial", size=10, bold=True)

            row += 1

    # Column widths
    col_widths = [22, 26, 16, 10, 12, 11, 35, 45, 14, 12, 18]
    for j, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # Freeze top two rows
    ws.freeze_panes = "A3"


# ── Validation column helpers for Excel writers ────────────────────────────────

def get_validation_cells(pv: ParishValidation, stmt_type: str) -> dict:
    """
    Return dict with validation_status, validation_notes, review_required
    for use when adding extra columns to parish cross-sectional rows.
    """
    sv = pv.statements.get(stmt_type)
    if sv is None:
        return {
            "validation_status": "NOT EXTRACTED",
            "validation_notes": "",
            "review_required": "YES",
        }
    errors = _validation_error_summary(sv.checks)
    missing = "; ".join(sv.missing_key_fields)
    notes_parts = [p for p in [errors, missing] if p]
    return {
        "validation_status": sv.status,
        "validation_notes": "; ".join(notes_parts)[:200],  # cap length for Excel cell
        "review_required": "NO" if sv.status == "PASS" else "YES",
    }
