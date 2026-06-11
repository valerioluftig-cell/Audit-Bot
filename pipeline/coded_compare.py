"""
Coded Excel Comparison Engine
Compares pipeline-extracted JSON values against student-coded ground-truth Excels.
Generates two types of flags:
  source="comparison"    — pipeline value differs from coded value
  source="coding_check"  — coded data fails an accounting identity (student error)
"""

import json
import os
import re
from pathlib import Path
from typing import Any, Optional


# ── helpers ──────────────────────────────────────────────────────────────────

def _num(v) -> Optional[float]:
    """Coerce a cell value to float, or None."""
    if v is None:
        return None
    try:
        return float(str(v).replace(",", "").replace("$", "").strip())
    except Exception:
        return None


def _label_matches(cell_val, *keywords) -> bool:
    """True if cell text contains ALL keywords (case-insensitive)."""
    if cell_val is None:
        return False
    t = str(cell_val).lower().strip()
    return all(kw.lower() in t for kw in keywords)


def _pct_diff(a: float, b: float) -> float:
    """Percentage difference between a and b, relative to b."""
    if b == 0:
        return 0.0 if a == 0 else 100.0
    return abs(a - b) / abs(b) * 100.0


def _severity(pct: float) -> str:
    if pct >= 5.0:
        return "high"
    if pct >= 1.0:
        return "medium"
    return "low"


def _fmt(v) -> str:
    if v is None:
        return "None"
    try:
        return f"{float(v):,.0f}"
    except Exception:
        return str(v)


# ── CBS parser ────────────────────────────────────────────────────────────────

def parse_cbs_sheet(ws) -> dict:
    """
    Parse a single parish CBS sheet from the coded Excel.
    Returns dict with keys matching pipeline CBS JSON structure.
    """
    # Find fund header row — row where col 11+ contains non-numeric text
    fund_start_col = None
    total_col = None
    fund_header_row = None

    for r in range(1, min(10, ws.max_row + 1)):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and "total governmental" in v.lower():
                fund_start_col = 11  # always starts at col 11 per format
                total_col = c
                fund_header_row = r
                break
        if total_col:
            break

    if total_col is None:
        # Fallback: assume col 18
        fund_start_col = 11
        total_col = 18

    def _row_total(row_idx: int) -> Optional[float]:
        """Get Total Governmental Funds value: use total_col if populated, else sum fund cols."""
        direct = _num(ws.cell(row_idx, total_col).value)
        if direct is not None:
            return direct
        # Sum all fund columns before the total column
        vals = [_num(ws.cell(row_idx, c).value) for c in range(fund_start_col, total_col)]
        vals = [v for v in vals if v is not None]
        return sum(vals) if vals else None

    result = {
        "total_assets": None,
        "total_liabilities": None,
        "total_fund_balance": None,
        "total_liabilities_and_fund_balances": None,
        "deferred_inflows": None,
    }

    for r in range(1, ws.max_row + 1):
        # Collect label from cols 1-6
        label = " ".join(
            str(ws.cell(r, c).value or "").strip()
            for c in range(1, 7)
        ).lower().strip()
        label = re.sub(r"\s+", " ", label)

        if result["total_assets"] is None and "total assets" in label:
            result["total_assets"] = _row_total(r)
        elif result["total_liabilities"] is None and "total liabilities" in label and "fund" not in label:
            result["total_liabilities"] = _row_total(r)
        elif result["total_fund_balance"] is None and "total fund balance" in label:
            result["total_fund_balance"] = _row_total(r)
        elif result["total_liabilities_and_fund_balances"] is None and "total liabilities and fund" in label:
            result["total_liabilities_and_fund_balances"] = _row_total(r)
        elif result["deferred_inflows"] is None and "deferred inflow" in label:
            # Look for a row that is a total of deferred inflows
            v = _row_total(r)
            if v is None:
                # Check if this row itself has the deferred-inflows line item
                vals = [_num(ws.cell(r, c).value) for c in range(fund_start_col, total_col + 1)]
                v = next((x for x in vals if x is not None), None)
            result["deferred_inflows"] = v

    return result


# ── SONA parser ───────────────────────────────────────────────────────────────

def parse_sona_sheet(ws) -> dict:
    """
    Parse a single parish SONA sheet.
    Governmental Activities is in the column whose header contains 'Governmental'.
    """
    gov_col = None
    for r in range(1, 10):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and "governmental" in v.lower() and "component" not in v.lower():
                gov_col = c
                break
        if gov_col:
            break

    if gov_col is None:
        gov_col = 8  # fallback

    result = {
        "total_assets": None,
        "total_liabilities": None,
        "total_net_assets": None,
    }

    for r in range(1, ws.max_row + 1):
        label = " ".join(
            str(ws.cell(r, c).value or "").strip()
            for c in range(1, 7)
        ).lower().strip()
        label = re.sub(r"\s+", " ", label)

        v = _num(ws.cell(r, gov_col).value)

        if result["total_assets"] is None and "total asset" in label:
            result["total_assets"] = v
        elif result["total_liabilities"] is None and "total liabilit" in label and "net" not in label:
            result["total_liabilities"] = v
        elif result["total_net_assets"] is None and (
            "total net asset" in label or "total net position" in label
        ):
            result["total_net_assets"] = v

    return result


# ── CA parser ─────────────────────────────────────────────────────────────────

def parse_ca_sheet(ws) -> dict:
    """
    Parse a single parish Capital Assets sheet.
    Returns beginning/increases/decreases/ending for total capital assets net.
    """
    # Find column positions from header row
    beg_col = inc_col = dec_col = end_col = None
    for r in range(1, 10):
        for c in range(1, ws.max_column + 1):
            v = str(ws.cell(r, c).value or "").lower().strip()
            if "beginning" in v and beg_col is None:
                beg_col = c
            elif "increase" in v and inc_col is None:
                inc_col = c
            elif "decrease" in v and dec_col is None:
                dec_col = c
            elif "ending" in v and end_col is None:
                end_col = c
        if end_col:
            break

    # Fallbacks from known format
    if beg_col is None: beg_col = 9
    if inc_col is None: inc_col = 10
    if dec_col is None: dec_col = 11
    if end_col is None: end_col = 12

    result = {
        "total_capital_assets_net_beginning": None,
        "total_capital_assets_net_increases": None,
        "total_capital_assets_net_decreases": None,
        "total_capital_assets_net_ending": None,
    }

    for r in range(1, ws.max_row + 1):
        label = " ".join(
            str(ws.cell(r, c).value or "").strip()
            for c in range(1, 7)
        ).lower().strip()
        label = re.sub(r"\s+", " ", label)

        if ("governmental activities capital assets, net" in label or
                ("capital assets" in label and "net" in label and "governmental" in label)):
            result["total_capital_assets_net_beginning"] = _num(ws.cell(r, beg_col).value)
            result["total_capital_assets_net_increases"] = _num(ws.cell(r, inc_col).value)
            result["total_capital_assets_net_decreases"] = _num(ws.cell(r, dec_col).value)
            result["total_capital_assets_net_ending"] = _num(ws.cell(r, end_col).value)
            break

    return result


# ── SOA parser ────────────────────────────────────────────────────────────────

def parse_soa_sheet(ws) -> dict:
    """
    Parse a single parish SOA sheet.
    'Total Governmental Activities' net revenue/expense in the governmental net column.
    """
    # Find the "Total Governmental Activities" column (net expense/revenue for gov)
    gov_net_col = None
    for r in range(1, 10):
        for c in range(1, ws.max_column + 1):
            v = str(ws.cell(r, c).value or "").lower()
            if "governmental" in v and "total" in v and "component" not in v:
                gov_net_col = c
                break
            # Also matches on the net (expense) revenue header
            if "net" in v and "expense" in v and "revenue" in v and "governmental" in v:
                gov_net_col = c
                break
        if gov_net_col:
            break

    if gov_net_col is None:
        gov_net_col = 13  # fallback

    result = {
        "total_governmental_activities_net": None,
        "total_general_revenues": None,
        "change_in_net_assets": None,
        "net_position_ending": None,
        "net_position_beginning": None,
    }

    for r in range(1, ws.max_row + 1):
        label = " ".join(
            str(ws.cell(r, c).value or "").strip()
            for c in range(1, 10)
        ).lower().strip()
        label = re.sub(r"\s+", " ", label)

        v = _num(ws.cell(r, gov_net_col).value)

        if result["total_governmental_activities_net"] is None and (
            "total government activit" in label or "total governmental activit" in label
        ):
            result["total_governmental_activities_net"] = v
        elif result["total_general_revenues"] is None and "total general revenue" in label:
            result["total_general_revenues"] = v
        elif result["change_in_net_assets"] is None and (
            "change in net asset" in label or "change in net position" in label
        ):
            result["change_in_net_assets"] = v
        elif result["net_position_ending"] is None and (
            "net asset" in label or "net position" in label
        ) and ("end" in label or "year" in label):
            result["net_position_ending"] = v
        elif result["net_position_beginning"] is None and (
            "net asset" in label or "net position" in label
        ) and "begin" in label and "restate" not in label:
            result["net_position_beginning"] = v

    return result


# ── comparison ────────────────────────────────────────────────────────────────

def _compare_field(flags, parish, year, stmt_type, field_path,
                   coded_val, pipeline_val, threshold_pct=0.5):
    """
    Compare a coded value against a pipeline-extracted value.
    Adds a flag dict to `flags` list if they differ.
    """
    if coded_val is None and pipeline_val is None:
        return
    if coded_val is None or pipeline_val is None:
        # One is missing
        flags.append({
            "stmt_type": stmt_type,
            "parish": parish,
            "year": year,
            "field_path": field_path,
            "source": "comparison",
            "severity": "medium",
            "extracted_value": pipeline_val,
            "alternative_value": coded_val,
            "reason": (
                f"Pipeline extracted {_fmt(pipeline_val)} but coded Excel has no value"
                if coded_val is None
                else f"Coded Excel has {_fmt(coded_val)} but pipeline extracted nothing"
            ),
            "text_snippet": None,
            "page_number": None,
        })
        return

    try:
        pct = _pct_diff(float(coded_val), float(pipeline_val))
    except Exception:
        # Non-numeric — do string compare
        if str(coded_val).strip() != str(pipeline_val).strip():
            flags.append({
                "stmt_type": stmt_type, "parish": parish, "year": year,
                "field_path": field_path, "source": "comparison", "severity": "medium",
                "extracted_value": str(pipeline_val),
                "alternative_value": str(coded_val),
                "reason": f"Pipeline: '{pipeline_val}' vs coded: '{coded_val}'",
                "text_snippet": None, "page_number": None,
            })
        return

    if pct >= threshold_pct:
        flags.append({
            "stmt_type": stmt_type, "parish": parish, "year": year,
            "field_path": field_path, "source": "comparison",
            "severity": _severity(pct),
            "extracted_value": pipeline_val,
            "alternative_value": coded_val,
            "reason": (
                f"Pipeline extracted {_fmt(pipeline_val)}, "
                f"coded Excel shows {_fmt(coded_val)} "
                f"({pct:.1f}% difference)"
            ),
            "text_snippet": None, "page_number": None,
        })


def _check_identity(flags, parish, year, stmt_type, description, lhs, rhs, tolerance_pct=0.5):
    """
    Flag if |lhs - rhs| / max(|lhs|, |rhs|) > tolerance_pct.
    Source = 'coding_check' (student error in coded Excel).
    """
    if lhs is None or rhs is None:
        return
    try:
        lhs, rhs = float(lhs), float(rhs)
    except Exception:
        return
    denom = max(abs(lhs), abs(rhs))
    pct = 0.0 if denom == 0 else abs(lhs - rhs) / denom * 100.0
    if pct >= tolerance_pct:
        flags.append({
            "stmt_type": stmt_type, "parish": parish, "year": year,
            "field_path": f"identity_check:{description.replace(' ', '_').lower()}",
            "source": "coding_check",
            "severity": _severity(pct),
            "extracted_value": lhs,
            "alternative_value": rhs,
            "reason": (
                f"Coding check failed: {description}. "
                f"LHS={_fmt(lhs)}, RHS={_fmt(rhs)}, diff={pct:.1f}%. "
                f"This may indicate a student coding error."
            ),
            "text_snippet": None, "page_number": None,
        })


# ── master comparison runner ──────────────────────────────────────────────────

def compare_parish(parish: str, year: int, cache_json: dict,
                   cbs_ws=None, sona_ws=None, ca_ws=None, soa_ws=None) -> list[dict]:
    """
    Run all comparisons for one parish.
    Returns list of flag dicts.
    """
    flags = []

    # ── CBS ──────────────────────────────────────────────────────────────────
    if cbs_ws is not None and cache_json.get("cbs"):
        coded = parse_cbs_sheet(cbs_ws)
        pipe = cache_json["cbs"]
        pipe_assets = (pipe.get("assets") or {})
        pipe_liab   = (pipe.get("liabilities") or {})
        pipe_fb     = (pipe.get("fund_balances") or {})

        pipe_total_assets     = ((pipe_assets.get("total_assets") or {}).get("Total Governmental Funds")
                                 or pipe_assets.get("total_assets_total"))
        pipe_total_liab       = ((pipe_liab.get("total_liabilities") or {}).get("Total Governmental Funds")
                                 or pipe_liab.get("total_liabilities_total"))
        # CBS fund balances key: "total_fund_balances" (plural) per pipeline JSON
        pipe_total_fb = None
        for fb_key in ("total_fund_balances", "total_fund_balance"):
            candidate = (pipe_fb.get(fb_key) or {})
            if isinstance(candidate, dict):
                pipe_total_fb = candidate.get("Total Governmental Funds")
            elif candidate is not None:
                pipe_total_fb = candidate
            if pipe_total_fb is not None:
                break
        pipe_total_liab_fb    = (pipe.get("total_liabilities_and_fund_balances") or {}).get(
                                    "Total Governmental Funds")

        _compare_field(flags, parish, year, "cbs", "assets.total_assets[Total Governmental Funds]",
                       coded["total_assets"], pipe_total_assets)
        _compare_field(flags, parish, year, "cbs", "liabilities.total_liabilities[Total Governmental Funds]",
                       coded["total_liabilities"], pipe_total_liab)
        _compare_field(flags, parish, year, "cbs", "fund_balances.total_fund_balance[Total Governmental Funds]",
                       coded["total_fund_balance"], pipe_total_fb)
        _compare_field(flags, parish, year, "cbs", "total_liabilities_and_fund_balances[Total Governmental Funds]",
                       coded["total_liabilities_and_fund_balances"], pipe_total_liab_fb)

        # Coding identity check: Assets = Liabilities + Deferred + Fund Balance
        coded_rhs = None
        if coded["total_liabilities"] is not None and coded["total_fund_balance"] is not None:
            deferred = coded.get("deferred_inflows") or 0.0
            coded_rhs = coded["total_liabilities"] + deferred + coded["total_fund_balance"]
        _check_identity(flags, parish, year, "cbs",
                        "Total Assets = Total Liabilities + Deferred + Fund Balance",
                        coded["total_assets"], coded_rhs)

        # Pipeline identity check via coded
        _check_identity(flags, parish, year, "cbs",
                        "Total Assets = Total Liabilities & Fund Balances (coded)",
                        coded["total_assets"], coded["total_liabilities_and_fund_balances"])

    # ── SONA ─────────────────────────────────────────────────────────────────
    if sona_ws is not None and cache_json.get("sona"):
        coded = parse_sona_sheet(sona_ws)
        pipe  = (cache_json["sona"].get("governmental_activities") or {})

        pipe_total_assets = pipe.get("total_assets")
        pipe_total_liab   = pipe.get("total_liabilities")
        # SONA net position: lives under net_position.total_net_position
        net_pos_block = pipe.get("net_position") or {}
        pipe_total_net = (net_pos_block.get("total_net_position")
                          or net_pos_block.get("total_net_assets")
                          or pipe.get("total_net_position")
                          or pipe.get("total_net_assets"))

        _compare_field(flags, parish, year, "sona",
                       "governmental_activities.total_assets",
                       coded["total_assets"], pipe_total_assets)
        _compare_field(flags, parish, year, "sona",
                       "governmental_activities.total_liabilities",
                       coded["total_liabilities"], pipe_total_liab)
        _compare_field(flags, parish, year, "sona",
                       "governmental_activities.total_net_position",
                       coded["total_net_assets"], pipe_total_net)

        # Coding check: Net Assets = Total Assets - Total Liabilities
        if coded["total_assets"] is not None and coded["total_liabilities"] is not None:
            expected_net = coded["total_assets"] - coded["total_liabilities"]
            _check_identity(flags, parish, year, "sona",
                            "Net Assets = Total Assets - Total Liabilities (coded)",
                            coded["total_net_assets"], expected_net)

    # ── CA ───────────────────────────────────────────────────────────────────
    if ca_ws is not None and cache_json.get("ca"):
        coded = parse_ca_sheet(ca_ws)
        pipe_ga = (cache_json["ca"].get("governmental_activities") or {})
        pipe_net = pipe_ga.get("total_capital_assets_net") or {}

        _compare_field(flags, parish, year, "ca",
                       "governmental_activities.total_capital_assets_net.ending",
                       coded["total_capital_assets_net_ending"],
                       pipe_net.get("ending") if isinstance(pipe_net, dict) else pipe_net)

        _compare_field(flags, parish, year, "ca",
                       "governmental_activities.total_capital_assets_net.beginning",
                       coded["total_capital_assets_net_beginning"],
                       pipe_net.get("beginning") if isinstance(pipe_net, dict) else None)

        # Coding check: Beginning + Increases + Decreases = Ending
        if (coded["total_capital_assets_net_beginning"] is not None and
                coded["total_capital_assets_net_increases"] is not None and
                coded["total_capital_assets_net_ending"] is not None):
            dec = coded["total_capital_assets_net_decreases"] or 0.0
            computed_end = (coded["total_capital_assets_net_beginning"]
                            + coded["total_capital_assets_net_increases"]
                            + dec)
            _check_identity(flags, parish, year, "ca",
                            "CA Net: Beginning + Changes = Ending (coded)",
                            coded["total_capital_assets_net_ending"], computed_end)

    # ── SOA ──────────────────────────────────────────────────────────────────
    if soa_ws is not None and cache_json.get("soa"):
        coded = parse_soa_sheet(soa_ws)
        pipe  = cache_json["soa"]
        pipe_gen_rev = (pipe.get("general_revenues") or {}).get("total_general_revenues")

        _compare_field(flags, parish, year, "soa",
                       "general_revenues.total_general_revenues",
                       coded["total_general_revenues"], pipe_gen_rev)
        _compare_field(flags, parish, year, "soa",
                       "change_in_net_position",
                       coded["change_in_net_assets"], pipe.get("change_in_net_position"))
        _compare_field(flags, parish, year, "soa",
                       "net_position_ending",
                       coded["net_position_ending"], pipe.get("net_position_ending"))

        # Coding check: Change in Net = Beginning + Change = Ending
        if (coded["net_position_beginning"] is not None and
                coded["change_in_net_assets"] is not None and
                coded["net_position_ending"] is not None):
            computed_end = coded["net_position_beginning"] + coded["change_in_net_assets"]
            _check_identity(flags, parish, year, "soa",
                            "SOA: Beginning + Change = Ending Net Position (coded)",
                            coded["net_position_ending"], computed_end)

    return flags


def run_comparison(coded_excel_dir: str, cache_dir: str, db_path: str,
                   year: int, job_id: str = "compare",
                   progress_callback=print) -> dict:
    """
    Main entry point.
    coded_excel_dir : folder containing 'Louisiana <year> CBS.xlsx' etc.
    cache_dir       : folder containing '<Parish>_<year>.json' files
    db_path         : path to feedback.db
    Returns summary dict.
    """
    import openpyxl

    coded_dir = Path(coded_excel_dir)
    cache_path = Path(cache_dir)

    # Load coded Excel workbooks
    def _wb(suffix):
        pattern = f"*{year}*{suffix}*.xlsx"
        matches = list(coded_dir.glob(pattern))
        if not matches:
            # Try case-insensitive approach with all xlsx files
            for f in coded_dir.glob("*.xlsx"):
                if suffix.lower() in f.name.lower() and str(year) in f.name:
                    return openpyxl.load_workbook(str(f), data_only=True)
            return None
        try:
            return openpyxl.load_workbook(str(matches[0]), data_only=True)
        except Exception as e:
            progress_callback(f"[compare] Could not load {matches[0].name}: {e}")
            return None

    wb_cbs  = _wb("CBS")
    wb_sona = _wb("SONA")
    wb_ca   = _wb("Capital Assets")
    wb_soa  = _wb("SOA")

    progress_callback(f"[compare] Loaded coded Excels — CBS:{wb_cbs is not None} "
                      f"SONA:{wb_sona is not None} CA:{wb_ca is not None} SOA:{wb_soa is not None}")

    # Load pipeline cache JSONs
    cache_files = list(cache_path.glob(f"*_{year}.json"))
    progress_callback(f"[compare] Found {len(cache_files)} pipeline cache files for {year}")

    # Save coded data snapshot to DB + run comparisons
    try:
        import sys, os as _os
        _pipeline_dir = str(Path(__file__).resolve().parent)
        if _pipeline_dir not in sys.path:
            sys.path.insert(0, _pipeline_dir)
        from feedback import save_uncertainty, save_coded_data, init_db
        init_db(db_path)
    except Exception as e:
        progress_callback(f"[compare] DB init warning: {e}")
        save_uncertainty = save_coded_data = None

    total_flags = 0
    comparison_flags = 0
    coding_flags = 0
    parishes_covered = []

    for cache_file in sorted(cache_files):
        try:
            with open(cache_file, encoding="utf-8") as f:
                cache_json = json.load(f)
        except Exception as e:
            progress_callback(f"[compare] Could not read {cache_file.name}: {e}")
            continue

        parish = (cache_json.get("cbs") or {}).get("parish") or cache_file.stem.split("_")[0]
        parishes_covered.append(parish)

        progress_callback(f"[compare] Comparing {parish} {year}...")

        # Get sheets for this parish
        def _ws(wb):
            if wb is None:
                return None
            # Try exact match, then partial match
            for name in wb.sheetnames:
                if name.lower() == parish.lower():
                    return wb[name]
            for name in wb.sheetnames:
                if parish.lower() in name.lower() or name.lower() in parish.lower():
                    return wb[name]
            return None

        cbs_ws  = _ws(wb_cbs)
        sona_ws = _ws(wb_sona)
        ca_ws   = _ws(wb_ca)
        soa_ws  = _ws(wb_soa)

        if cbs_ws is None and sona_ws is None and ca_ws is None and soa_ws is None:
            progress_callback(f"[compare]   No coded sheet found for '{parish}' — skipping")
            continue

        flags = compare_parish(parish, year, cache_json,
                               cbs_ws=cbs_ws, sona_ws=sona_ws,
                               ca_ws=ca_ws, soa_ws=soa_ws)

        for flag in flags:
            total_flags += 1
            if flag["source"] == "comparison":
                comparison_flags += 1
            else:
                coding_flags += 1

            if save_uncertainty:
                try:
                    save_uncertainty(
                        db_path, job_id, flag["parish"], flag["year"],
                        flag["stmt_type"],
                        field_path=flag["field_path"],
                        reason=flag["reason"],
                        extracted_value=flag["extracted_value"],
                        alternative_value=flag["alternative_value"],
                        page_number=flag["page_number"],
                        text_snippet=flag["text_snippet"],
                        severity=flag["severity"],
                        source=flag["source"],
                    )
                except Exception as e:
                    progress_callback(f"[compare]   DB save error: {e}")

        progress_callback(f"[compare]   {parish}: {len(flags)} flags "
                          f"({sum(1 for f in flags if f['source']=='comparison')} comparison, "
                          f"{sum(1 for f in flags if f['source']=='coding_check')} coding checks)")

    # Record comparison run metadata
    if save_coded_data:
        try:
            save_coded_data(db_path, year, str(coded_excel_dir),
                            parishes_covered, total_flags,
                            comparison_flags, coding_flags)
        except Exception:
            pass

    summary = {
        "year": year,
        "parishes": parishes_covered,
        "total_flags": total_flags,
        "comparison_flags": comparison_flags,
        "coding_check_flags": coding_flags,
        "job_id": job_id,
    }
    progress_callback(f"[compare] Done. {total_flags} flags across {len(parishes_covered)} parishes.")
    return summary
