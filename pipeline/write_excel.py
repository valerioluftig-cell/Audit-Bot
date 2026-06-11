"""
Excel writers for all 4 statement types.
Matches the layout of the 2013 coded Excel files exactly.
Each file: 64 parish tabs + 1 Cross Sectional tab.
"""
import os
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ── Styling helpers ────────────────────────────────────────────────────────────

def _bold(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=True, name="Arial", size=10)
    return cell

def _val(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Arial", size=10)
    if isinstance(value, (int, float)) and value is not None:
        cell.number_format = '#,##0;(#,##0);"-"'
    return cell

def _header(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=True, name="Arial", size=10)
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    return cell

def _title(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=True, name="Arial", size=12)
    cell.alignment = Alignment(horizontal="center")
    return cell

def _set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def _n(v):
    """Return value or None (skip 0 display as null)."""
    return v if v is not None else None

def _safe(d, *keys):
    """Safe nested dict access."""
    for k in keys:
        if not isinstance(d, dict):
            return None
        d = d.get(k)
    return d


# ── CBS Writer ─────────────────────────────────────────────────────────────────

def write_cbs_parish_tab(ws, data: dict):
    """Write one parish CBS tab matching the 2013 format."""
    parish = data.get("parish", "Unknown")
    year = data.get("year", "")
    funds = data.get("funds", ["Total Governmental Funds"])

    # Ensure Total Governmental Funds is last
    if "Total Governmental Funds" not in funds:
        funds = funds + ["Total Governmental Funds"]

    # Column layout: cols 1-5 are label columns, data starts at col 6
    DATA_START_COL = 6
    n_funds = len(funds)

    # Row 1-2: Title
    _title(ws, 1, 4, f"{parish} Combined Balance Sheet")
    _title(ws, 2, 4, f"Issued: December 31, {year}")

    # Row 4: Fund headers
    for j, fund in enumerate(funds):
        _header(ws, 4, DATA_START_COL + j, fund)

    def write_section_rows(start_row, section_key, items_config):
        """Write a flat section like assets liabilities."""
        row = start_row
        for label, field_key in items_config:
            is_subtotal = label.startswith("Total")
            values = _safe(data, section_key, field_key) or {}
            wrote_any = False
            for v in values.values():
                if v is not None:
                    wrote_any = True
                    break
            if not wrote_any and not is_subtotal:
                continue  # Skip empty rows entirely
            if is_subtotal:
                _bold(ws, row, 5, label + ":")
            else:
                _val(ws, row, 2, label)
            for j, fund in enumerate(funds):
                v = values.get(fund)
                _val(ws, row, DATA_START_COL + j, v)
            row += 1
        return row

    # ── ASSETS ──
    row = 5
    _bold(ws, row, 1, "Assets:")
    row += 1

    asset_rows = [
        ("Cash and interest bearing deposits", "cash_and_deposits"),
        ("Investments", "investments"),
        ("Taxes Receivable", "taxes_receivable"),
        ("Special assessments receivable", "special_assessments_receivable"),
        ("Other receivables", "other_receivables"),
        ("Due from other governmental agencies", "due_from_other_governments"),
        ("Due from other funds", "due_from_other_funds"),
        ("Due from component units", "due_from_component_units"),
        ("Inventory", "inventory"),
        ("Prepaid items", "prepaid_items"),
        ("Other assets", "other_assets"),
    ]
    for label, field in asset_rows:
        values = _safe(data, "assets", field) or {}
        has_data = any(v is not None for v in values.values())
        if not has_data:
            continue
        _val(ws, row, 2, label)
        for j, fund in enumerate(funds):
            _val(ws, row, DATA_START_COL + j, values.get(fund))
        row += 1

    row += 1  # blank
    total_assets = _safe(data, "assets", "total_assets") or {}
    _bold(ws, row, 5, "Total assets:")
    for j, fund in enumerate(funds):
        _val(ws, row, DATA_START_COL + j, total_assets.get(fund))
    row += 2

    # ── LIABILITIES AND FUND BALANCES ──
    _bold(ws, row, 1, "LIABILITIES AND FUND BALANCES")
    row += 1
    _bold(ws, row, 2, "Liabilities:")
    row += 1

    liab_rows = [
        ("Accounts payable", "accounts_payable"),
        ("Retainage payable", "retainage_payable"),
        ("Accrued liabilities and other payables", "accrued_liabilities"),
        ("Deposits payable", "deposits_payable"),
        ("Unearned revenue", "unearned_revenue"),
        ("Due to other funds", "due_to_other_funds"),
        ("Other liabilities", "other_liabilities"),
    ]
    for label, field in liab_rows:
        values = _safe(data, "liabilities", field) or {}
        has_data = any(v is not None for v in values.values())
        if not has_data:
            continue
        _val(ws, row, 3, label)
        for j, fund in enumerate(funds):
            _val(ws, row, DATA_START_COL + j, values.get(fund))
        row += 1

    row += 1
    total_liab = _safe(data, "liabilities", "total_liabilities") or {}
    _bold(ws, row, 5, "Total liabilities:")
    for j, fund in enumerate(funds):
        _val(ws, row, DATA_START_COL + j, total_liab.get(fund))
    row += 2

    # ── DEFERRED INFLOWS ──
    di_items = _safe(data, "deferred_inflows", "items") or []
    di_total = _safe(data, "deferred_inflows", "total") or {}
    if di_items or any(v for v in di_total.values() if v):
        _bold(ws, row, 1, "Deferred Inflows of Resources:")
        row += 1
        for item in di_items:
            values = item.get("values", {})
            if any(v is not None for v in values.values()):
                _val(ws, row, 2, item.get("label", ""))
                for j, fund in enumerate(funds):
                    _val(ws, row, DATA_START_COL + j, values.get(fund))
                row += 1
        row += 1

    # ── FUND BALANCES ──
    _bold(ws, row, 2, "Fund Balances")
    row += 1

    for fb_cat in ["nonspendable", "restricted", "committed", "assigned"]:
        items = _safe(data, "fund_balances", fb_cat) or []
        if not items:
            continue
        if isinstance(items, list):
            for item in items:
                values = item.get("values", {})
                if any(v is not None for v in values.values()):
                    _val(ws, row, 3, fb_cat.capitalize() + ":")
                    _val(ws, row, 4, item.get("label", ""))
                    for j, fund in enumerate(funds):
                        _val(ws, row, DATA_START_COL + j, values.get(fund))
                    row += 1
        elif isinstance(items, dict):
            if any(v is not None for v in items.values()):
                _val(ws, row, 3, fb_cat.capitalize())
                for j, fund in enumerate(funds):
                    _val(ws, row, DATA_START_COL + j, items.get(fund))
                row += 1

    unassigned = _safe(data, "fund_balances", "unassigned") or {}
    if any(v is not None for v in unassigned.values()):
        _val(ws, row, 3, "Unassigned")
        for j, fund in enumerate(funds):
            _val(ws, row, DATA_START_COL + j, unassigned.get(fund))
        row += 1

    row += 1
    total_fb = _safe(data, "fund_balances", "total_fund_balances") or {}
    _bold(ws, row, 5, "Total fund balances:")
    for j, fund in enumerate(funds):
        _val(ws, row, DATA_START_COL + j, total_fb.get(fund))
    row += 2

    total_l_fb = data.get("total_liabilities_and_fund_balances", {}) or {}
    _bold(ws, row, 5, "Total liabilities and fund balances:")
    for j, fund in enumerate(funds):
        _val(ws, row, DATA_START_COL + j, total_l_fb.get(fund))

    # Column widths
    _set_col_width(ws, 1, 28)
    _set_col_width(ws, 2, 32)
    _set_col_width(ws, 3, 32)
    _set_col_width(ws, 4, 30)
    _set_col_width(ws, 5, 28)
    for j in range(n_funds):
        _set_col_width(ws, DATA_START_COL + j, 18)


def write_cbs_cross_sectional_row(ws, data: dict, row: int):
    """Append one parish row to the CBS Cross Sectional sheet."""
    cs = data.get("cross_sectional", {}) or {}
    parish = data.get("parish", "")

    cols = [
        parish,
        cs.get("cash"),
        cs.get("investments"),
        cs.get("receivables"),
        cs.get("inventory"),
        cs.get("other_assets"),
        cs.get("transfers_in"),
        cs.get("prepaid_items"),
        cs.get("total_assets"),          # Sum col
        None,                             # Accuracy placeholder
        cs.get("total_assets"),           # Accuracy value
        cs.get("deferred_outflows"),
        cs.get("accounts_payable"),
        cs.get("deferred_revenues"),
        cs.get("government_transfers"),
        cs.get("other_liabilities"),
        cs.get("total_liabilities"),      # Sum col
        None,
        cs.get("total_liabilities"),
        cs.get("deferred_inflows"),
        cs.get("reserved"),
        cs.get("unreserved_designated"),
        cs.get("unreserved_undesignated"),
        cs.get("total_fund_balances"),    # Sum col
        cs.get("total_liabilities_and_fund_balances"),  # Grand total
        None,
        cs.get("total_liabilities_and_fund_balances"),
    ]
    for j, v in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=j, value=v)
        cell.font = Font(name="Arial", size=10)
        if isinstance(v, (int, float)):
            cell.number_format = '#,##0;(#,##0);"-"'


def init_cbs_cross_sectional(ws, year: int):
    """Write headers for CBS Cross Sectional sheet."""
    ws.title = "Cross Sectional"
    headers_r1 = ["", "Assets:", "", "", "", "", "Non-Current Assets:", "", "", "", "",
                  "", "Liabilities:", "", "", "", "", "", "", "", "Fund Equity:", "", "", "", "", "", ""]
    headers_r2 = ["", "Current Assets:", "", "", "", "", "", "", "", "", "",
                  "", "Current Liabilities:", "", "", "", "", "", "", "", "", "", "", "", "", "", ""]
    headers_r3 = [
        f"{year}-CBS",
        "Cash", "Investments", "Receivables", "Inventory", "Other Assets",
        "Transfers", "Prepaid Items", "Total Assets (Sum)", "Accuracy", "Accuracy",
        "Deferred Outflows of Resources",
        "Accounts Payable", "Deferred Revenues", "Government Transfers", "Other Liabilities",
        "Total Liabilities (Sum)", "Accuracy", "Accuracy",
        "Deferred Inflow of Resources",
        "Reserved", "Unreserved (Designated)", "Unreserved (Undesignated)",
        "Total Fund Balances (Sum)", "Total Liabilities and Fund Balances (Sum)",
        "Accuracy", "Accuracy",
    ]
    headers_r4 = ["Denotes Error Found in Statement"] + [""] * (len(headers_r3) - 1)
    headers_r4[11] = "PARISH"
    headers_r4[19] = "PARISH"
    headers_r5 = ["County"] + [""] * (len(headers_r3) - 1)

    for j, v in enumerate(headers_r1, 1):
        ws.cell(row=1, column=j, value=v).font = Font(bold=True, name="Arial", size=10)
    for j, v in enumerate(headers_r2, 1):
        ws.cell(row=2, column=j, value=v).font = Font(bold=True, name="Arial", size=10)
    for j, v in enumerate(headers_r3, 1):
        c = ws.cell(row=3, column=j, value=v)
        c.font = Font(bold=True, name="Arial", size=10)
        c.alignment = Alignment(wrap_text=True, horizontal="center")
        ws.column_dimensions[get_column_letter(j)].width = 16
    for j, v in enumerate(headers_r4, 1):
        ws.cell(row=4, column=j, value=v).font = Font(bold=True, name="Arial", size=10)
    for j, v in enumerate(headers_r5, 1):
        ws.cell(row=5, column=j, value=v).font = Font(bold=True, name="Arial", size=10)
    ws.column_dimensions["A"].width = 24


# ── SOA Writer ─────────────────────────────────────────────────────────────────

def write_soa_parish_tab(ws, data: dict):
    """Write one parish Statement of Activities tab."""
    parish = data.get("parish", "Unknown")
    year = data.get("year", "")

    _title(ws, 1, 6, f"{parish} {year} Statement of Activities")
    _title(ws, 2, 6, f"Issued: December 31, {year}")

    # Header row
    headers = [
        "Functions/Programs", "", "", "", "",
        "Expenses",
        "Charges for Services",
        "Operating Grants and Contributions",
        "Capital Grants and Contributions",
        "Total Governmental Activities",
        "Component Units",
    ]
    for j, h in enumerate(headers, 1):
        _header(ws, 5, j, h)

    row = 6
    ga = data.get("governmental_activities", {}) or {}
    _bold(ws, row, 1, "Primary Government:")
    row += 1
    _bold(ws, row, 2, "Government Activities:")
    row += 1

    activity_rows = [
        ("General government", "general_government"),
        ("Public safety", "public_safety"),
        ("Public works", "public_works"),
        ("Economic development and assistance", "economic_development"),
        ("Health and welfare", "health_and_welfare"),
        ("Culture and recreation", "culture_and_recreation"),
        ("Interest on long-term debt", "interest_on_long_term_debt"),
    ]
    for label, key in activity_rows:
        act = ga.get(key, {}) or {}
        if not act:
            continue
        _val(ws, row, 3, label)
        _val(ws, row, 6, act.get("expenses"))
        _val(ws, row, 7, act.get("charges_for_services"))
        _val(ws, row, 8, act.get("operating_grants"))
        _val(ws, row, 9, act.get("capital_grants"))
        _val(ws, row, 10, act.get("net_expense_revenue"))
        row += 1

    # Other activities
    for other in (ga.get("other_activities") or []):
        _val(ws, row, 3, other.get("label", ""))
        _val(ws, row, 6, other.get("expenses"))
        _val(ws, row, 7, other.get("charges_for_services"))
        _val(ws, row, 8, other.get("operating_grants"))
        _val(ws, row, 9, other.get("capital_grants"))
        _val(ws, row, 10, other.get("net_expense_revenue"))
        row += 1

    row += 1
    total = ga.get("total", {}) or {}
    _bold(ws, row, 5, "Total Government Activities")
    _val(ws, row, 6, total.get("expenses"))
    _val(ws, row, 7, total.get("charges_for_services"))
    _val(ws, row, 8, total.get("operating_grants"))
    _val(ws, row, 9, total.get("capital_grants"))
    _val(ws, row, 10, total.get("net_expense_revenue"))
    row += 2

    # General Revenues
    _bold(ws, row, 6, "General Revenues:")
    row += 1
    gr = data.get("general_revenues", {}) or {}

    gr_rows = [
        ("Property taxes", "property_taxes"),
        ("Sales and use taxes", "sales_taxes"),
        ("Severance taxes", "severance_taxes"),
        ("Fire insurance premiums", "fire_insurance_premiums"),
        ("Franchise fees", "franchise_fees"),
        ("Other taxes", "other_taxes"),
        ("Occupational licenses and permits", "occupational_licenses"),
        ("Gaming revenues", "gaming_revenues"),
        ("State revenue sharing", "state_revenue_sharing"),
        ("State shared revenue", "state_shared_revenue"),
        ("Non-employer pension contribution", "non_employer_pension_contribution"),
        ("Investment income", "investment_income"),
        ("Miscellaneous", "miscellaneous"),
        ("Transfers", "transfers"),
    ]
    for label, key in gr_rows:
        v = gr.get(key)
        if v is None:
            continue
        _val(ws, row, 7, label)
        _val(ws, row, 10, v)
        row += 1

    for other in (gr.get("other_items") or []):
        _val(ws, row, 7, other.get("label", ""))
        _val(ws, row, 10, other.get("amount"))
        row += 1

    row += 1
    _bold(ws, row, 6, "Total General Revenues")
    _val(ws, row, 10, gr.get("total_general_revenues"))
    row += 2

    _bold(ws, row, 6, "Change in Net Assets")
    _val(ws, row, 10, data.get("change_in_net_position"))
    row += 2

    _bold(ws, row, 6, "Net position, beginning, as previously stated")
    _val(ws, row, 10, data.get("net_position_beginning"))
    row += 1

    _bold(ws, row, 6, "Net position, ending")
    _val(ws, row, 10, data.get("net_position_ending"))

    for col in [1, 2, 3, 4, 5]:
        _set_col_width(ws, col, 5)
    _set_col_width(ws, 6, 30)
    _set_col_width(ws, 7, 22)
    _set_col_width(ws, 8, 22)
    _set_col_width(ws, 9, 22)
    _set_col_width(ws, 10, 22)
    _set_col_width(ws, 11, 16)


def write_soa_cross_sectional_row(ws, data: dict, row: int):
    """Append one parish row to the SOA Governmental Cross Sectional sheet."""
    cs = data.get("cross_sectional", {}) or {}
    parish = data.get("parish", "")

    cols = [
        parish,
        cs.get("property_ad_valorem"),
        cs.get("sales_use_taxes"),
        cs.get("severance_taxes"),
        cs.get("other_tax_revenue"),
        cs.get("total_tax_revenue"),
        cs.get("state_revenue_sharing"),
        cs.get("state_intergovernmental"),
        cs.get("federal_intergovernmental"),
        cs.get("local_transfer"),
        cs.get("all_other_revenue"),
        cs.get("total_other_revenue"),
        cs.get("total_program_revenue"),
        cs.get("total_revenues"),
        None, None,
        cs.get("total_general_government"),
        cs.get("legislative"),
        cs.get("judicial"),
        cs.get("elections"),
        cs.get("finance_and_administration"),
        cs.get("other_general_government"),
        cs.get("total_general_government"),
        cs.get("public_safety"),
        cs.get("public_works"),
        cs.get("economic_development"),
        cs.get("health_and_welfare"),
        cs.get("culture_and_recreation"),
        cs.get("interest_debt_service"),
        cs.get("all_other_expenditures"),
        cs.get("total_expenditures"),
        None, None,
    ]
    for j, v in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=j, value=v)
        cell.font = Font(name="Arial", size=10)
        if isinstance(v, (int, float)):
            cell.number_format = '#,##0;(#,##0);"-"'


def init_soa_cross_sectional(ws, year: int):
    ws.title = "Governmental Cross Sectional"
    r3 = [
        f"{year}-SOA",
        "Property/Ad Valorem", "Sales/Use (includes gaming taxes, hotel-motel etc)",
        "Severance", "Other Tax Revenue", "Total Tax Revenue (SUM)",
        "State Revenue Sharing",
        "State Intergovernmental",
        "Federal Intergovernmental",
        "Local Transfer", "All Other Revenue", "Total Other Revenue (SUM)",
        "Total Program Revenue (SUM)", "Total Revenues (Sum)", "Accuracy", "Accuracy",
        "General Government (not detailed)", "Legislative", "Judicial", "Elections",
        "Finance and Administration", "Other General Government", "Total General Government (SUM)",
        "Public Safety", "Public Works",
        "Economic Development and Assistance", "Health and Welfare", "Culture and Recreation",
        "Interest and fiscal charges on long-term debt (debt service)",
        "All Other Expenditures", "Total Expenditures (Sum)", "Accuracy", "Accuracy",
    ]
    for j, v in enumerate(r3, 1):
        c = ws.cell(row=3, column=j, value=v)
        c.font = Font(bold=True, name="Arial", size=10)
        c.alignment = Alignment(wrap_text=True, horizontal="center")
        ws.column_dimensions[get_column_letter(j)].width = 16
    ws.cell(row=1, column=1, value=f"{year}-SOA").font = Font(bold=True, name="Arial", size=10)
    ws.cell(row=4, column=1, value="Error Found in Statement").font = Font(bold=True, name="Arial", size=10)
    ws.cell(row=5, column=1, value="").font = Font(bold=True, name="Arial", size=10)
    ws.cell(row=5, column=2, value="PARISH").font = Font(bold=True, name="Arial", size=10)
    ws.column_dimensions["A"].width = 28


# ── SONA Writer ────────────────────────────────────────────────────────────────

def write_sona_parish_tab(ws, data: dict):
    """Write one parish Statement of Net Position tab."""
    parish = data.get("parish", "Unknown")
    year = data.get("year", "")

    _title(ws, 1, 4, f"{parish} Parish {year} Statement of Net Assets")
    _title(ws, 2, 4, f"Issued: December 31, {year}")

    _header(ws, 4, 8, "Primary Government")
    _header(ws, 5, 8, "Governmental Activities")
    _header(ws, 5, 9, "Component Unit")

    ga = data.get("governmental_activities", {}) or {}
    cu = data.get("component_units", {}) or {}

    row = 6
    _bold(ws, row, 1, "ASSETS")
    row += 1
    _bold(ws, row, 2, "Current Assets")
    row += 1

    current_asset_rows = [
        ("Cash", "cash_and_deposits"),
        ("Investments", "investments"),
        ("Taxes receivable", "taxes_receivable"),
        ("Other receivables", "other_receivables"),
        ("Due from other governmental entities", "due_from_other_governments"),
        ("Due from component units", "due_from_component_units"),
        ("Inventory", "inventory"),
        ("Prepaid items", "prepaid_items"),
        ("Other current assets", "other_current_assets"),
    ]
    ca_current = ga.get("current_assets", {}) or {}
    for label, key in current_asset_rows:
        v = ca_current.get(key)
        if v is None:
            continue
        _val(ws, row, 3, label)
        _val(ws, row, 8, v)
        row += 1

    # Capital assets
    ca_data = ga.get("capital_assets", {}) or {}
    if ca_data:
        _val(ws, row, 3, "Capital assets:")
        row += 1
        for label, key in [("Non-depreciable", "non_depreciable"), ("Depreciable, net", "depreciable_net"),
                            ("Right-to-use lease assets, net", "right_to_use_net")]:
            v = ca_data.get(key)
            if v is None:
                continue
            _val(ws, row, 4, label)
            _val(ws, row, 8, v)
            row += 1

    row += 1
    _bold(ws, row, 3, "Total Assets")
    _val(ws, row, 8, ga.get("total_assets"))
    _val(ws, row, 9, cu.get("total_assets"))
    row += 2

    # Deferred Outflows
    do_data = ga.get("deferred_outflows", {}) or {}
    if do_data.get("total"):
        _bold(ws, row, 1, "DEFERRED OUTFLOWS OF RESOURCES")
        row += 1
        for item in (do_data.get("items") or []):
            _val(ws, row, 3, item.get("label", ""))
            _val(ws, row, 8, item.get("amount"))
            row += 1
        row += 1

    # Liabilities
    _bold(ws, row, 1, "LIABILITIES")
    row += 1

    curr_liab = ga.get("current_liabilities", {}) or {}
    liab_rows = [
        ("Accounts payable", "accounts_payable"),
        ("Retainage payable", "retainage_payable"),
        ("Accrued liabilities and other payables", "accrued_liabilities"),
        ("Deposits payable", "deposits_payable"),
        ("Unearned revenue", "unearned_revenue"),
        ("Accrued interest payable", "accrued_interest"),
        ("Other current liabilities", "other_current_liabilities"),
    ]
    for label, key in liab_rows:
        v = curr_liab.get(key)
        if v is None:
            continue
        _val(ws, row, 3, label)
        _val(ws, row, 8, v)
        row += 1

    # Long-term liabilities
    lt_liab = ga.get("long_term_liabilities", {}) or {}
    if lt_liab:
        _val(ws, row, 3, "Long-term liabilities:")
        row += 1
        lt_rows = [
            ("Bonds payable - current portion", "bonds_payable_current"),
            ("Bonds payable - long-term", "bonds_payable_noncurrent"),
            ("Compensated absences - current", "compensated_absences_current"),
            ("Compensated absences - long-term", "compensated_absences_noncurrent"),
            ("Net pension liability", "net_pension_liability"),
            ("Landfill closure and post-closure care costs", "landfill_closure"),
            ("Lease liability - current", "lease_liability_current"),
            ("Lease liability - long-term", "lease_liability_noncurrent"),
            ("Other long-term liabilities", "other_long_term"),
        ]
        for label, key in lt_rows:
            v = lt_liab.get(key)
            if v is None:
                continue
            _val(ws, row, 4, label)
            _val(ws, row, 8, v)
            row += 1

    row += 1
    _bold(ws, row, 4, "Total Liabilities")
    _val(ws, row, 8, ga.get("total_liabilities"))
    _val(ws, row, 9, cu.get("total_liabilities"))
    row += 2

    # Deferred Inflows
    di_data = ga.get("deferred_inflows", {}) or {}
    if di_data.get("total"):
        _bold(ws, row, 1, "DEFERRED INFLOWS OF RESOURCES")
        row += 1
        for item in (di_data.get("items") or []):
            _val(ws, row, 3, item.get("label", ""))
            _val(ws, row, 8, item.get("amount"))
            row += 1
        row += 1

    # Net Position
    np_data = ga.get("net_position", {}) or {}
    _bold(ws, row, 1, "NET POSITION")
    row += 1
    for label, key in [
        ("Net investment in capital assets", "net_investment_in_capital_assets"),
        ("Restricted", "restricted"),
        ("Unrestricted", "unrestricted"),
    ]:
        v = np_data.get(key)
        if v is None:
            continue
        _val(ws, row, 3, label)
        _val(ws, row, 8, v)
        row += 1
    row += 1
    _bold(ws, row, 4, "Total Net Position")
    _val(ws, row, 8, np_data.get("total_net_position"))
    _val(ws, row, 9, cu.get("total_net_position"))

    for col in range(1, 8):
        _set_col_width(ws, col, 6)
    _set_col_width(ws, 8, 22)
    _set_col_width(ws, 9, 18)


# ── Capital Assets Writer ──────────────────────────────────────────────────────

def write_ca_parish_tab(ws, data: dict):
    """Write one parish Capital Assets tab."""
    parish = data.get("parish", "Unknown")
    year = data.get("year", "")

    _title(ws, 1, 4, f"{parish.upper()} PARISH POLICE JURY")
    _title(ws, 2, 4, "Capital Assets")
    _title(ws, 3, 4, f"For the Year Ended December 31, {year}")

    for j, h in enumerate(["Beginning Balance", "Increases", "Decreases", "Ending Balance", "Cross Sectional"], start=9):
        _header(ws, 5, j, h)

    ga = data.get("governmental_activities", {}) or {}
    row = 7

    def write_ca_section(title, section_key, items_config):
        nonlocal row
        _bold(ws, row, 1, title)
        row += 1
        section = ga.get(section_key, {}) or {}

        for label, field in items_config:
            d = section.get(field)
            if isinstance(d, dict) and any(v is not None for v in d.values()):
                _val(ws, row, 2, label)
                for col_off, key in enumerate(["beginning", "increases", "decreases", "ending"], start=9):
                    _val(ws, row, col_off, d.get(key))
                cs_val = d.get("ending")
                _val(ws, row, 13, cs_val)
                row += 1

        # Other items (dynamic)
        other_key = "other_non_depreciable" if section_key == "not_depreciated" else "other_depreciable"
        for item in (section.get(other_key) or []):
            d_item = item
            if isinstance(d_item, dict):
                _val(ws, row, 2, d_item.get("label", ""))
                for col_off, key in enumerate(["beginning", "increases", "decreases", "ending"], start=9):
                    _val(ws, row, col_off, d_item.get(key))
                row += 1

    write_ca_section(
        "Capital assets, not being depreciated:",
        "not_depreciated",
        [("Land", "land"), ("Construction in progress", "construction_in_progress")],
    )

    # Total not depreciated
    tnd = _safe(ga, "not_depreciated", "total_not_depreciated") or {}
    if tnd:
        _val(ws, row, 3, "Total assets not being depreciated")
        for col_off, key in enumerate(["beginning", "increases", "decreases", "ending"], start=9):
            _val(ws, row, col_off, tnd.get(key))
        row += 2

    write_ca_section(
        "Capital assets being depreciated:",
        "depreciable",
        [
            ("Buildings and improvements", "buildings_and_improvements"),
            ("Machinery and equipment", "machinery_and_equipment"),
            ("Improvements, other than buildings", "improvements_other_than_buildings"),
            ("Infrastructure", "infrastructure"),
            ("Vehicles", "vehicles"),
            ("Furniture and fixtures", "furniture_and_fixtures"),
            ("Books and periodicals", "books_and_periodicals"),
            ("Leased property under capital leases", "leased_property"),
        ],
    )
    tdep = _safe(ga, "depreciable", "total_depreciable") or {}
    if tdep:
        _val(ws, row, 3, "Total capital assets being depreciated")
        for col_off, key in enumerate(["beginning", "increases", "decreases", "ending"], start=9):
            _val(ws, row, col_off, tdep.get(key))
        row += 2

    # Accumulated depreciation
    _bold(ws, row, 1, "Less accumulated depreciation for:")
    row += 1
    acc_dep = ga.get("accumulated_depreciation", {}) or {}
    acc_rows = [
        ("Buildings and improvements", "buildings_and_improvements"),
        ("Machinery and equipment", "machinery_and_equipment"),
        ("Improvements, other than buildings", "improvements_other_than_buildings"),
        ("Infrastructure", "infrastructure"),
        ("Vehicles", "vehicles"),
        ("Furniture and fixtures", "furniture_and_fixtures"),
        ("Books and periodicals", "books_and_periodicals"),
        ("Leased property under capital leases", "leased_property"),
    ]
    for label, key in acc_rows:
        d = acc_dep.get(key)
        if isinstance(d, dict) and any(v is not None for v in d.values()):
            _val(ws, row, 2, label)
            for col_off, fk in enumerate(["beginning", "increases", "decreases", "ending"], start=9):
                _val(ws, row, col_off, d.get(fk))
            ending = d.get("ending")
            # Net = ending gross depreciable - this depreciation → handled in cross sectional
            _val(ws, row, 13, ending)
            row += 1

    tacc = acc_dep.get("total_accumulated_depreciation") or {}
    if tacc:
        _val(ws, row, 3, "Total accumulated depreciation")
        for col_off, key in enumerate(["beginning", "increases", "decreases", "ending"], start=9):
            _val(ws, row, col_off, tacc.get(key))
        row += 2

    tnet = ga.get("total_capital_assets_net") or {}
    if tnet:
        _bold(ws, row, 2, "Governmental activities capital assets, net")
        for col_off, key in enumerate(["beginning", "increases", "decreases", "ending"], start=9):
            _val(ws, row, col_off, tnet.get(key))
        _val(ws, row, 13, tnet.get("ending"))

    for col in range(1, 9):
        _set_col_width(ws, col, 6 if col < 4 else 26)
    for col in range(9, 14):
        _set_col_width(ws, col, 18)


def write_ca_cross_sectional_row(ws, data: dict, row: int):
    """Append one parish row to Capital Assets Cross Sectional."""
    cs = data.get("cross_sectional", {}) or {}
    parish = data.get("parish", "")

    cols = [
        parish,
        cs.get("land"),
        cs.get("construction_in_progress"),
        cs.get("other_non_depreciable"),
        cs.get("buildings_net"),
        cs.get("improvements_net"),
        cs.get("machinery_net"),
        cs.get("other_depreciable_net"),
        cs.get("books_net"),
        cs.get("furniture_net"),
        cs.get("vehicles_net"),
        cs.get("bridges_net"),
        cs.get("leased_property_net"),
        cs.get("infrastructure_net"),
        cs.get("total_governmental_net"),
        cs.get("total_governmental_net"),  # Check
        None,                               # Okay?
    ]
    for j, v in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=j, value=v)
        cell.font = Font(name="Arial", size=10)
        if isinstance(v, (int, float)):
            cell.number_format = '#,##0;(#,##0);"-"'


def init_ca_cross_sectional(ws, year: int):
    ws.title = "Cross Sectional"
    headers = [
        "Parish",
        "Land", "Construction in progress", "Other non-depreciable",
        "Buildings and improvements", "Improvements other than buildings",
        "Machinery and equipment", "Other depreciable",
        "Books and periodicals", "Furniture and fixtures", "Vehicles",
        "Bridges", "Leased Property Under Capital Leases", "Infrastructure",
        "Total (Net)", "Check", "Okay?",
    ]
    ws.cell(row=1, column=1, value="Louisiana Counties").font = Font(bold=True, name="Arial", size=11)
    ws.cell(row=2, column=1, value=f"{year} Capital Assets, Net of Accumulated Depreciation").font = Font(bold=True, name="Arial", size=10)
    ws.cell(row=3, column=1, value="Governmental Activities").font = Font(bold=True, name="Arial", size=10)
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=j, value=h)
        c.font = Font(bold=True, name="Arial", size=10)
        c.alignment = Alignment(wrap_text=True, horizontal="center")
        ws.column_dimensions[get_column_letter(j)].width = 16
    ws.column_dimensions["A"].width = 22


# ── Workbook management ────────────────────────────────────────────────────────

STMT_CONFIGS = {
    "cbs": {
        "filename": "Louisiana {year} CBS.xlsx",
        "tab_writer": write_cbs_parish_tab,
        "cs_writer": write_cbs_cross_sectional_row,
        "cs_init": init_cbs_cross_sectional,
    },
    "soa": {
        "filename": "Louisiana {year} SOA.xlsx",
        "tab_writer": write_soa_parish_tab,
        "cs_writer": write_soa_cross_sectional_row,
        "cs_init": init_soa_cross_sectional,
    },
    "sona": {
        "filename": "Louisiana {year} SONA.xlsx",
        "tab_writer": write_sona_parish_tab,
        "cs_writer": None,  # SONA has no cross sectional in the 2013 data
        "cs_init": None,
    },
    "ca": {
        "filename": "Louisiana {year} Capital Assets.xlsx",
        "tab_writer": write_ca_parish_tab,
        "cs_writer": write_ca_cross_sectional_row,
        "cs_init": init_ca_cross_sectional,
    },
}


def get_or_create_workbook(output_dir: str, stmt_type: str, year: int) -> tuple[Workbook, str]:
    """Load existing workbook or create a new one."""
    cfg = STMT_CONFIGS[stmt_type]
    filename = cfg["filename"].format(year=year)
    path = os.path.join(output_dir, filename)

    if os.path.exists(path):
        wb = load_workbook(path)
    else:
        wb = Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        # For SONA there's no cross sectional — create a placeholder sheet so wb is valid
        if cfg["cs_init"]:
            cs_ws = wb.create_sheet("Cross Sectional")
            cfg["cs_init"](cs_ws, year)
        else:
            wb.create_sheet("Overview")  # placeholder so workbook is never empty

    return wb, path


def add_parish_to_workbook(
    wb: Workbook,
    stmt_type: str,
    data: dict,
    year: int,
):
    """Add/update parish tab and cross sectional row in the workbook."""
    cfg = STMT_CONFIGS[stmt_type]
    parish = data.get("parish", "Unknown")

    # Create or overwrite parish tab
    if parish in wb.sheetnames:
        del wb[parish]
    ws = wb.create_sheet(parish)

    cfg["tab_writer"](ws, data)

    # Add to Cross Sectional
    if cfg["cs_writer"]:
        cs_sheet_name = "Cross Sectional" if stmt_type != "soa" else "Governmental Cross Sectional"
        if cs_sheet_name not in wb.sheetnames:
            cs_ws = wb.create_sheet(cs_sheet_name)
            cfg["cs_init"](cs_ws, year)
        else:
            cs_ws = wb[cs_sheet_name]

        # Find next empty row (after header rows 1-5)
        next_row = 6
        while cs_ws.cell(row=next_row, column=1).value:
            next_row += 1

        cfg["cs_writer"](cs_ws, data, next_row)


def save_workbook(wb: Workbook, path: str):
    wb.save(path)
    print(f"    Saved: {os.path.basename(path)}")


def write_parish_combined_workbook(
    output_dir: str,
    parish: str,
    results: dict,
    year: int,
    parish_validation=None,   # validate.ParishValidation or None
    manual_pages: dict = None,
) -> str:
    """
    Create a single workbook with up to 4 sheets — one per statement type —
    for a single parish. Sheet order: CBS, SOA, SONA, Capital Assets, Quality Report.
    Sheets are omitted silently when the statement was not found.
    Returns the filename (not the full path).
    """
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    sheet_config = [
        ("cbs",  "CBS"),
        ("soa",  "SOA"),
        ("sona", "SONA"),
        ("ca",   "Capital Assets"),
    ]

    wrote_any = False
    for stmt_type, sheet_name in sheet_config:
        data = results.get(stmt_type)
        if not data:
            continue
        ws = wb.create_sheet(sheet_name)
        STMT_CONFIGS[stmt_type]["tab_writer"](ws, data)
        wrote_any = True

    if not wrote_any:
        return None  # nothing to save

    # Add per-parish quality report tab if validation data available
    if parish_validation is not None:
        try:
            from validate import write_quality_report_excel_tab
            write_quality_report_excel_tab(
                wb,
                [parish_validation],
                manual_pages or {},
                year,
            )
        except Exception:
            pass  # never let quality report crash the parish workbook

    filename = f"{parish} {year}.xlsx"
    path = os.path.join(output_dir, filename)
    wb.save(path)
    return filename
