import json, sys, traceback, pathlib
sys.path.insert(0, r'C:\Users\valer\Downloads\audit_project\pipeline')
from coded_compare import compare_parish
import openpyxl

coded_dir = pathlib.Path(r'C:\Users\valer\Downloads\audit_project\coded\2013')
wb_cbs  = openpyxl.load_workbook(str(coded_dir / 'Louisiana 2013 CBS.xlsx'),  data_only=True)
wb_sona = openpyxl.load_workbook(str(coded_dir / 'Louisiana 2013 SONA.xlsx'), data_only=True)
wb_ca   = openpyxl.load_workbook(str(coded_dir / 'Louisiana 2013 Capital Assets.xlsx'), data_only=True)
wb_soa  = openpyxl.load_workbook(str(coded_dir / 'Louisiana 2013 SOA.xlsx'),  data_only=True)

cache_dir = pathlib.Path(r'C:\Users\valer\Downloads\audit_project\runs\e44ac3c9\output\cache')
files = sorted(cache_dir.glob('*_2013.json'))

errors = []
for jf in files:
    parish = jf.stem.rsplit('_', 1)[0]
    cache = json.loads(jf.read_text())
    try:
        flags = compare_parish(parish, 2013, cache,
            cbs_ws=wb_cbs[parish] if parish in wb_cbs.sheetnames else None,
            sona_ws=wb_sona[parish] if parish in wb_sona.sheetnames else None,
            ca_ws=wb_ca[parish] if parish in wb_ca.sheetnames else None,
            soa_ws=wb_soa[parish] if parish in wb_soa.sheetnames else None)
        print(f'OK  {parish}: {len(flags)} flags')
    except Exception as e:
        tb = traceback.format_exc()
        errors.append((parish, str(e), tb))
        print(f'ERR {parish}: {e}')

print()
print(f'{len(errors)} parishes crashed:')
for parish, msg, tb in errors:
    print(f'\n=== {parish} ===')
    print(tb)
