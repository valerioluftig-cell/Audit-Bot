"""
Check what parishes are in the coded Excels vs what's in cache,
to understand the comparison coverage gap.
"""
import pathlib, sys
sys.path.insert(0, r"C:\Users\valer\Downloads\audit_project\pipeline")

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

coded_dir = pathlib.Path(r"C:\Users\valer\Downloads\audit_project\coded\2013")

for xlfile in sorted(coded_dir.glob("Louisiana 2013*.xlsx")):
    print(f"\n{xlfile.name}")
    if not HAS_OPENPYXL:
        print("  (openpyxl not available)")
        continue
    wb = openpyxl.load_workbook(xlfile, read_only=True, data_only=True)
    sheets = [s for s in wb.sheetnames if s not in ("Cross Sectional","Instructions","Notes")]
    print(f"  Sheets ({len(sheets)}): {sheets[:5]}{'...' if len(sheets)>5 else ''}")
    wb.close()

print()
# Also check 2014
coded_dir_14 = pathlib.Path(r"C:\Users\valer\Downloads\audit_project\coded\2014")
for xlfile in sorted(coded_dir_14.glob("Louisiana 2014*.xlsx")):
    print(f"\n{xlfile.name}")
    if not HAS_OPENPYXL:
        print("  (openpyxl not available)")
        continue
    wb = openpyxl.load_workbook(xlfile, read_only=True, data_only=True)
    sheets = [s for s in wb.sheetnames if s not in ("Cross Sectional","Instructions","Notes")]
    print(f"  Sheets ({len(sheets)}): {sheets[:5]}{'...' if len(sheets)>5 else ''}")
    wb.close()
