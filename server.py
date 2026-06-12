#!/usr/bin/env python3
"""
Parish Audit Pipeline — single file.

1. Run:  python server.py
2. Browser opens automatically.
3. Paste your Anthropic API key in the Settings panel (gear icon, top-right).
"""
import io, json, os, queue, re, shutil, sqlite3, sys, threading, time, uuid, webbrowser
from pathlib import Path


def _seed_data():
    """On first run, copy seeded/ data into working dirs so Training is pre-populated."""
    seed = _resource_dir() / "seeded"
    if not seed.exists():
        return
    app = _app_dir()

    # feedback.db — copy only if none exists yet
    seed_db = seed / "feedback.db"
    target_db = app / "feedback.db"
    if seed_db.exists() and not target_db.exists():
        try:
            shutil.copy2(seed_db, target_db)
            print("[seed] Copied feedback.db", flush=True)
        except Exception as e:
            print(f"[seed] feedback.db copy failed: {e}", flush=True)

    # coded/ — copy year subdirs that don't exist yet
    seed_coded = seed / "coded"
    if seed_coded.exists():
        for yr_dir in seed_coded.iterdir():
            target = app / "coded" / yr_dir.name
            if not target.exists():
                try:
                    shutil.copytree(yr_dir, target)
                    print(f"[seed] Copied coded/{yr_dir.name}", flush=True)
                except Exception as e:
                    print(f"[seed] coded/{yr_dir.name} failed: {e}", flush=True)

    # runs/ — copy seeded run dirs that don't exist yet
    seed_runs = seed / "runs"
    if seed_runs.exists():
        target_runs = app / "runs"
        target_runs.mkdir(exist_ok=True)
        for run_dir in seed_runs.iterdir():
            target = target_runs / f"seeded_{run_dir.name}"
            if not target.exists():
                try:
                    shutil.copytree(run_dir, target)
                    print(f"[seed] Copied runs/{run_dir.name} → seeded_{run_dir.name}", flush=True)
                except Exception as e:
                    print(f"[seed] runs/{run_dir.name} failed: {e}", flush=True)

def _app_dir() -> Path:
    """
    Returns the writable app directory — works both as a plain .py script
    and when frozen by PyInstaller.
    - Frozen:  directory containing AuditPipeline.exe
    - Script:  directory containing server.py
    """
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent

def _resource_dir() -> Path:
    """
    Returns the directory where bundled read-only assets live.
    - Frozen:  sys._MEIPASS  (PyInstaller extraction folder)
    - Script:  same as _app_dir()
    """
    if getattr(sys, "frozen", False):
        return Path(sys._MEIPASS)
    return Path(__file__).resolve().parent

# ── API key: loaded from config.json next to the exe / server.py ─────────────
def _config_path() -> Path:
    return _app_dir() / "config.json"

def _load_api_key() -> str:
    # Azure / Railway: key set as environment variable
    if os.environ.get("ANTHROPIC_API_KEY"):
        return os.environ["ANTHROPIC_API_KEY"]
    try:
        cfg = json.loads(_config_path().read_text())
        return cfg.get("api_key", "")
    except Exception:
        return ""

def _save_api_key(key: str):
    cfg = {}
    try:
        cfg = json.loads(_config_path().read_text())
    except Exception:
        pass
    cfg["api_key"] = key.strip()
    _config_path().write_text(json.dumps(cfg, indent=2))

def _get_api_key() -> str:
    """Always returns the freshest key (re-reads config each call)."""
    return _load_api_key()

# Feedback database — initialised on first run, path set at startup
_DB_PATH: str | None = None

def _get_db_path() -> str | None:
    return _DB_PATH

def _search_candidates(value_str: str) -> list[str]:
    """
    Build a list of number strings to search for on the PDF page.
    Tries full value, in-thousands form, and parentheses-negative form.
    """
    candidates = []
    try:
        raw = str(value_str).strip().replace(",", "").replace("$", "").replace(" ", "")
        negative = raw.startswith("(") and raw.endswith(")")
        if negative:
            raw = "-" + raw[1:-1]
        v = float(raw)
        a = abs(v)
        # Full integer forms
        candidates += [f"{a:,.0f}", f"{int(a)}"]
        # In-thousands forms (PDF shows 1,235 but stored value is 1,234,567)
        if a >= 1000:
            t = a / 1000
            candidates += [f"{t:,.0f}", f"{t:.0f}",
                           f"{t:,.1f}", f"{t:,.2f}"]
        # Negative / parentheses forms
        if v < 0 or negative:
            candidates += [f"({a:,.0f})", f"({int(a)})"]
            if a >= 1000:
                t = a / 1000
                candidates += [f"({t:,.0f})", f"({t:.0f})"]
    except Exception:
        candidates.append(str(value_str))
    # De-duplicate preserving order
    seen = set()
    return [c for c in candidates if not (c in seen or seen.add(c))]


def find_text_boxes(pdf_path: str, page_0indexed: int,
                    value: str | None = None,
                    snippet: str | None = None) -> list[dict]:
    """
    Return pdfplumber bounding boxes for the target value / snippet on the page.
    Tries word-level matching first, then falls back to joined-token search
    to handle numbers like "(1,234,567)" that PDF engines split across words.
    """
    import pdfplumber as _plb
    boxes = []
    try:
        with _plb.open(pdf_path) as pdf:
            if page_0indexed >= len(pdf.pages):
                return boxes
            pg = pdf.pages[page_0indexed]
            words = pg.extract_words(keep_blank_chars=True,
                                     extra_attrs=["fontname", "size"])

            targets = _search_candidates(value) if value else []
            if snippet:
                for tok in snippet.split():
                    clean = tok.strip("$(),.")
                    if len(clean) >= 3:
                        targets.append(clean)

            # Pass 1 – single-word matching
            matched_indices = set()
            for idx, word in enumerate(words):
                text = word["text"].strip()
                for tgt in targets:
                    if not tgt:
                        continue
                    # Normalise: strip surrounding parens/dollar/comma for comparison
                    norm_text = text.lstrip("$(").rstrip("),").replace(",", "")
                    norm_tgt  = tgt.lstrip("$(").rstrip("),").replace(",", "")
                    if (text == tgt
                            or text.lstrip("$").strip() == tgt
                            or tgt in text
                            or norm_text == norm_tgt):
                        boxes.append({
                            "x0": word["x0"], "top": word["top"],
                            "x1": word["x1"], "bottom": word["bottom"],
                        })
                        matched_indices.add(idx)
                        break

            # Pass 2 – sliding window of 2–3 adjacent words (handles split tokens)
            if not boxes and words:
                for win in (2, 3):
                    for i in range(len(words) - win + 1):
                        joined = "".join(w["text"] for w in words[i:i+win]).strip()
                        for tgt in targets:
                            if tgt and (tgt == joined or tgt in joined):
                                grp = words[i:i+win]
                                boxes.append({
                                    "x0": grp[0]["x0"], "top": min(w["top"] for w in grp),
                                    "x1": grp[-1]["x1"], "bottom": max(w["bottom"] for w in grp),
                                })
                                break
                    if boxes:
                        break

    except Exception as e:
        print(f"[find_text_boxes] {e}", flush=True)
    return boxes


def find_row_col_labels(pdf_path: str, page_0: int, value_box: dict) -> dict:
    """
    Given a matched value's bounding box, locate:
      row_boxes  – words to the LEFT on the same horizontal row (the line-item label)
      col_boxes  – words ABOVE in the same horizontal column band (the column header)
    Returns {"row_boxes": [...], "col_boxes": [...]}
    """
    import pdfplumber as _plb
    result: dict = {"row_boxes": [], "col_boxes": []}
    try:
        with _plb.open(pdf_path) as pdf:
            if page_0 >= len(pdf.pages):
                return result
            pg = pdf.pages[page_0]
            words = pg.extract_words(keep_blank_chars=True)

            v_cy   = (value_box["top"]  + value_box["bottom"]) / 2
            v_cx   = (value_box["x0"]   + value_box["x1"])     / 2
            v_h    = value_box["bottom"] - value_box["top"]
            v_w    = value_box["x1"]    - value_box["x0"]

            row_tol = max(9, v_h * 0.65)
            col_tol = max(18, v_w * 0.9)

            # Row label: words to the LEFT of value on the same horizontal band
            row_words = [
                w for w in words
                if w["x1"] < value_box["x0"] - 2
                and abs((w["top"] + w["bottom"]) / 2 - v_cy) < row_tol
            ]
            if row_words:
                row_words.sort(key=lambda w: w["x0"])
                result["row_boxes"] = [
                    {"x0": w["x0"], "top": w["top"], "x1": w["x1"], "bottom": w["bottom"]}
                    for w in row_words
                ]

            # Column header: words ABOVE the value in the same horizontal band
            col_words = [
                w for w in words
                if w["bottom"] < value_box["top"] - 2
                and abs((w["x0"] + w["x1"]) / 2 - v_cx) < col_tol
            ]
            if col_words:
                # Take the group of words closest to the value (largest bottom value)
                col_words.sort(key=lambda w: -w["bottom"])
                closest_bottom = col_words[0]["bottom"]
                col_group = [w for w in col_words if abs(w["bottom"] - closest_bottom) < 7]
                result["col_boxes"] = [
                    {"x0": w["x0"], "top": w["top"], "x1": w["x1"], "bottom": w["bottom"]}
                    for w in col_group
                ]
    except Exception as e:
        print(f"[find_row_col_labels] {e}", flush=True)
    return result


def render_pdf_page(pdf_path: str, page_0indexed: int,
                    highlight_boxes: list[dict] | None = None,
                    trace_row_boxes: list[dict] | None = None,
                    trace_col_boxes: list[dict] | None = None,
                    second_page_0indexed: int | None = None,
                    dpi: int = 150) -> bytes | None:
    """
    Render a PDF page (optionally stitched with a second page below) to PNG.
    Overlay layers (all optional):
      highlight_boxes  – yellow/orange fill for the matched value
      trace_row_boxes  – blue fill for the row label words (left of value)
      trace_col_boxes  – green fill for the column header words (above value)
    When both row/col traces AND highlight boxes are present, dashed connector
    lines are drawn: horizontal (row→value) and vertical (col→value).
    """
    try:
        import pypdfium2 as pdfium
        from PIL import Image, ImageDraw

        doc   = pdfium.PdfDocument(pdf_path)
        scale = dpi / 72

        def _one(p_idx) -> Image.Image:
            pg  = doc[p_idx]
            bmp = pg.render(scale=scale, rotation=0)
            return bmp.to_pil().convert("RGBA")

        img1 = _one(page_0indexed)
        page1_h = img1.height

        # --- stitch second page below if requested ---
        if second_page_0indexed is not None and second_page_0indexed < len(doc):
            img2 = _one(second_page_0indexed)
            sep_h = 6  # separator height
            full_w = max(img1.width, img2.width)
            full_h = img1.height + sep_h + img2.height
            canvas = Image.new("RGBA", (full_w, full_h), (30, 32, 40, 255))
            canvas.paste(img1, (0, 0))
            # thin separator
            sep_draw = ImageDraw.Draw(canvas)
            sep_draw.rectangle([0, img1.height, full_w, img1.height + sep_h],
                               fill=(80, 80, 90, 255))
            canvas.paste(img2, (0, img1.height + sep_h))
            pil_img = canvas
        else:
            pil_img = img1

        # --- draw all annotation layers ---
        overlay = Image.new("RGBA", pil_img.size, (0, 0, 0, 0))
        draw    = ImageDraw.Draw(overlay)
        pad     = 4

        def px(box, y_off=0):
            """Convert PDF coords → pixel rect, with optional vertical offset."""
            return (
                max(0,             int(box["x0"]     * scale) - pad),
                max(0,             int(box["top"]     * scale) - pad + y_off),
                min(pil_img.width, int(box["x1"]     * scale) + pad),
                min(pil_img.height,int(box["bottom"] * scale) + pad + y_off),
            )

        # Row label — blue
        if trace_row_boxes:
            for b in trace_row_boxes:
                r = px(b)
                draw.rectangle(r, fill=(88, 166, 255, 60))
                draw.rectangle(r, outline=(88, 166, 255, 210), width=2)

        # Column header — green
        if trace_col_boxes:
            for b in trace_col_boxes:
                r = px(b)
                draw.rectangle(r, fill=(63, 185, 80, 60))
                draw.rectangle(r, outline=(63, 185, 80, 210), width=2)

        # Value highlight — yellow/orange
        if highlight_boxes:
            for b in highlight_boxes:
                r = px(b)
                draw.rectangle(r, fill=(255, 210, 0, 110))
                draw.rectangle(r, outline=(255, 130, 0, 240), width=2)

        # Connector line: row label → value (horizontal)
        if trace_row_boxes and highlight_boxes:
            row_rx = max(int(b["x1"] * scale) + pad for b in trace_row_boxes)
            val_lx = min(int(b["x0"] * scale) - pad for b in highlight_boxes)
            val_cy = int(((highlight_boxes[0]["top"] + highlight_boxes[0]["bottom"]) / 2) * scale)
            if row_rx < val_lx - 4:
                # dashed line
                x = row_rx
                while x < val_lx - 2:
                    draw.line([(x, val_cy), (min(x + 8, val_lx), val_cy)],
                              fill=(88, 166, 255, 200), width=2)
                    x += 14

        # Connector line: col header → value (vertical)
        if trace_col_boxes and highlight_boxes:
            col_by = max(int(b["bottom"] * scale) + pad for b in trace_col_boxes)
            val_ty = min(int(b["top"]    * scale) - pad for b in highlight_boxes)
            val_cx = int(((highlight_boxes[0]["x0"] + highlight_boxes[0]["x1"]) / 2) * scale)
            if col_by < val_ty - 4:
                y = col_by
                while y < val_ty - 2:
                    draw.line([(val_cx, y), (val_cx, min(y + 8, val_ty))],
                              fill=(63, 185, 80, 200), width=2)
                    y += 14

        pil_img = Image.alpha_composite(pil_img, overlay)
        pil_img = pil_img.convert("RGB")
        buf = io.BytesIO()
        pil_img.save(buf, format="PNG")
        return buf.getvalue()
    except Exception as e:
        print(f"[render_pdf_page] {e}", flush=True)
        return None

# ══════════════════════════════════════════════════════════════════════════════
# PROMPTS
# ══════════════════════════════════════════════════════════════════════════════

CBS_PROMPT = """You are a government finance data extraction expert. Extract ALL data from the Combined Balance Sheet of Governmental Funds from this Louisiana parish audit PDF text.

CRITICAL RULES:
1. Extract from the MAIN Balance Sheet only (not combining/nonmajor schedules)
2. Remove $ signs and commas from numbers → return plain integers
3. A dash "-" or blank means the value is null (not 0)
4. Numbers in parentheses like (123,456) are NEGATIVE: -123456
5. If the statement says "(in thousands)" set "in_thousands": true — do NOT multiply the numbers yourself; return them exactly as printed in the PDF. Our code will apply the scaling.
6. The last fund column is always "Total Governmental Funds"

Return ONLY valid JSON (no markdown, no explanation) with this structure:

{
  "parish": "<parish name>",
  "year": <year as integer>,
  "in_thousands": <true/false>,
  "funds": ["<Fund 1>", "<Fund 2>", ..., "Total Governmental Funds"],
  "assets": {
    "cash_and_deposits": {"<Fund 1>": <int or null>, ...},
    "investments": {"<Fund 1>": <int or null>, ...},
    "taxes_receivable": {"<Fund 1>": <int or null>, ...},
    "special_assessments_receivable": {"<Fund 1>": <int or null>, ...},
    "other_receivables": {"<Fund 1>": <int or null>, ...},
    "due_from_other_governments": {"<Fund 1>": <int or null>, ...},
    "due_from_other_funds": {"<Fund 1>": <int or null>, ...},
    "due_from_component_units": {"<Fund 1>": <int or null>, ...},
    "inventory": {"<Fund 1>": <int or null>, ...},
    "prepaid_items": {"<Fund 1>": <int or null>, ...},
    "other_assets": {"<Fund 1>": <int or null>, ...},
    "total_assets": {"<Fund 1>": <int or null>, ...}
  },
  "liabilities": {
    "accounts_payable": {"<Fund 1>": <int or null>, ...},
    "retainage_payable": {"<Fund 1>": <int or null>, ...},
    "accrued_liabilities": {"<Fund 1>": <int or null>, ...},
    "deposits_payable": {"<Fund 1>": <int or null>, ...},
    "unearned_revenue": {"<Fund 1>": <int or null>, ...},
    "due_to_other_funds": {"<Fund 1>": <int or null>, ...},
    "other_liabilities": {"<Fund 1>": <int or null>, ...},
    "total_liabilities": {"<Fund 1>": <int or null>, ...}
  },
  "deferred_inflows": {
    "items": [{"label": "<description>", "values": {"<Fund 1>": <int or null>, ...}}],
    "total": {"<Fund 1>": <int or null>, ...}
  },
  "fund_balances": {
    "nonspendable": [{"label": "<description>", "values": {"<Fund 1>": <int or null>, ...}}],
    "restricted": [{"label": "<description>", "values": {"<Fund 1>": <int or null>, ...}}],
    "committed": [{"label": "<description>", "values": {"<Fund 1>": <int or null>, ...}}],
    "assigned": [{"label": "<description>", "values": {"<Fund 1>": <int or null>, ...}}],
    "unassigned": {"<Fund 1>": <int or null>, ...},
    "total_fund_balances": {"<Fund 1>": <int or null>, ...}
  },
  "total_liabilities_and_fund_balances": {"<Fund 1>": <int or null>, ...},
  "cross_sectional": {
    "cash": <General Fund cash>, "investments": <General Fund investments>,
    "receivables": <General Fund taxes+assessments+other+due_from_govts>,
    "inventory": <General Fund inventory>, "other_assets": <General Fund other assets>,
    "transfers_in": <General Fund due_from_other_funds>, "prepaid_items": <General Fund prepaid>,
    "total_assets": <General Fund total assets>, "deferred_outflows": <General Fund deferred outflows>,
    "accounts_payable": <General Fund AP+retainage>, "deferred_revenues": <General Fund unearned>,
    "government_transfers": <General Fund due_to_other_funds>,
    "other_liabilities": <General Fund accrued+deposits+other liabilities>,
    "total_liabilities": <General Fund total liabilities>, "deferred_inflows": <General Fund deferred inflows>,
    "reserved": <General Fund nonspendable+restricted>, "unreserved_designated": <General Fund committed+assigned>,
    "unreserved_undesignated": <General Fund unassigned>, "total_fund_balances": <General Fund total fund balances>,
    "total_liabilities_and_fund_balances": <General Fund total>
  }
}

Also include "_uncertainties": an array of any fields where you are not fully confident. Each entry:
{"field":"<field name>","fund":"<fund name or null>","reason":"<why uncertain>","extracted":<value you chose>,"alternative":<other value seen or null>,"page":<page number from [Page N] marker>,"text_snippet":"<exact PDF text causing uncertainty, max 150 chars>","severity":"high|medium|low"}
Return "_uncertainties":[] if everything is clear. Do NOT fabricate uncertainties.

PDF TEXT:
"""

SOA_PROMPT = """You are a government finance data extraction expert. Extract ALL data from the Statement of Activities from this Louisiana parish audit PDF text.

CRITICAL RULES:
1. Net (Expense) Revenue values are typically shown as negative numbers
2. Remove $ signs and commas → plain integers
3. Dashes "-" = null
4. Parentheses (123,456) = negative: -123456
5. If "(in thousands)" set "in_thousands": true — do NOT multiply the numbers yourself.
6. Extract the governmental activities section

Return ONLY valid JSON (no markdown, no explanation):

{
  "parish": "<parish name>", "year": <year>, "in_thousands": <true/false>,
  "governmental_activities": {
    "general_government": {"expenses": <int or null>, "charges_for_services": <int or null>, "operating_grants": <int or null>, "capital_grants": <int or null>, "net_expense_revenue": <int or null>},
    "public_safety": {"expenses": <int or null>, "charges_for_services": <int or null>, "operating_grants": <int or null>, "capital_grants": <int or null>, "net_expense_revenue": <int or null>},
    "public_works": {"expenses": <int or null>, "charges_for_services": <int or null>, "operating_grants": <int or null>, "capital_grants": <int or null>, "net_expense_revenue": <int or null>},
    "economic_development": {"expenses": <int or null>, "charges_for_services": <int or null>, "operating_grants": <int or null>, "capital_grants": <int or null>, "net_expense_revenue": <int or null>},
    "health_and_welfare": {"expenses": <int or null>, "charges_for_services": <int or null>, "operating_grants": <int or null>, "capital_grants": <int or null>, "net_expense_revenue": <int or null>},
    "culture_and_recreation": {"expenses": <int or null>, "charges_for_services": <int or null>, "operating_grants": <int or null>, "capital_grants": <int or null>, "net_expense_revenue": <int or null>},
    "interest_on_long_term_debt": {"expenses": <int or null>, "net_expense_revenue": <int or null>},
    "other_activities": [{"label": "<name>", "expenses": <int or null>, "charges_for_services": <int or null>, "operating_grants": <int or null>, "capital_grants": <int or null>, "net_expense_revenue": <int or null>}],
    "total": {"expenses": <int or null>, "charges_for_services": <int or null>, "operating_grants": <int or null>, "capital_grants": <int or null>, "net_expense_revenue": <int or null>}
  },
  "general_revenues": {
    "property_taxes": <int or null>, "sales_taxes": <int or null>, "severance_taxes": <int or null>,
    "fire_insurance_premiums": <int or null>, "franchise_fees": <int or null>, "other_taxes": <int or null>,
    "occupational_licenses": <int or null>, "gaming_revenues": <int or null>,
    "state_revenue_sharing": <int or null>, "state_shared_revenue": <int or null>,
    "non_employer_pension_contribution": <int or null>, "investment_income": <int or null>,
    "miscellaneous": <int or null>, "transfers": <int or null>,
    "other_items": [{"label": "<name>", "amount": <int or null>}],
    "total_general_revenues": <int or null>
  },
  "change_in_net_position": <int or null>, "net_position_beginning": <int or null>, "net_position_ending": <int or null>,
  "cross_sectional": {
    "property_ad_valorem": <property taxes>, "sales_use_taxes": <sales+gaming>,
    "severance_taxes": <severance>, "other_tax_revenue": <fire+franchise+occupational+other taxes>,
    "total_tax_revenue": <all taxes>, "state_revenue_sharing": <state revenue sharing>,
    "state_intergovernmental": <state grants>, "federal_intergovernmental": <federal grants>,
    "local_transfer": <transfers>, "all_other_revenue": <investment+misc+other>,
    "total_other_revenue": <state+intergovernmental+transfers+other>,
    "total_program_revenue": <charges+operating+capital grants from total>,
    "total_revenues": <tax+other+program>,
    "general_government": <general govt net expense>, "legislative": null, "judicial": null,
    "elections": null, "finance_and_administration": null, "other_general_government": null,
    "total_general_government": <general govt net expense>, "public_safety": <public safety net>,
    "public_works": <public works net>, "economic_development": <econ dev net>,
    "health_and_welfare": <health net>, "culture_and_recreation": <culture net>,
    "interest_debt_service": <interest net>, "all_other_expenditures": <other activities net>,
    "total_expenditures": <total governmental expenses>
  }
}

Also include "_uncertainties": an array of any fields where you are not fully confident. Each entry:
{"field":"<field name>","fund":"<fund name or null>","reason":"<why uncertain>","extracted":<value you chose>,"alternative":<other value seen or null>,"page":<page number from [Page N] marker>,"text_snippet":"<exact PDF text causing uncertainty, max 150 chars>","severity":"high|medium|low"}
Return "_uncertainties":[] if everything is clear. Do NOT fabricate uncertainties.

PDF TEXT:
"""

SONA_PROMPT = """You are a government finance data extraction expert. Extract ALL data from the Statement of Net Position (or Statement of Net Assets) from this Louisiana parish audit PDF text.

CRITICAL RULES:
1. Remove $ signs and commas → plain integers
2. Dashes "-" = null
3. Parentheses (123,456) = negative: -123456
4. If "(in thousands)" set "in_thousands": true — do NOT multiply the numbers yourself.
5. Extract the government-wide statement (Governmental Activities column)

Return ONLY valid JSON (no markdown, no explanation):

{
  "parish": "<parish name>", "year": <year>, "in_thousands": <true/false>,
  "governmental_activities": {
    "current_assets": {
      "cash_and_deposits": <int or null>, "investments": <int or null>, "taxes_receivable": <int or null>,
      "other_receivables": <int or null>, "due_from_other_governments": <int or null>,
      "due_from_component_units": <int or null>, "inventory": <int or null>,
      "prepaid_items": <int or null>, "other_current_assets": <int or null>
    },
    "capital_assets": {
      "non_depreciable": <int or null>, "depreciable_net": <int or null>,
      "right_to_use_net": <int or null>, "total_capital_assets_net": <int or null>
    },
    "other_noncurrent_assets": <int or null>, "total_assets": <int or null>,
    "deferred_outflows": {"items": [{"label": "<description>", "amount": <int or null>}], "total": <int or null>},
    "current_liabilities": {
      "accounts_payable": <int or null>, "retainage_payable": <int or null>,
      "accrued_liabilities": <int or null>, "deposits_payable": <int or null>,
      "unearned_revenue": <int or null>, "accrued_interest": <int or null>,
      "other_current_liabilities": <int or null>
    },
    "long_term_liabilities": {
      "bonds_payable_current": <int or null>, "bonds_payable_noncurrent": <int or null>,
      "compensated_absences_current": <int or null>, "compensated_absences_noncurrent": <int or null>,
      "net_pension_liability": <int or null>, "landfill_closure": <int or null>,
      "lease_liability_current": <int or null>, "lease_liability_noncurrent": <int or null>,
      "other_long_term": <int or null>
    },
    "total_liabilities": <int or null>,
    "deferred_inflows": {"items": [{"label": "<description>", "amount": <int or null>}], "total": <int or null>},
    "net_position": {
      "net_investment_in_capital_assets": <int or null>, "restricted": <int or null>,
      "unrestricted": <int or null>, "total_net_position": <int or null>
    }
  },
  "component_units": {"total_assets": <int or null>, "total_liabilities": <int or null>, "total_net_position": <int or null>}
}

Also include "_uncertainties": an array of any fields where you are not fully confident. Each entry:
{"field":"<field name>","fund":"<fund name or null>","reason":"<why uncertain>","extracted":<value you chose>,"alternative":<other value seen or null>,"page":<page number from [Page N] marker>,"text_snippet":"<exact PDF text causing uncertainty, max 150 chars>","severity":"high|medium|low"}
Return "_uncertainties":[] if everything is clear. Do NOT fabricate uncertainties.

PDF TEXT:
"""

CA_PROMPT = """You are a government finance data extraction expert. Extract ALL data from the Capital Assets schedule from this Louisiana parish audit PDF text.

CRITICAL RULES:
1. Remove $ signs and commas → plain integers
2. Dashes "-" = null
3. Parentheses (123,456) = negative: -123456
4. If "(in thousands)" set "in_thousands": true — do NOT multiply the numbers yourself.
5. Decreases/disposals are typically shown as negative numbers
6. Extract Governmental Activities section

Return ONLY valid JSON (no markdown, no explanation):

{
  "parish": "<parish name>", "year": <year>, "in_thousands": <true/false>,
  "governmental_activities": {
    "not_depreciated": {
      "land": {"beginning": <int or null>, "increases": <int or null>, "decreases": <int or null>, "ending": <int or null>},
      "construction_in_progress": {"beginning": <int or null>, "increases": <int or null>, "decreases": <int or null>, "ending": <int or null>},
      "other_non_depreciable": [{"label": "<name>", "beginning": <int or null>, "increases": <int or null>, "decreases": <int or null>, "ending": <int or null>}],
      "total_not_depreciated": {"beginning": <int or null>, "increases": <int or null>, "decreases": <int or null>, "ending": <int or null>}
    },
    "depreciable": {
      "buildings_and_improvements": {"beginning": <int or null>, "increases": <int or null>, "decreases": <int or null>, "ending": <int or null>},
      "machinery_and_equipment": {"beginning": <int or null>, "increases": <int or null>, "decreases": <int or null>, "ending": <int or null>},
      "improvements_other_than_buildings": {"beginning": <int or null>, "increases": <int or null>, "decreases": <int or null>, "ending": <int or null>},
      "infrastructure": {"beginning": <int or null>, "increases": <int or null>, "decreases": <int or null>, "ending": <int or null>},
      "vehicles": {"beginning": <int or null>, "increases": <int or null>, "decreases": <int or null>, "ending": <int or null>},
      "furniture_and_fixtures": {"beginning": <int or null>, "increases": <int or null>, "decreases": <int or null>, "ending": <int or null>},
      "books_and_periodicals": {"beginning": <int or null>, "increases": <int or null>, "decreases": <int or null>, "ending": <int or null>},
      "leased_property": {"beginning": <int or null>, "increases": <int or null>, "decreases": <int or null>, "ending": <int or null>},
      "other_depreciable": [{"label": "<name>", "beginning": <int or null>, "increases": <int or null>, "decreases": <int or null>, "ending": <int or null>}],
      "total_depreciable": {"beginning": <int or null>, "increases": <int or null>, "decreases": <int or null>, "ending": <int or null>}
    },
    "accumulated_depreciation": {
      "buildings_and_improvements": {"beginning": <int or null>, "increases": <int or null>, "decreases": <int or null>, "ending": <int or null>},
      "machinery_and_equipment": {"beginning": <int or null>, "increases": <int or null>, "decreases": <int or null>, "ending": <int or null>},
      "improvements_other_than_buildings": {"beginning": <int or null>, "increases": <int or null>, "decreases": <int or null>, "ending": <int or null>},
      "infrastructure": {"beginning": <int or null>, "increases": <int or null>, "decreases": <int or null>, "ending": <int or null>},
      "vehicles": {"beginning": <int or null>, "increases": <int or null>, "decreases": <int or null>, "ending": <int or null>},
      "furniture_and_fixtures": {"beginning": <int or null>, "increases": <int or null>, "decreases": <int or null>, "ending": <int or null>},
      "books_and_periodicals": {"beginning": <int or null>, "increases": <int or null>, "decreases": <int or null>, "ending": <int or null>},
      "leased_property": {"beginning": <int or null>, "increases": <int or null>, "decreases": <int or null>, "ending": <int or null>},
      "other": [{"label": "<name>", "beginning": <int or null>, "increases": <int or null>, "decreases": <int or null>, "ending": <int or null>}],
      "total_accumulated_depreciation": {"beginning": <int or null>, "increases": <int or null>, "decreases": <int or null>, "ending": <int or null>}
    },
    "total_depreciable_net": {"beginning": <int or null>, "increases": <int or null>, "decreases": <int or null>, "ending": <int or null>},
    "total_capital_assets_net": {"beginning": <int or null>, "increases": <int or null>, "decreases": <int or null>, "ending": <int or null>}
  },
  "cross_sectional": {
    "land": <land ending>, "construction_in_progress": <CIP ending>, "other_non_depreciable": <other non-dep ending>,
    "buildings_net": <buildings net>, "improvements_net": <improvements net>, "machinery_net": <machinery net>,
    "other_depreciable_net": <other dep net>, "books_net": <books net>, "furniture_net": <furniture net>,
    "vehicles_net": <vehicles net>, "bridges_net": <bridges net>, "leased_property_net": <leased net>,
    "infrastructure_net": <infrastructure net>, "total_governmental_net": <total capital assets net ending>
  }
}

Also include "_uncertainties": an array of any fields where you are not fully confident. Each entry:
{"field":"<field name>","fund":"<fund name or null>","reason":"<why uncertain>","extracted":<value you chose>,"alternative":<other value seen or null>,"page":<page number from [Page N] marker>,"text_snippet":"<exact PDF text causing uncertainty, max 150 chars>","severity":"high|medium|low"}
Return "_uncertainties":[] if everything is clear. Do NOT fabricate uncertainties.

PDF TEXT:
"""

PROMPT_MAP = {"cbs": CBS_PROMPT, "soa": SOA_PROMPT, "sona": SONA_PROMPT, "ca": CA_PROMPT}

# ══════════════════════════════════════════════════════════════════════════════
# PDF EXTRACTION
# ══════════════════════════════════════════════════════════════════════════════

import pdfplumber

MDA_EXCLUDE = ["management's discussion", "management discussion", "managements discussion"]

def _is_notes_page(text):
    first400, very_top = text[:400], text[:200]
    return (
        "notes to financial" in first400
        or "notes to the financial" in first400
        or "notes to the basic" in first400
        or "notes to basic" in first400
        or ("notes to the" in very_top and "statements" in very_top)
        or ("notes to" in text[:150] and "statements" in text[:150])
    )

STATEMENT_PATTERNS = {
    "cbs": {
        "require_any": [
            ["balance sheet", "governmental fund"],
            ["balance sheet", "governmental activities", "assets"],
            ["balance sheet", "governmental fn"],
            ["balance sheet", "govemmental fund"],
            ["balance sheet", "govemmental fn"],
            ["balance sheet", "major fund", "assets"],
            ["balance sheet", "eund"],
            ["balance sheet", "ernmental fund"],
            ["balance sheel", "governmental fund"],
            ["balance sheel", "ernmental fund"],
            ["balance. sheet", "governmental fund"],
            ["balancesheet", "governmental fund"],
            ["balanc e sheet", "ernmental fund"],
            ["balance sheet", "govemmenia"],
            ["governmental fund", "comparative totals", "assets"],
        ],
        "exclude": ["combining balance sheet", "combining statement", "combining balance. sheet",
                    "balance sheet nonmajor", "reconciliation", "reconcl",
                    "discretely presented", "presented separately", "fiduciary"] + MDA_EXCLUDE,
    },
    "soa": {
        "require_any": [
            ["functions/programs"],
            ["statement of activities", "program revenues"],
            ["statement of activities", "charges for"],
            ["net (expense)", "program revenues"],
            ["net (expense) revenue", "general revenues"],
            ["statement of activities", "net (expense)"],
            ["statement of activi", "program revenu"],
            ["statement of activi", "general revenues", "charges f"],
        ],
        "exclude": MDA_EXCLUDE,
    },
    "sona": {
        "require_any": [["statement of net pos"], ["statement of net assets"]],
        "exclude": ["reconciliation", "condensed", "changes in net"] + MDA_EXCLUDE,
    },
    "ca": {
        "require_any": [
            ["capital assets", "not being depreciated"],
            ["capital assets", "depreciable"],
            ["capital assets", "accumulated depreciation"],
        ],
        "exclude": ["policy", "policies", "summary of significant",
                    "fiduciary funds", "internal service funds",
                    "statement of net pos", "statement of net assets",
                    "reconciliation", "capital asset administration",
                    "debt administration", "invested in capital assets", "fiduciary"] + MDA_EXCLUDE,
    },
}

CONFIRMATION = {
    "cbs":  ["assets", "liabilities", "fund balance"],
    "soa":  ["expenses", "general revenues"],
    "sona": ["primary", "assets", "liabilities"],
    "ca":   ["depreciation", "beginning balance", "increases"],
}

CA_NOTES_PATTERNS = [
    ["capital assets", "accumulated depreciation", "beginning"],
    ["capital assets", "beginning balance"],
    ["note", "capital assets", "beginning"],
    ["capital assets", "accumulated depreciation", "additions", "not being depreciated"],
    ["capital assets", "accumulated depreciation", "ending", "not being depreciated"],
    ["capital assets", "accumulated depreciation", "not being depreciated", "increases"],
    ["capital assets", "accumulated depreciation", "cost at december"],
    ["capital asset", "depreciable", "additions", "disposals"],
    ["capital assets", "additions", "deletions"],
]

def _matches_any(text, patterns, allow_notes=False):
    if not allow_notes and _is_notes_page(text):
        return False
    if "table of contents" in text[:400] or ".........." in text:
        return False
    for excl in patterns.get("exclude", []):
        if excl in text:
            return False
    for req_group in patterns.get("require_any", []):
        if all(kw in text for kw in req_group):
            return True
    return False

def _page_text(page):
    return (page.extract_text() or "").lower().replace('\n', ' ')

def find_statement_pages(pdf_path):
    result = {"cbs": [], "soa": [], "sona": [], "ca": []}
    with pdfplumber.open(pdf_path) as pdf:
        n = len(pdf.pages)
        texts = [_page_text(p) for p in pdf.pages]
        for stmt_type, patterns in STATEMENT_PATTERNS.items():
            anchor = None
            for i, text in enumerate(texts):
                if _matches_any(text, patterns):
                    window = " ".join(texts[max(0,i-1):min(n,i+4)])
                    if any(kw in window for kw in CONFIRMATION[stmt_type]):
                        anchor = i; break
            if anchor is None and stmt_type == "ca":
                for i, text in enumerate(texts):
                    if not _is_notes_page(text): continue
                    for req_group in CA_NOTES_PATTERNS:
                        if all(kw in text for kw in req_group):
                            window = " ".join(texts[max(0,i-1):min(n,i+4)])
                            if any(kw in window for kw in CONFIRMATION["ca"]):
                                anchor = i; break
                    if anchor is not None: break
            if anchor is None: continue
            max_cont = 8 if stmt_type == "soa" else 6
            collected = [anchor]
            blank_streak = 0
            for j in range(anchor+1, min(n, anchor+max_cont+1)):
                t = texts[j]
                stop = any([
                    "independent auditor" in t,
                    "management's discussion" in t,
                    _is_notes_page(t),
                    "required supplementary" in t,
                    ("combining balance sheet" in t and stmt_type == "cbs"),
                    ("combining" in t and "nonmajor" in t and stmt_type == "cbs"),
                    "table of contents" in t,
                    ("reconciliation of" in t and stmt_type == "cbs"),
                    ("balance sheet" in t[:300] and stmt_type == "soa"),
                    ("balance sheel" in t[:300] and stmt_type == "soa"),
                    ("balance. sheet" in t[:300] and stmt_type == "soa"),
                    ("functions/programs" in t and stmt_type in ("sona","ca")),
                    ("statement of activities" in t[:300] and stmt_type in ("sona","ca")),
                    ("statement of net pos" in t[:300] and stmt_type == "ca"),
                    ("balance sheet" in t and "governmental fund" in t and stmt_type in ("sona","ca")),
                    ("balance sheet" in t and "ernmental fund" in t and stmt_type in ("sona","ca")),
                ])
                if stop: break
                if len(t) > 150:
                    collected.append(j); blank_streak = 0
                elif len(t) < 50:
                    blank_streak += 1
                    if blank_streak >= 2: break
            result[stmt_type] = collected
    return result

def extract_statement_text(pdf_path, page_indices):
    parts = []
    with pdfplumber.open(pdf_path) as pdf:
        for i in page_indices:
            text = pdf.pages[i].extract_text() or ""
            if text.strip():
                parts.append(f"[Page {i+1}]\n{text}")
    return "\n\n".join(parts)

def get_parish_name(pdf_path):
    return os.path.splitext(os.path.basename(pdf_path))[0]

def get_all_statement_texts(pdf_path):
    page_map = find_statement_pages(pdf_path)
    return {t: (extract_statement_text(pdf_path, pages) if pages else None)
            for t, pages in page_map.items()}

# ══════════════════════════════════════════════════════════════════════════════
# CLAUDE API
# ══════════════════════════════════════════════════════════════════════════════

import anthropic as _anthropic

_client = None

def _get_client():
    global _client
    key = _get_api_key()
    if not key:
        raise RuntimeError("No API key configured. Open Settings (gear icon) and paste your Anthropic key.")
    # Re-create client if key changed
    if _client is None or getattr(_client, '_api_key_used', None) != key:
        _client = _anthropic.Anthropic(api_key=key)
        _client._api_key_used = key
    return _client

def _scale_dict(d, factor):
    if isinstance(d, dict):   return {k: _scale_dict(v, factor) for k, v in d.items()}
    if isinstance(d, list):   return [_scale_dict(i, factor) for i in d]
    if isinstance(d, (int, float)): return int(d * factor)
    return d

def _apply_thousands(data):
    if not data or not data.get("in_thousands"): return data
    skip = {"parish", "year", "in_thousands", "funds"}
    return {k: (_scale_dict(v, 1000) if k not in skip else v) for k, v in data.items()}

def extract_statement(stmt_type, text, parish, year, past_corrections=None, statement_notes=None):
    if not text or not text.strip(): return None
    try:
        tokens = 8192 if stmt_type == "cbs" else 4096
        prompt = PROMPT_MAP[stmt_type]
        # Build human-reviewer preamble: statement-level notes FIRST, then field corrections
        if statement_notes or past_corrections:
            preamble = [f"\nIMPORTANT — Human reviewer instructions for {parish} {stmt_type.upper()}:",
                        "(These override any conflicting interpretation of the PDF text.)"]
            if statement_notes:
                preamble.append("")
                preamble.append("Statement-level context from auditor:")
                for note in statement_notes:
                    preamble.append(f"  ★ {note}")
            if past_corrections:
                preamble.append("")
                preamble.append("Prior field-level corrections (human-verified):")
                for c in past_corrections:
                    line = f"  • {c['field_path']}: extracted {c['extracted_value']}"
                    if c["action"] == "corrected":
                        line += f" → correct value is {c['corrected_value']}"
                    elif c["action"] == "confirmed":
                        line += " (confirmed correct)"
                    if c.get("user_note"):
                        line += f"  [{c['user_note']}]"
                    preamble.append(line)
            prompt = prompt + "\n".join(preamble) + "\n\n"
        _model = "claude-haiku-4-5"
        resp = _get_client().messages.create(
            model=_model,
            max_tokens=tokens,
            messages=[{"role": "user", "content": prompt + text[:14000]}],
        )
        raw = resp.content[0].text.strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"): raw = raw[4:]
        data = json.loads(raw.strip())
        data["parish"] = parish
        data["year"] = year
        # Extract uncertainties BEFORE scaling so values stay raw
        uncertainties = data.pop("_uncertainties", None) or []
        scaled = _apply_thousands(data)
        if scaled:
            # Scale uncertainty extracted/alternative values if in_thousands
            if data.get("in_thousands"):
                for u in uncertainties:
                    for k in ("extracted", "alternative"):
                        if isinstance(u.get(k), (int, float)):
                            u[k] = int(u[k] * 1000)
            scaled["_uncertainties"] = uncertainties
        return scaled
    except Exception as e:
        print(f"  [ERROR] {stmt_type} ({parish}): {e}")
        return None

_STMT_LABELS = {"cbs": "Balance Sheet", "soa": "Activities", "sona": "Net Position", "ca": "Capital Assets"}

def extract_all_statements(texts, parish, year, db_path=None, log_fn=None):
    _log = log_fn or (lambda msg: print(msg, flush=True))
    results = {}
    for stmt_type in ["cbs", "soa", "sona", "ca"]:
        corrections = []
        stmt_notes = []
        if db_path:
            try:
                from feedback import get_corrections_for_parish, get_statement_notes
                corrections = get_corrections_for_parish(db_path, parish, year, stmt_type)
                stmt_notes = get_statement_notes(db_path, parish, year, stmt_type)
            except Exception:
                pass
        label = _STMT_LABELS.get(stmt_type, stmt_type.upper())
        has_text = bool(texts.get(stmt_type) and texts[stmt_type].strip())
        if has_text:
            _log(f"    [haiku] {label}...")
        else:
            _log(f"    [haiku] {label} — no text found, skipping")
        results[stmt_type] = extract_statement(
            stmt_type, texts.get(stmt_type), parish, year,
            past_corrections=corrections or None,
            statement_notes=stmt_notes or None,
        )
        time.sleep(0.5)
    return results

# ══════════════════════════════════════════════════════════════════════════════
# EXCEL WRITERS
# ══════════════════════════════════════════════════════════════════════════════

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def _bold(ws, r, c, v):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(bold=True, name="Arial", size=10)
    return cell

def _val(ws, r, c, v):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name="Arial", size=10)
    if isinstance(v, (int, float)):
        cell.number_format = '#,##0;(#,##0);"-"'
    return cell

def _hdr(ws, r, c, v):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(bold=True, name="Arial", size=10)
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    return cell

def _title(ws, r, c, v):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(bold=True, name="Arial", size=12)
    cell.alignment = Alignment(horizontal="center")
    return cell

def _cw(ws, c, w):
    ws.column_dimensions[get_column_letter(c)].width = w

def _safe(d, *keys):
    for k in keys:
        if not isinstance(d, dict): return None
        d = d.get(k)
    return d

# ── CBS ────────────────────────────────────────────────────────────────────────

def write_cbs_tab(ws, data):
    parish, year = data.get("parish",""), data.get("year","")
    funds = data.get("funds", ["Total Governmental Funds"])
    if "Total Governmental Funds" not in funds:
        funds = funds + ["Total Governmental Funds"]
    DC = 6
    _title(ws,1,4,f"{parish} Combined Balance Sheet")
    _title(ws,2,4,f"Issued: December 31, {year}")
    for j,f in enumerate(funds): _hdr(ws,4,DC+j,f)
    row = 5
    _bold(ws,row,1,"Assets:"); row+=1
    for label,field in [
        ("Cash and interest bearing deposits","cash_and_deposits"),("Investments","investments"),
        ("Taxes Receivable","taxes_receivable"),("Special assessments receivable","special_assessments_receivable"),
        ("Other receivables","other_receivables"),("Due from other governmental agencies","due_from_other_governments"),
        ("Due from other funds","due_from_other_funds"),("Due from component units","due_from_component_units"),
        ("Inventory","inventory"),("Prepaid items","prepaid_items"),("Other assets","other_assets"),
    ]:
        vals = _safe(data,"assets",field) or {}
        if not any(v is not None for v in vals.values()): continue
        _val(ws,row,2,label)
        for j,f in enumerate(funds): _val(ws,row,DC+j,vals.get(f))
        row+=1
    row+=1
    ta = _safe(data,"assets","total_assets") or {}
    _bold(ws,row,5,"Total assets:")
    for j,f in enumerate(funds): _val(ws,row,DC+j,ta.get(f))
    row+=2
    _bold(ws,row,1,"LIABILITIES AND FUND BALANCES"); row+=1
    _bold(ws,row,2,"Liabilities:"); row+=1
    for label,field in [
        ("Accounts payable","accounts_payable"),("Retainage payable","retainage_payable"),
        ("Accrued liabilities and other payables","accrued_liabilities"),("Deposits payable","deposits_payable"),
        ("Unearned revenue","unearned_revenue"),("Due to other funds","due_to_other_funds"),("Other liabilities","other_liabilities"),
    ]:
        vals = _safe(data,"liabilities",field) or {}
        if not any(v is not None for v in vals.values()): continue
        _val(ws,row,3,label)
        for j,f in enumerate(funds): _val(ws,row,DC+j,vals.get(f))
        row+=1
    row+=1
    tl = _safe(data,"liabilities","total_liabilities") or {}
    _bold(ws,row,5,"Total liabilities:")
    for j,f in enumerate(funds): _val(ws,row,DC+j,tl.get(f))
    row+=2
    di_items = _safe(data,"deferred_inflows","items") or []
    di_total = _safe(data,"deferred_inflows","total") or {}
    if di_items or any(v for v in di_total.values() if v):
        _bold(ws,row,1,"Deferred Inflows of Resources:"); row+=1
        for item in di_items:
            vals = item.get("values",{})
            if any(v is not None for v in vals.values()):
                _val(ws,row,2,item.get("label",""))
                for j,f in enumerate(funds): _val(ws,row,DC+j,vals.get(f))
                row+=1
        row+=1
    _bold(ws,row,2,"Fund Balances"); row+=1
    for cat in ["nonspendable","restricted","committed","assigned"]:
        items = _safe(data,"fund_balances",cat) or []
        if not items: continue
        if isinstance(items,list):
            for item in items:
                vals = item.get("values",{})
                if any(v is not None for v in vals.values()):
                    _val(ws,row,3,cat.capitalize()+":")
                    _val(ws,row,4,item.get("label",""))
                    for j,f in enumerate(funds): _val(ws,row,DC+j,vals.get(f))
                    row+=1
        elif isinstance(items,dict):
            if any(v is not None for v in items.values()):
                _val(ws,row,3,cat.capitalize())
                for j,f in enumerate(funds): _val(ws,row,DC+j,items.get(f))
                row+=1
    ua = _safe(data,"fund_balances","unassigned") or {}
    if any(v is not None for v in ua.values()):
        _val(ws,row,3,"Unassigned")
        for j,f in enumerate(funds): _val(ws,row,DC+j,ua.get(f))
        row+=1
    row+=1
    tfb = _safe(data,"fund_balances","total_fund_balances") or {}
    _bold(ws,row,5,"Total fund balances:")
    for j,f in enumerate(funds): _val(ws,row,DC+j,tfb.get(f))
    row+=2
    tlfb = data.get("total_liabilities_and_fund_balances",{}) or {}
    _bold(ws,row,5,"Total liabilities and fund balances:")
    for j,f in enumerate(funds): _val(ws,row,DC+j,tlfb.get(f))
    for c,w in [(1,28),(2,32),(3,32),(4,30),(5,28)]: _cw(ws,c,w)
    for j in range(len(funds)): _cw(ws,DC+j,18)

def _cbs_cs_row(ws, data, row):
    cs = data.get("cross_sectional",{}) or {}
    cols = [data.get("parish",""),cs.get("cash"),cs.get("investments"),cs.get("receivables"),
            cs.get("inventory"),cs.get("other_assets"),cs.get("transfers_in"),cs.get("prepaid_items"),
            cs.get("total_assets"),None,cs.get("total_assets"),cs.get("deferred_outflows"),
            cs.get("accounts_payable"),cs.get("deferred_revenues"),cs.get("government_transfers"),
            cs.get("other_liabilities"),cs.get("total_liabilities"),None,cs.get("total_liabilities"),
            cs.get("deferred_inflows"),cs.get("reserved"),cs.get("unreserved_designated"),
            cs.get("unreserved_undesignated"),cs.get("total_fund_balances"),
            cs.get("total_liabilities_and_fund_balances"),None,cs.get("total_liabilities_and_fund_balances")]
    for j,v in enumerate(cols,1):
        c=ws.cell(row=row,column=j,value=v); c.font=Font(name="Arial",size=10)
        if isinstance(v,(int,float)): c.number_format='#,##0;(#,##0);"-"'

def _cbs_cs_init(ws, year):
    ws.title="Cross Sectional"
    r3=[f"{year}-CBS","Cash","Investments","Receivables","Inventory","Other Assets","Transfers","Prepaid Items",
        "Total Assets (Sum)","Accuracy","Accuracy","Deferred Outflows","Accounts Payable","Deferred Revenues",
        "Government Transfers","Other Liabilities","Total Liabilities (Sum)","Accuracy","Accuracy","Deferred Inflows",
        "Reserved","Unreserved (Designated)","Unreserved (Undesignated)","Total Fund Balances (Sum)",
        "Total Liabilities and Fund Balances (Sum)","Accuracy","Accuracy"]
    for j,v in enumerate(r3,1):
        c=ws.cell(row=3,column=j,value=v); c.font=Font(bold=True,name="Arial",size=10)
        c.alignment=Alignment(wrap_text=True,horizontal="center")
        ws.column_dimensions[get_column_letter(j)].width=16
    ws.column_dimensions["A"].width=24

# ── SOA ───────────────────────────────────────────────────────────────────────

def write_soa_tab(ws, data):
    parish, year = data.get("parish",""), data.get("year","")
    _title(ws,1,6,f"{parish} {year} Statement of Activities")
    _title(ws,2,6,f"Issued: December 31, {year}")
    for j,h in enumerate(["Functions/Programs","","","","","Expenses","Charges for Services",
                           "Operating Grants","Capital Grants","Total Governmental Activities","Component Units"],1):
        _hdr(ws,5,j,h)
    row=6
    ga=data.get("governmental_activities",{}) or {}
    _bold(ws,row,1,"Primary Government:"); row+=1
    _bold(ws,row,2,"Government Activities:"); row+=1
    for label,key in [("General government","general_government"),("Public safety","public_safety"),
                      ("Public works","public_works"),("Economic development","economic_development"),
                      ("Health and welfare","health_and_welfare"),("Culture and recreation","culture_and_recreation"),
                      ("Interest on long-term debt","interest_on_long_term_debt")]:
        act=ga.get(key,{}) or {}
        if not act: continue
        _val(ws,row,3,label); _val(ws,row,6,act.get("expenses")); _val(ws,row,7,act.get("charges_for_services"))
        _val(ws,row,8,act.get("operating_grants")); _val(ws,row,9,act.get("capital_grants"))
        _val(ws,row,10,act.get("net_expense_revenue")); row+=1
    for other in (ga.get("other_activities") or []):
        _val(ws,row,3,other.get("label","")); _val(ws,row,6,other.get("expenses"))
        _val(ws,row,7,other.get("charges_for_services")); _val(ws,row,8,other.get("operating_grants"))
        _val(ws,row,9,other.get("capital_grants")); _val(ws,row,10,other.get("net_expense_revenue")); row+=1
    row+=1; tot=ga.get("total",{}) or {}
    _bold(ws,row,5,"Total Government Activities")
    _val(ws,row,6,tot.get("expenses")); _val(ws,row,7,tot.get("charges_for_services"))
    _val(ws,row,8,tot.get("operating_grants")); _val(ws,row,9,tot.get("capital_grants"))
    _val(ws,row,10,tot.get("net_expense_revenue")); row+=2
    _bold(ws,row,6,"General Revenues:"); row+=1
    gr=data.get("general_revenues",{}) or {}
    for label,key in [("Property taxes","property_taxes"),("Sales and use taxes","sales_taxes"),
                      ("Severance taxes","severance_taxes"),("Fire insurance premiums","fire_insurance_premiums"),
                      ("Franchise fees","franchise_fees"),("Other taxes","other_taxes"),
                      ("Occupational licenses","occupational_licenses"),("Gaming revenues","gaming_revenues"),
                      ("State revenue sharing","state_revenue_sharing"),("Investment income","investment_income"),
                      ("Miscellaneous","miscellaneous"),("Transfers","transfers")]:
        v=gr.get(key)
        if v is None: continue
        _val(ws,row,7,label); _val(ws,row,10,v); row+=1
    for other in (gr.get("other_items") or []):
        _val(ws,row,7,other.get("label","")); _val(ws,row,10,other.get("amount")); row+=1
    row+=1; _bold(ws,row,6,"Total General Revenues"); _val(ws,row,10,gr.get("total_general_revenues")); row+=2
    _bold(ws,row,6,"Change in Net Assets"); _val(ws,row,10,data.get("change_in_net_position")); row+=2
    _bold(ws,row,6,"Net position, beginning"); _val(ws,row,10,data.get("net_position_beginning")); row+=1
    _bold(ws,row,6,"Net position, ending"); _val(ws,row,10,data.get("net_position_ending"))
    for c in range(1,6): _cw(ws,c,5)
    for c,w in [(6,30),(7,22),(8,22),(9,22),(10,22),(11,16)]: _cw(ws,c,w)

def _soa_cs_row(ws, data, row):
    cs=data.get("cross_sectional",{}) or {}
    cols=[data.get("parish",""),cs.get("property_ad_valorem"),cs.get("sales_use_taxes"),
          cs.get("severance_taxes"),cs.get("other_tax_revenue"),cs.get("total_tax_revenue"),
          cs.get("state_revenue_sharing"),cs.get("state_intergovernmental"),cs.get("federal_intergovernmental"),
          cs.get("local_transfer"),cs.get("all_other_revenue"),cs.get("total_other_revenue"),
          cs.get("total_program_revenue"),cs.get("total_revenues"),None,None,
          cs.get("total_general_government"),cs.get("legislative"),cs.get("judicial"),cs.get("elections"),
          cs.get("finance_and_administration"),cs.get("other_general_government"),cs.get("total_general_government"),
          cs.get("public_safety"),cs.get("public_works"),cs.get("economic_development"),
          cs.get("health_and_welfare"),cs.get("culture_and_recreation"),cs.get("interest_debt_service"),
          cs.get("all_other_expenditures"),cs.get("total_expenditures"),None,None]
    for j,v in enumerate(cols,1):
        c=ws.cell(row=row,column=j,value=v); c.font=Font(name="Arial",size=10)
        if isinstance(v,(int,float)): c.number_format='#,##0;(#,##0);"-"'

def _soa_cs_init(ws, year):
    ws.title="Governmental Cross Sectional"
    r3=[f"{year}-SOA","Property/Ad Valorem","Sales/Use","Severance","Other Tax","Total Tax (SUM)",
        "State Revenue Sharing","State Intergovernmental","Federal Intergovernmental","Local Transfer",
        "All Other Revenue","Total Other Revenue (SUM)","Total Program Revenue (SUM)","Total Revenues (Sum)",
        "Accuracy","Accuracy","General Government","Legislative","Judicial","Elections",
        "Finance and Admin","Other General Govt","Total General Govt (SUM)","Public Safety","Public Works",
        "Economic Development","Health and Welfare","Culture and Recreation","Interest/Debt Service",
        "All Other Expenditures","Total Expenditures (Sum)","Accuracy","Accuracy"]
    for j,v in enumerate(r3,1):
        c=ws.cell(row=3,column=j,value=v); c.font=Font(bold=True,name="Arial",size=10)
        c.alignment=Alignment(wrap_text=True,horizontal="center")
        ws.column_dimensions[get_column_letter(j)].width=16
    ws.column_dimensions["A"].width=28

# ── SONA ──────────────────────────────────────────────────────────────────────

def write_sona_tab(ws, data):
    parish, year = data.get("parish",""), data.get("year","")
    _title(ws,1,4,f"{parish} Parish {year} Statement of Net Assets")
    _title(ws,2,4,f"Issued: December 31, {year}")
    _hdr(ws,4,8,"Primary Government"); _hdr(ws,5,8,"Governmental Activities"); _hdr(ws,5,9,"Component Unit")
    ga=data.get("governmental_activities",{}) or {}; cu=data.get("component_units",{}) or {}
    row=6; _bold(ws,row,1,"ASSETS"); row+=1; _bold(ws,row,2,"Current Assets"); row+=1
    ca_cur=ga.get("current_assets",{}) or {}
    for label,key in [("Cash","cash_and_deposits"),("Investments","investments"),
                      ("Taxes receivable","taxes_receivable"),("Other receivables","other_receivables"),
                      ("Due from other governmental entities","due_from_other_governments"),
                      ("Due from component units","due_from_component_units"),("Inventory","inventory"),
                      ("Prepaid items","prepaid_items"),("Other current assets","other_current_assets")]:
        v=ca_cur.get(key)
        if v is None: continue
        _val(ws,row,3,label); _val(ws,row,8,v); row+=1
    ca_d=ga.get("capital_assets",{}) or {}
    if ca_d:
        _val(ws,row,3,"Capital assets:"); row+=1
        for label,key in [("Non-depreciable","non_depreciable"),("Depreciable, net","depreciable_net"),
                          ("Right-to-use lease assets, net","right_to_use_net")]:
            v=ca_d.get(key)
            if v is None: continue
            _val(ws,row,4,label); _val(ws,row,8,v); row+=1
    row+=1; _bold(ws,row,3,"Total Assets"); _val(ws,row,8,ga.get("total_assets")); _val(ws,row,9,cu.get("total_assets")); row+=2
    do_d=ga.get("deferred_outflows",{}) or {}
    if do_d.get("total"):
        _bold(ws,row,1,"DEFERRED OUTFLOWS OF RESOURCES"); row+=1
        for item in (do_d.get("items") or []):
            _val(ws,row,3,item.get("label","")); _val(ws,row,8,item.get("amount")); row+=1
        row+=1
    _bold(ws,row,1,"LIABILITIES"); row+=1
    cl=ga.get("current_liabilities",{}) or {}
    for label,key in [("Accounts payable","accounts_payable"),("Retainage payable","retainage_payable"),
                      ("Accrued liabilities","accrued_liabilities"),("Deposits payable","deposits_payable"),
                      ("Unearned revenue","unearned_revenue"),("Accrued interest","accrued_interest"),
                      ("Other current liabilities","other_current_liabilities")]:
        v=cl.get(key)
        if v is None: continue
        _val(ws,row,3,label); _val(ws,row,8,v); row+=1
    lt=ga.get("long_term_liabilities",{}) or {}
    if lt:
        _val(ws,row,3,"Long-term liabilities:"); row+=1
        for label,key in [("Bonds payable - current","bonds_payable_current"),("Bonds payable - long-term","bonds_payable_noncurrent"),
                          ("Compensated absences - current","compensated_absences_current"),("Compensated absences - long-term","compensated_absences_noncurrent"),
                          ("Net pension liability","net_pension_liability"),("Landfill closure","landfill_closure"),
                          ("Lease liability - current","lease_liability_current"),("Lease liability - long-term","lease_liability_noncurrent"),
                          ("Other long-term","other_long_term")]:
            v=lt.get(key)
            if v is None: continue
            _val(ws,row,4,label); _val(ws,row,8,v); row+=1
    row+=1; _bold(ws,row,4,"Total Liabilities"); _val(ws,row,8,ga.get("total_liabilities")); _val(ws,row,9,cu.get("total_liabilities")); row+=2
    di_d=ga.get("deferred_inflows",{}) or {}
    if di_d.get("total"):
        _bold(ws,row,1,"DEFERRED INFLOWS OF RESOURCES"); row+=1
        for item in (di_d.get("items") or []):
            _val(ws,row,3,item.get("label","")); _val(ws,row,8,item.get("amount")); row+=1
        row+=1
    np=ga.get("net_position",{}) or {}
    _bold(ws,row,1,"NET POSITION"); row+=1
    for label,key in [("Net investment in capital assets","net_investment_in_capital_assets"),
                      ("Restricted","restricted"),("Unrestricted","unrestricted")]:
        v=np.get(key)
        if v is None: continue
        _val(ws,row,3,label); _val(ws,row,8,v); row+=1
    row+=1; _bold(ws,row,4,"Total Net Position"); _val(ws,row,8,np.get("total_net_position")); _val(ws,row,9,cu.get("total_net_position"))
    for c in range(1,8): _cw(ws,c,6)
    _cw(ws,8,22); _cw(ws,9,18)

# ── Capital Assets ─────────────────────────────────────────────────────────────

def write_ca_tab(ws, data):
    parish, year = data.get("parish",""), data.get("year","")
    _title(ws,1,4,f"{parish.upper()} PARISH POLICE JURY")
    _title(ws,2,4,"Capital Assets")
    _title(ws,3,4,f"For the Year Ended December 31, {year}")
    for j,h in enumerate(["Beginning Balance","Increases","Decreases","Ending Balance","Cross Sectional"],start=9):
        _hdr(ws,5,j,h)
    ga=data.get("governmental_activities",{}) or {}; row=[7]
    def ca_section(title, sec_key, items_cfg):
        _bold(ws,row[0],1,title); row[0]+=1
        sec=ga.get(sec_key,{}) or {}
        for label,field in items_cfg:
            d=sec.get(field)
            if isinstance(d,dict) and any(v is not None for v in d.values()):
                _val(ws,row[0],2,label)
                for co,k in enumerate(["beginning","increases","decreases","ending"],start=9):
                    _val(ws,row[0],co,d.get(k))
                _val(ws,row[0],13,d.get("ending")); row[0]+=1
        other_key="other_non_depreciable" if sec_key=="not_depreciated" else "other_depreciable"
        for item in (sec.get(other_key) or []):
            if isinstance(item,dict):
                _val(ws,row[0],2,item.get("label",""))
                for co,k in enumerate(["beginning","increases","decreases","ending"],start=9):
                    _val(ws,row[0],co,item.get(k))
                row[0]+=1
    ca_section("Capital assets, not being depreciated:","not_depreciated",[("Land","land"),("Construction in progress","construction_in_progress")])
    tnd=_safe(ga,"not_depreciated","total_not_depreciated") or {}
    if tnd:
        _val(ws,row[0],3,"Total assets not being depreciated")
        for co,k in enumerate(["beginning","increases","decreases","ending"],start=9): _val(ws,row[0],co,tnd.get(k))
        row[0]+=2
    ca_section("Capital assets being depreciated:","depreciable",[
        ("Buildings and improvements","buildings_and_improvements"),("Machinery and equipment","machinery_and_equipment"),
        ("Improvements, other than buildings","improvements_other_than_buildings"),("Infrastructure","infrastructure"),
        ("Vehicles","vehicles"),("Furniture and fixtures","furniture_and_fixtures"),
        ("Books and periodicals","books_and_periodicals"),("Leased property","leased_property")])
    tdep=_safe(ga,"depreciable","total_depreciable") or {}
    if tdep:
        _val(ws,row[0],3,"Total capital assets being depreciated")
        for co,k in enumerate(["beginning","increases","decreases","ending"],start=9): _val(ws,row[0],co,tdep.get(k))
        row[0]+=2
    _bold(ws,row[0],1,"Less accumulated depreciation for:"); row[0]+=1
    acc=ga.get("accumulated_depreciation",{}) or {}
    for label,key in [("Buildings and improvements","buildings_and_improvements"),("Machinery and equipment","machinery_and_equipment"),
                      ("Improvements, other than buildings","improvements_other_than_buildings"),("Infrastructure","infrastructure"),
                      ("Vehicles","vehicles"),("Furniture and fixtures","furniture_and_fixtures"),
                      ("Books and periodicals","books_and_periodicals"),("Leased property","leased_property")]:
        d=acc.get(key)
        if isinstance(d,dict) and any(v is not None for v in d.values()):
            _val(ws,row[0],2,label)
            for co,fk in enumerate(["beginning","increases","decreases","ending"],start=9): _val(ws,row[0],co,d.get(fk))
            _val(ws,row[0],13,d.get("ending")); row[0]+=1
    tacc=acc.get("total_accumulated_depreciation") or {}
    if tacc:
        _val(ws,row[0],3,"Total accumulated depreciation")
        for co,k in enumerate(["beginning","increases","decreases","ending"],start=9): _val(ws,row[0],co,tacc.get(k))
        row[0]+=2
    tnet=ga.get("total_capital_assets_net") or {}
    if tnet:
        _bold(ws,row[0],2,"Governmental activities capital assets, net")
        for co,k in enumerate(["beginning","increases","decreases","ending"],start=9): _val(ws,row[0],co,tnet.get(k))
        _val(ws,row[0],13,tnet.get("ending"))
    for c in range(1,9): _cw(ws,c,6 if c<4 else 26)
    for c in range(9,14): _cw(ws,c,18)

def _ca_cs_row(ws, data, row):
    cs=data.get("cross_sectional",{}) or {}
    cols=[data.get("parish",""),cs.get("land"),cs.get("construction_in_progress"),cs.get("other_non_depreciable"),
          cs.get("buildings_net"),cs.get("improvements_net"),cs.get("machinery_net"),cs.get("other_depreciable_net"),
          cs.get("books_net"),cs.get("furniture_net"),cs.get("vehicles_net"),cs.get("bridges_net"),
          cs.get("leased_property_net"),cs.get("infrastructure_net"),cs.get("total_governmental_net"),
          cs.get("total_governmental_net"),None]
    for j,v in enumerate(cols,1):
        c=ws.cell(row=row,column=j,value=v); c.font=Font(name="Arial",size=10)
        if isinstance(v,(int,float)): c.number_format='#,##0;(#,##0);"-"'

def _ca_cs_init(ws, year):
    ws.title="Cross Sectional"
    headers=["Parish","Land","Construction in progress","Other non-depreciable","Buildings and improvements",
             "Improvements other than buildings","Machinery and equipment","Other depreciable","Books and periodicals",
             "Furniture and fixtures","Vehicles","Bridges","Leased Property","Infrastructure","Total (Net)","Check","Okay?"]
    ws.cell(row=1,column=1,value="Louisiana Counties").font=Font(bold=True,name="Arial",size=11)
    ws.cell(row=2,column=1,value=f"{year} Capital Assets, Net").font=Font(bold=True,name="Arial",size=10)
    for j,h in enumerate(headers,1):
        c=ws.cell(row=4,column=j,value=h); c.font=Font(bold=True,name="Arial",size=10)
        c.alignment=Alignment(wrap_text=True,horizontal="center")
        ws.column_dimensions[get_column_letter(j)].width=16
    ws.column_dimensions["A"].width=22

STMT_CONFIGS = {
    "cbs":  {"filename":"{label} {year} CBS.xlsx",           "tab_writer":write_cbs_tab,  "cs_writer":_cbs_cs_row, "cs_init":_cbs_cs_init},
    "soa":  {"filename":"{label} {year} SOA.xlsx",           "tab_writer":write_soa_tab,  "cs_writer":_soa_cs_row, "cs_init":_soa_cs_init},
    "sona": {"filename":"{label} {year} SONA.xlsx",          "tab_writer":write_sona_tab, "cs_writer":None,        "cs_init":None},
    "ca":   {"filename":"{label} {year} Capital Assets.xlsx","tab_writer":write_ca_tab,   "cs_writer":_ca_cs_row,  "cs_init":_ca_cs_init},
}

def _get_or_create_wb(output_dir, stmt_type, year, label):
    cfg = STMT_CONFIGS[stmt_type]
    path = os.path.join(output_dir, cfg["filename"].format(year=year, label=label))
    if os.path.exists(path):
        wb = load_workbook(path)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames: del wb["Sheet"]
        if cfg["cs_init"]:
            cfg["cs_init"](wb.create_sheet("Cross Sectional" if stmt_type!="soa" else "Governmental Cross Sectional"), year)
        else:
            wb.create_sheet("Overview")
    return wb, path

def _add_parish_to_wb(wb, stmt_type, data, year):
    cfg = STMT_CONFIGS[stmt_type]
    parish = data.get("parish","Unknown")
    if parish in wb.sheetnames: del wb[parish]
    cfg["tab_writer"](wb.create_sheet(parish), data)
    if cfg["cs_writer"]:
        cs_name = "Cross Sectional" if stmt_type!="soa" else "Governmental Cross Sectional"
        if cs_name not in wb.sheetnames:
            cfg["cs_init"](wb.create_sheet(cs_name), year)
        cs_ws = wb[cs_name]
        next_row = 6
        while cs_ws.cell(row=next_row,column=1).value: next_row+=1
        cfg["cs_writer"](cs_ws, data, next_row)

def _write_parish_combined(output_dir, parish, results, year,
                           parish_validation=None, manual_pages=None):
    wb = Workbook()
    if "Sheet" in wb.sheetnames: del wb["Sheet"]
    wrote = False
    for stmt_type, sheet_name in [("cbs","CBS"),("soa","SOA"),("sona","SONA"),("ca","Capital Assets")]:
        data = results.get(stmt_type)
        if not data: continue
        STMT_CONFIGS[stmt_type]["tab_writer"](wb.create_sheet(sheet_name), data)
        wrote = True
    if not wrote: return None
    if parish_validation is not None:
        try:
            from validate import write_quality_report_excel_tab as _wqrx
            _wqrx(wb, [parish_validation], manual_pages or {}, year)
        except Exception:
            pass
    fname = f"{parish} {year}.xlsx"
    wb.save(os.path.join(output_dir, fname))
    return fname

# ══════════════════════════════════════════════════════════════════════════════
# PIPELINE ORCHESTRATOR
# ══════════════════════════════════════════════════════════════════════════════

def _load_manual_pages():
    path = _resource_dir() / "manual_pages.json"
    if path.exists():
        data = json.loads(path.read_text())
        return {k: v for k, v in data.items() if not k.startswith("_")}
    return {}

def _apply_manual(pdf_path, parish, manual_pages):
    if parish not in manual_pages: return None
    result = {}
    for stmt_type in ["cbs","soa","sona","ca"]:
        overrides = manual_pages[parish]
        if stmt_type in overrides:
            pages = [p-1 for p in overrides[stmt_type]]
            text = extract_statement_text(pdf_path, pages)
            result[stmt_type] = text if text.strip() else None
        else:
            result[stmt_type] = None
    return result

def run_pipeline(input_dir, year, output_dir, cache_dir, skip_cache=False,
                 progress_callback=print, job_id=None):
    log = progress_callback

    # ── Optional pipeline modules ────────────────────────────────────────────
    _pipeline_dir = str(_resource_dir() / "pipeline")
    if _pipeline_dir not in sys.path:
        sys.path.insert(0, _pipeline_dir)

    try:
        from validate import (validate_parish as _vp,
                              write_quality_report_csv as _wqrc,
                              write_quality_report_excel_tab as _wqrx)
        _validate_available = True
    except Exception:
        _validate_available = False

    try:
        from feedback import (init_db as _fb_init, save_uncertainty as _fb_save,
                              get_uncertainties_for_job as _fb_get)
        _db_path = str(_app_dir() / "feedback.db")
        _fb_init(_db_path)
        _feedback_available = True
    except Exception as _e:
        print(f"[feedback] unavailable: {_e}", flush=True)
        _feedback_available = False
        _db_path = None

    os.makedirs(output_dir, exist_ok=True)
    os.makedirs(cache_dir, exist_ok=True)
    pdfs = sorted(os.path.join(input_dir,f) for f in os.listdir(input_dir) if f.lower().endswith(".pdf"))
    if not pdfs: log(f"No PDFs found in {input_dir}"); return
    manual_pages = _load_manual_pages()
    if manual_pages: log(f"Manual overrides loaded for: {', '.join(manual_pages.keys())}")
    log(f"Processing {len(pdfs)} parishes for year {year}...")
    log(f"  Input:  {input_dir}")
    log(f"  Output: {output_dir}")
    log("")

    # Determine output file label: single parish name, or "Multiple Parishes Run N"
    if len(pdfs) == 1:
        label = get_parish_name(pdfs[0])
    else:
        existing_runs = [f for f in os.listdir(output_dir)
                         if re.match(r"Multiple Parishes Run \d+.*CBS", f)]
        run_num = len(existing_runs) + 1
        label = f"Multiple Parishes Run {run_num}"
    log(f"  Label:  {label}")
    log("")

    workbooks = {t: _get_or_create_wb(output_dir, t, year, label) for t in ["cbs","soa","sona","ca"]}
    parish_validations = []

    for i, pdf_path in enumerate(pdfs, 1):
        parish = get_parish_name(pdf_path)
        log(f"[{i:2d}/{len(pdfs)}] {parish}")
        cache_file = os.path.join(cache_dir, f"{parish}_{year}.json")
        if not skip_cache and os.path.exists(cache_file):
            log("  Loading from cache...")
            results = json.loads(open(cache_file).read())
        else:
            log("  Extracting pages from PDF...")
            try:
                manual = _apply_manual(pdf_path, parish, manual_pages)
                if manual is not None:
                    log(f"  [MANUAL] Using hardcoded page overrides for {parish}")
                    texts = manual
                else:
                    texts = get_all_statement_texts(pdf_path)
            except Exception as e:
                log(f"  [ERROR] PDF extraction failed: {e}"); continue
            log("  Calling Claude API (claude-haiku-4-5) — 4 statements...")
            results = extract_all_statements(texts, parish, year, db_path=_db_path, log_fn=log)
            open(cache_file,"w").write(json.dumps(results, indent=2))
            log(f"  Cached to {os.path.basename(cache_file)}")

        # Validate
        if _validate_available:
            try:
                pv = _vp(parish, year, results)
                parish_validations.append(pv)
                sym = {"PASS":"[OK]","WARNING":"[WARN]","REVIEW REQUIRED":"[!]","FAILED":"[FAIL]"}.get(pv.overall_status,"[?]")
                log(f"  {sym} Validation: {pv.overall_status}")
                for stype, sv in pv.statements.items():
                    if sv.status != "PASS":
                        errs = "; ".join(c.name for c in sv.checks if not c.passed)
                        log(f"    {stype.upper()}: {sv.status}" + (f" - {errs}" if errs else ""))
            except Exception as exc:
                log(f"  [ERROR] Validation: {exc}")

        # ── Uncertainty processing ─────────────────────────────────────────
        if _feedback_available:
            for stmt_type in ["cbs", "soa", "sona", "ca"]:
                stmt_data = results.get(stmt_type)
                if not stmt_data:
                    continue
                # 1) Claude-reported uncertainties
                claude_items = stmt_data.pop("_uncertainties", None) or []
                for u in claude_items:
                    try:
                        uid = _fb_save(
                            _db_path, job_id or "cli", parish, year, stmt_type,
                            field_path=u.get("field") or u.get("fund"),
                            reason=u.get("reason", ""),
                            extracted_value=u.get("extracted"),
                            alternative_value=u.get("alternative"),
                            page_number=u.get("page"),
                            text_snippet=u.get("text_snippet"),
                            severity=u.get("severity", "medium"),
                            source="claude",
                        )
                        log({"type": "uncertainty", "data": {
                            "id": uid, "job_id": job_id or "cli",
                            "parish": parish, "year": year,
                            "statement_type": stmt_type,
                            "field_path": u.get("field"),
                            "fund": u.get("fund"),
                            "reason": u.get("reason", ""),
                            "extracted_value": str(u.get("extracted")) if u.get("extracted") is not None else None,
                            "alternative_value": str(u.get("alternative")) if u.get("alternative") is not None else None,
                            "page_number": u.get("page"),
                            "text_snippet": u.get("text_snippet"),
                            "severity": u.get("severity", "medium"),
                            "source": "claude",
                            "resolved": 0,
                        }})
                    except Exception:
                        pass
                # 2) Auto-generate uncertainties from validation failures
                if _validate_available:
                    try:
                        pv_item = next((p for p in parish_validations if p.parish == parish), None)
                        if pv_item and stmt_type in pv_item.statements:
                            sv = pv_item.statements[stmt_type]
                            for chk in sv.checks:
                                if not chk.passed:
                                    uid = _fb_save(
                                        _db_path, job_id or "cli", parish, year, stmt_type,
                                        field_path=chk.name,
                                        reason=(f"{chk.lhs_label} ({chk.lhs_value:,}) ≠ "
                                                f"{chk.rhs_label} ({chk.rhs_value:,}), "
                                                f"diff={chk.difference:,}"
                                                if chk.lhs_value is not None else chk.name),
                                        extracted_value=chk.lhs_value,
                                        alternative_value=chk.rhs_value,
                                        severity="high" if abs(chk.difference or 0) > 1000 else "medium",
                                        source="validation",
                                    )
                                    log({"type": "uncertainty", "data": {
                                        "id": uid, "job_id": job_id or "cli",
                                        "parish": parish, "year": year,
                                        "statement_type": stmt_type,
                                        "field_path": chk.name,
                                        "fund": None,
                                        "reason": (f"{chk.lhs_label} ≠ {chk.rhs_label} "
                                                   f"(diff {chk.difference:,})"
                                                   if chk.lhs_value is not None else chk.name),
                                        "extracted_value": str(chk.lhs_value) if chk.lhs_value is not None else None,
                                        "alternative_value": str(chk.rhs_value) if chk.rhs_value is not None else None,
                                        "page_number": None,
                                        "text_snippet": None,
                                        "severity": "high" if abs(chk.difference or 0) > 1000 else "medium",
                                        "source": "validation",
                                        "resolved": 0,
                                    }})
                    except Exception:
                        pass

        log("  Writing to Excel...")
        for stmt_type in ["cbs","soa","sona","ca"]:
            data = results.get(stmt_type)
            if data:
                try: _add_parish_to_wb(workbooks[stmt_type][0], stmt_type, data, year)
                except Exception as e: log(f"    [ERROR] {stmt_type}: {e}")
            else:
                log(f"    [SKIP] No data for {stmt_type}")
        log("")

    # Add Quality Report tabs before saving
    if _validate_available and parish_validations:
        for stmt_type,(wb,path) in workbooks.items():
            try: _wqrx(wb, parish_validations, manual_pages, year)
            except Exception: pass

    log("Saving Excel files...")
    saved_files = []
    for stmt_type,(wb,path) in workbooks.items():
        try:
            for special in ["cross sectional","governmental cross sectional","quality report"]:
                matches = [s for s in wb.sheetnames if s.lower() == special]
                for m in matches: wb.move_sheet(m, offset=len(wb.sheetnames))
            wb.save(path); saved_files.append(os.path.basename(path))
            log(f"  Saved: {os.path.basename(path)}")
        except Exception as e: log(f"  [ERROR] Save failed for {stmt_type}: {e}")

    log("Building parish workbooks...")
    parish_files = []
    pv_index = {pv.parish: pv for pv in parish_validations} if parish_validations else {}
    for pdf_path in pdfs:
        parish = get_parish_name(pdf_path)
        cache_file = os.path.join(cache_dir, f"{parish}_{year}.json")
        if not os.path.exists(cache_file): continue
        results = json.loads(open(cache_file).read())
        fname = _write_parish_combined(output_dir, parish, results, year,
                                       parish_validation=pv_index.get(parish) if _validate_available else None,
                                       manual_pages=manual_pages)
        if fname: parish_files.append(fname); log(f"  Saved: {fname}")

    # Quality report CSV
    if _validate_available and parish_validations:
        try:
            csv_path = _wqrc(parish_validations, output_dir, manual_pages, year)
            log(f"  Saved: {os.path.basename(csv_path)}")
            statuses = [pv.overall_status for pv in parish_validations]
            log(f"  Summary: {statuses.count('PASS')} PASS | {statuses.count('WARNING')} WARN | "
                f"{statuses.count('REVIEW REQUIRED')} REVIEW | {statuses.count('FAILED')} FAILED")
        except Exception as exc:
            log(f"  [ERROR] Quality report CSV: {exc}")

    log(""); log("Done.")
    return {"statement_files": saved_files, "parish_files": parish_files}

# ══════════════════════════════════════════════════════════════════════════════
# WEB UI (embedded HTML)
# ══════════════════════════════════════════════════════════════════════════════

_HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8" /><meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>Parish Audit Pipeline</title>
  <style>
    :root{--bg:#ffffff;--surf:#f6f8fa;--surf2:#eaeef2;--border:#d0d7de;--text:#1f2328;--dim:#57606a;--dimmer:#8c959f;--purple:#6f42c1;--blue:#6f42c1;--green:#1a7f37;--orange:#9a6700;--red:#cf222e;--cyan:#0550ae;--kw:#cf222e;--str:#0a3069;--cm:#8c959f;--fn:#6f42c1;--nm:#0550ae;--sp:#953800}
    *,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
    body{background:var(--bg);color:var(--text);font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",system-ui,sans-serif;font-size:14px;line-height:1.6;min-height:100vh;padding-bottom:80px}
    .topbar{background:#6f42c1;border-bottom:none;padding:0 28px;display:flex;align-items:center;justify-content:space-between;position:sticky;top:0;z-index:10;box-shadow:0 1px 3px rgba(0,0,0,.15)}
    .topbar-title{font-size:13px;font-weight:700;color:#ffffff;padding:15px 0;letter-spacing:-.1px}
    .tabs{display:flex}
    .tab-btn{background:none;border:none;border-bottom:2px solid transparent;padding:17px 18px 15px;font-size:13px;font-weight:500;color:rgba(255,255,255,.7);cursor:pointer;transition:color .15s,border-color .15s}
    .tab-btn:hover{color:#ffffff}.tab-btn.active{color:#ffffff;border-bottom-color:#ffffff;font-weight:600}
    .panel{display:none}.panel.active{display:block}
    .wrap{max-width:780px;margin:0 auto;padding:36px 28px 0}.wrap-wide{max-width:920px;margin:0 auto;padding:36px 28px 0}
    .page-title{margin-bottom:28px}.page-title h2{font-size:20px;font-weight:700;letter-spacing:-.3px}.page-title p{margin-top:5px;font-size:13px;color:var(--dim)}
    .upload-zone{background:var(--surf);border:1px dashed var(--border);border-radius:8px;padding:48px 24px;text-align:center;cursor:pointer;transition:border-color .18s,background .18s;user-select:none}
    .upload-zone:hover,.upload-zone.drag-over{border-color:var(--purple);background:rgba(111,66,193,.04)}
    .uz-icon{display:block;margin:0 auto 14px;color:var(--dimmer)}
    .uz-main{font-size:13.5px;color:var(--dim)}.uz-browse{color:var(--purple);font-weight:600;cursor:pointer;text-decoration:underline;text-underline-offset:2px}
    .uz-hint{margin-top:6px;font-size:11.5px;color:var(--dimmer)}
    .file-badge{display:inline-block;margin-top:12px;padding:4px 12px;background:rgba(111,66,193,.1);border:1px solid rgba(111,66,193,.3);color:var(--purple);border-radius:20px;font-size:11px;font-weight:600}
    .controls{display:flex;align-items:center;gap:10px;margin-top:12px}
    .year-field{display:flex;align-items:center;gap:8px;background:var(--surf);border:1px solid var(--border);border-radius:7px;padding:0 14px;height:42px;font-size:13px;color:var(--dim);transition:border-color .15s}
    .year-field.unset{border-color:var(--orange);background:rgba(210,100,0,.06)}
    @keyframes shake{0%,100%{transform:translateX(0)}20%{transform:translateX(-6px)}40%{transform:translateX(6px)}60%{transform:translateX(-4px)}80%{transform:translateX(4px)}}
    #filename-warnings{margin-top:10px;font-size:11.5px;display:none}
    .fw-row{display:flex;align-items:flex-start;gap:7px;padding:6px 10px;border-radius:5px;margin-bottom:5px;line-height:1.45}
    .fw-row.bad{background:rgba(248,81,73,.08);border:1px solid rgba(248,81,73,.3)}
    .fw-row.warn{background:rgba(210,100,0,.07);border:1px solid rgba(210,100,0,.28)}
    .fw-icon{flex-shrink:0;font-size:13px;margin-top:1px}
    .fw-suggest{color:var(--green);font-weight:700}
    .year-field select{border:none;outline:none;font-size:13px;font-weight:600;color:var(--text);background:transparent;cursor:pointer;appearance:none;-webkit-appearance:none;padding-right:4px}
    .year-field select.placeholder{color:var(--dimmer);font-weight:400}
    .run-btn{flex:1;height:42px;background:#6f42c1;color:#ffffff;border:none;border-radius:7px;font-size:13px;font-weight:700;cursor:pointer;transition:background .18s;letter-spacing:.2px}
    .run-btn:hover:not(:disabled){background:#5a35a0}.run-btn:disabled{opacity:.35;cursor:default}
    .card{background:var(--surf);border:1px solid var(--border);border-radius:8px;margin-top:18px;overflow:hidden}
    .card-hdr{padding:10px 16px;border-bottom:1px solid var(--border);display:flex;align-items:center;gap:8px;background:var(--surf2)}
    .card-hdr h3{font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.7px;color:var(--dim)}
    .status-dot{width:8px;height:8px;border-radius:50%;background:var(--dimmer);flex-shrink:0;transition:background .3s}
    .status-dot.running{background:var(--orange);animation:pulse 1.2s ease-in-out infinite}
    .status-dot.done{background:var(--green);animation:none}
    .status-dot.error{background:var(--red);animation:none}
    .status-dot.lost{background:#f0c040;animation:pulse .8s ease-in-out infinite}
    #status-label{font-size:10.5px;font-weight:700;letter-spacing:.5px;text-transform:uppercase;transition:color .3s}
    #status-label.running{color:var(--orange)}
    #status-label.done{color:var(--green)}
    #status-label.error{color:var(--red)}
    #status-label.lost{color:#d4a800}
    @keyframes pulse{0%,100%{opacity:1}50%{opacity:.25}}
    .log{font-family:"Cascadia Code","Fira Code","Consolas","Menlo",monospace;font-size:12px;line-height:1.7;background:#f6f8fa;color:var(--dim);padding:14px 18px;height:300px;overflow-y:auto;white-space:pre-wrap;word-break:break-word;border-top:1px solid var(--border)}
    .line-ok{color:var(--green)}.line-skip{color:var(--orange)}.line-error{color:var(--red)}.line-header{color:var(--purple);font-weight:600}.line-manual{color:var(--purple)}
    .dl-section{padding:14px 14px 4px}.dl-section+.dl-section{border-top:1px solid var(--border);padding-top:14px}
    .dl-section-label{font-size:10px;font-weight:700;letter-spacing:.9px;text-transform:uppercase;color:var(--dimmer);margin-bottom:10px}
    .dl-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(190px,1fr));gap:8px;margin-bottom:10px}
    .dl-btn{display:flex;align-items:center;gap:10px;padding:10px 12px;border:1px solid var(--border);border-radius:7px;text-decoration:none;color:var(--dim);font-size:12px;font-weight:500;transition:border-color .18s,color .18s,background .18s}
    .dl-btn:hover{border-color:var(--purple);color:var(--purple);background:rgba(111,66,193,.04)}
    .dl-icon{flex-shrink:0;width:28px;height:28px;background:var(--surf2);border:1px solid var(--border);border-radius:5px;display:flex;align-items:center;justify-content:center;font-size:9px;font-weight:700;letter-spacing:.3px}
    .dl-icon.parish{color:var(--purple)}.dl-icon.statement{color:var(--green)}.dl-name{word-break:break-word;line-height:1.3}
    .section{margin-bottom:64px}
    .sec-kicker{font-size:11px;font-weight:700;letter-spacing:1.4px;text-transform:uppercase;color:var(--purple);margin-bottom:5px}
    .sec-title{font-size:19px;font-weight:700;letter-spacing:-.2px;color:var(--text);margin-bottom:8px}
    .sec-body{font-size:13px;color:var(--dim);line-height:1.75;max-width:680px;margin-bottom:22px}
    .flow{display:grid;grid-template-columns:repeat(5,1fr);gap:2px}
    .flow-step{background:var(--surf);border:1px solid var(--border);padding:16px 14px 18px;position:relative}
    .flow-step:first-child{border-radius:8px 0 0 8px}.flow-step:last-child{border-radius:0 8px 8px 0}
    .flow-step::before{content:"";position:absolute;top:0;left:0;right:0;height:2px}
    .s1::before{background:var(--purple)}.s2::before{background:var(--orange)}.s3::before{background:var(--green)}.s4::before{background:#0550ae}.s5::before{background:var(--cyan)}
    .step-n{font-size:10px;font-weight:700;letter-spacing:1px;text-transform:uppercase;margin-bottom:7px}
    .s1 .step-n{color:var(--purple)}.s2 .step-n{color:var(--orange)}.s3 .step-n{color:var(--green)}.s4 .step-n{color:#0550ae}.s5 .step-n{color:var(--cyan)}
    .step-title{font-size:12px;font-weight:700;color:var(--text);margin-bottom:2px}
    .step-lib{font-size:10px;color:var(--dimmer);text-transform:uppercase;letter-spacing:.5px;margin-bottom:9px}
    .step-pts{list-style:none}.step-pts li{font-size:11px;color:var(--dim);line-height:1.55;padding:1px 0 1px 10px;position:relative}.step-pts li::before{content:"–";position:absolute;left:0;color:var(--dimmer)}
    .code-wrap{background:var(--surf);border:1px solid var(--border);border-radius:8px;overflow:hidden;margin:16px 0}
    .code-bar{display:flex;justify-content:space-between;align-items:center;padding:7px 14px;background:var(--surf2);border-bottom:1px solid var(--border)}
    .code-label{font-size:11px;font-weight:600;color:var(--dim)}.code-file{font-family:monospace;font-size:10px;color:var(--dimmer)}
    pre{padding:16px 18px;overflow-x:auto;font-family:"Cascadia Code","Fira Code","Consolas","Menlo",monospace;font-size:12px;line-height:1.7;color:var(--text);tab-size:4}
    .kw{color:var(--kw)}.str{color:var(--str)}.cm{color:var(--cm);font-style:italic}.fn{color:var(--fn)}.nm{color:var(--nm)}.sp{color:var(--sp)}
    .hl{display:block;background:rgba(111,66,193,.06);margin:0 -18px;padding:0 18px;border-left:2px solid var(--purple)}
    .hl-g{display:block;background:rgba(26,127,55,.06);margin:0 -18px;padding:0 18px;border-left:2px solid var(--green)}
    .hl-r{display:block;background:rgba(207,34,46,.06);margin:0 -18px;padding:0 18px;border-left:2px solid var(--red)}
    .stmt-grid{display:grid;grid-template-columns:1fr 1fr;gap:12px}
    .stmt-card{background:var(--surf);border:1px solid var(--border);border-radius:8px;padding:18px;border-left:3px solid}
    .cbs{border-left-color:var(--purple)}.soa{border-left-color:var(--orange)}.sona{border-left-color:var(--green)}.ca{border-left-color:#0550ae}
    .stmt-abbr{font-size:10px;font-weight:700;letter-spacing:1px;text-transform:uppercase;margin-bottom:3px}
    .cbs .stmt-abbr{color:var(--purple)}.soa .stmt-abbr{color:var(--orange)}.sona .stmt-abbr{color:var(--green)}.ca .stmt-abbr{color:#0550ae}
    .stmt-name{font-size:12.5px;font-weight:700;color:var(--text);margin-bottom:6px}.stmt-desc{font-size:11.5px;color:var(--dim);line-height:1.65}
    .stmt-fields{margin-top:10px;display:flex;flex-wrap:wrap;gap:5px}
    .sf{font-size:10.5px;padding:2px 7px;border-radius:4px;background:var(--surf2);border:1px solid var(--border);color:var(--dim)}
    .tbl-wrap{background:var(--surf);border:1px solid var(--border);border-radius:8px;overflow:hidden;margin:16px 0}
    table{width:100%;border-collapse:collapse}
    th{font-size:10px;font-weight:700;letter-spacing:.8px;text-transform:uppercase;color:var(--dim);text-align:left;padding:9px 14px;background:var(--surf2);border-bottom:1px solid var(--border)}
    td{font-size:12.5px;padding:9px 14px;border-bottom:1px solid var(--border);vertical-align:middle}
    tr:last-child td{border-bottom:none}
    .td-orig{font-family:monospace;font-size:12px;color:var(--text)}.td-bad{font-family:monospace;font-size:12px;color:var(--red)}.td-fix{font-family:monospace;font-size:11px;color:var(--green)}.td-parish{font-size:11px;color:var(--dimmer)}
    .td-name{font-weight:700;color:var(--text)}.td-role{font-size:10px;font-weight:600;text-transform:uppercase;letter-spacing:.4px;color:var(--purple)}.td-desc{color:var(--dim);font-size:12.5px}.td-mono{font-family:monospace;font-size:11px;color:var(--dimmer)}
    .callout{background:var(--surf);border:1px solid var(--border);border-left:3px solid var(--orange);border-radius:8px;padding:16px 18px;margin:16px 0;font-size:12.5px;color:var(--dim);line-height:1.7}
    .callout strong{color:var(--text)}
    code{font-family:"Cascadia Code","Fira Code","Consolas",monospace;font-size:11px;background:var(--surf2);border:1px solid var(--border);border-radius:3px;padding:1px 5px;color:var(--text)}
    /* rv-btn kept for PDF page modal buttons */
    .rv-btn{border:none;border-radius:6px;padding:5px 13px;font-size:12px;font-weight:600;cursor:pointer;transition:opacity .15s}
    .rv-btn:hover{opacity:.8}
    .btn-confirm{background:rgba(26,127,55,.12);color:var(--green);border:1px solid rgba(26,127,55,.25)}
    .btn-page{background:rgba(111,66,193,.1);color:var(--purple);border:1px solid rgba(111,66,193,.25)}
    .btn-correct{background:rgba(154,103,0,.1);color:var(--orange);border:1px solid rgba(154,103,0,.2)}
    .btn-skip{background:var(--surf2);color:var(--dimmer);border:1px solid var(--border)}
    .badge-stmt{font-size:10px;font-weight:700;letter-spacing:.6px;text-transform:uppercase;padding:2px 7px;border-radius:4px;background:var(--surf2);border:1px solid var(--border);color:var(--dim)}
    .badge-sev{font-size:10px;font-weight:700;letter-spacing:.4px;padding:2px 7px;border-radius:4px}
    .sev-high{background:rgba(207,34,46,.1);color:var(--red);border:1px solid rgba(207,34,46,.25)}
    .sev-medium{background:rgba(154,103,0,.1);color:var(--orange);border:1px solid rgba(154,103,0,.25)}
    .sev-low{background:rgba(26,127,55,.1);color:var(--green);border:1px solid rgba(26,127,55,.2)}
    /* Chat modal */
    .chat-box{display:flex;flex-direction:column;height:420px}
    .chat-ctx{padding:10px 14px;background:var(--surf2);border-bottom:1px solid var(--border);font-size:11.5px;color:var(--dim);line-height:1.6;flex-shrink:0}
    .chat-ctx strong{color:var(--text)}
    .chat-msgs{flex:1;overflow-y:auto;padding:14px;display:flex;flex-direction:column;gap:10px;background:#ffffff}
    .chat-bubble{max-width:88%;padding:9px 13px;border-radius:10px;font-size:12.5px;line-height:1.65;white-space:pre-wrap;word-break:break-word}
    .chat-bubble.user{background:rgba(111,66,193,.12);border:1px solid rgba(111,66,193,.2);color:var(--text);align-self:flex-end;border-bottom-right-radius:3px}
    .chat-bubble.assistant{background:var(--surf);border:1px solid var(--border);color:var(--text);align-self:flex-start;border-bottom-left-radius:3px}
    .chat-bubble.thinking{color:var(--dimmer);font-style:italic}
    .chat-input-row{display:flex;gap:8px;padding:10px 14px;border-top:1px solid var(--border);flex-shrink:0;background:#ffffff}
    .chat-input{flex:1;background:#ffffff;border:1px solid var(--border);border-radius:7px;padding:7px 12px;font-size:13px;color:var(--text);outline:none;font-family:inherit;resize:none;height:38px}
    .chat-input:focus{border-color:var(--purple);box-shadow:0 0 0 3px rgba(111,66,193,.1)}
    .chat-send{background:#6f42c1;color:#ffffff;border:none;border-radius:7px;padding:0 16px;font-size:13px;font-weight:700;cursor:pointer;white-space:nowrap}
    .chat-send:disabled{opacity:.4;cursor:default}
    /* Page image modal */
    .modal-overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,.5);z-index:100;align-items:center;justify-content:center}
    .modal-overlay.open{display:flex}
    .modal-box{background:#ffffff;border:1px solid var(--border);border-radius:10px;max-width:90vw;max-height:90vh;overflow:hidden;display:flex;flex-direction:column;box-shadow:0 8px 32px rgba(0,0,0,.12)}
    .modal-hdr{display:flex;justify-content:space-between;align-items:center;padding:12px 16px;border-bottom:1px solid var(--border);background:var(--surf)}
    .modal-hdr h3{font-size:13px;font-weight:600;color:var(--text)}
    .modal-close{background:none;border:none;color:var(--dim);cursor:pointer;font-size:18px;line-height:1;padding:2px 6px}
    .modal-close:hover{color:var(--text)}
    .modal-body{overflow:auto;padding:12px}
    .modal-body img{max-width:100%;display:block;border-radius:4px}
    .modal-loading{padding:40px;text-align:center;color:var(--dimmer);font-size:13px}
    /* Library tab */
    /* ── Library ── */
    .lib-layout{display:flex;gap:0;height:calc(100vh - 52px);overflow:hidden}
    .lib-sidebar{width:240px;min-width:200px;border-right:1px solid var(--border);overflow-y:auto;flex-shrink:0;background:var(--surf)}
    .lib-main{flex:1;display:flex;flex-direction:column;overflow:hidden;background:#fff}
    /* sidebar year accordion */
    .lib-year-item{}
    .lib-year-btn{display:flex;align-items:center;gap:8px;width:100%;padding:10px 14px;cursor:pointer;border:none;border-bottom:1px solid var(--border);background:var(--surf);font-size:13px;font-weight:700;color:var(--text);text-align:left;transition:background .1s}
    .lib-year-btn:hover{background:var(--surf2)}
    .lib-year-btn.active{background:rgba(111,66,193,.1);color:var(--purple)}
    .lib-year-caret{font-size:10px;margin-left:auto;transition:transform .2s;color:var(--dimmer)}
    .lib-year-count{font-size:10.5px;font-weight:400;color:var(--dimmer);margin-left:4px}
    .lib-parishes{display:none}
    .lib-parishes.open{display:block}
    .lib-parish-row{display:flex;align-items:center;gap:6px;padding:6px 14px 6px 24px;cursor:pointer;border-bottom:1px solid rgba(0,0,0,.03);font-size:12px;color:var(--text);transition:background .1s}
    .lib-parish-row:hover{background:var(--surf2)}
    .lib-parish-row.active{background:rgba(111,66,193,.1);color:var(--purple);font-weight:600}
    .lib-badges{display:flex;gap:2px;margin-left:auto;flex-shrink:0}
    .lib-badge{font-size:8.5px;font-weight:700;padding:1px 3px;border-radius:2px;background:var(--surf2);color:var(--dim)}
    .lib-badge.ok{background:rgba(26,127,55,.12);color:#1a7f37}
    .lib-badge.pdf{background:rgba(111,66,193,.12);color:var(--purple)}
    /* search */
    .lib-search-input{width:100%;box-sizing:border-box;padding:7px 12px;border:none;border-bottom:1px solid var(--border);font-size:12px;color:var(--text);background:var(--surf);outline:none}
    .lib-search-input:focus{background:#fff}
    /* toolbar */
    .lib-toolbar{display:flex;align-items:center;gap:10px;padding:9px 16px;border-bottom:1px solid var(--border);background:var(--surf);flex-shrink:0}
    .lib-toolbar h3{font-size:13px;font-weight:600;color:var(--text);margin:0}
    .lib-back-btn{background:none;border:1px solid var(--border);border-radius:5px;padding:3px 10px;cursor:pointer;font-size:12px;color:var(--dim)}
    .lib-back-btn:hover{background:var(--surf2)}
    .lib-page-ctrl{display:flex;align-items:center;gap:6px;margin-left:auto}
    .lib-page-btn{background:none;border:1px solid var(--border);border-radius:5px;padding:3px 10px;cursor:pointer;font-size:13px;color:var(--text)}
    .lib-page-btn:hover{background:var(--surf2)}
    .lib-page-btn:disabled{opacity:.35;cursor:default}
    .lib-page-label{font-size:12px;color:var(--dim);min-width:70px;text-align:center}
    /* content */
    .lib-content{flex:1;display:flex;overflow:hidden}
    .lib-pdf-pane{flex:1;overflow-y:auto;display:flex;align-items:flex-start;justify-content:center;padding:16px;background:#e8e8e8}
    .lib-pdf-pane img{max-width:100%;border-radius:4px;box-shadow:0 2px 12px rgba(0,0,0,.2)}
    .lib-data-pane{width:270px;border-left:1px solid var(--border);overflow-y:auto;background:#fff;flex-shrink:0}
    .lib-data-section{border-bottom:1px solid var(--border);padding:10px 14px}
    .lib-data-section h4{font-size:10.5px;font-weight:700;color:var(--dimmer);text-transform:uppercase;letter-spacing:.06em;margin:0 0 7px}
    .lib-data-row{display:flex;justify-content:space-between;align-items:baseline;gap:6px;padding:3px 0;font-size:11.5px}
    .lib-data-label{color:var(--dim);flex:1;line-height:1.3}
    .lib-data-val{color:var(--text);font-weight:600;font-family:monospace;font-size:11px;white-space:nowrap}
    .lib-empty{flex:1;display:flex;flex-direction:column;align-items:center;justify-content:center;color:var(--dimmer);gap:8px;font-size:13px}
    /* dashboard */
    .dash-scroll{flex:1;overflow-y:auto;padding:24px}
    .dash-kpi-row{display:flex;gap:14px;flex-wrap:wrap;margin-bottom:20px}
    .dash-kpi{flex:1;min-width:130px;background:var(--surf);border:1px solid var(--border);border-radius:8px;padding:14px 16px}
    .dash-kpi-val{font-size:22px;font-weight:700;color:var(--text);margin-bottom:2px}
    .dash-kpi-label{font-size:11.5px;color:var(--dimmer)}
    .dash-table{width:100%;border-collapse:collapse;font-size:12px}
    .dash-table th{text-align:left;padding:6px 10px;color:var(--dim);font-weight:600;border-bottom:2px solid var(--border);white-space:nowrap}
    .dash-table th.num{text-align:right}
    .dash-table td{padding:5px 10px;border-bottom:1px solid var(--border);color:var(--text);vertical-align:middle}
    .dash-table td.num{text-align:right;font-family:monospace;font-size:11.5px}
    .dash-table tr:hover td{background:var(--surf)}
    .dash-table tr.clickable{cursor:pointer}
    .dash-table tr.dup-row td{opacity:.6;font-style:italic}
    .dash-stmt-dot{display:inline-block;width:7px;height:7px;border-radius:50%;background:var(--border);margin-right:1px}
    .dash-stmt-dot.ok{background:#1a7f37}
    .dash-section-hdr{font-size:12px;font-weight:700;color:var(--text);margin:20px 0 10px;padding-bottom:6px;border-bottom:1px solid var(--border)}
    /* ── Spreadsheet split view ── */
    .lib-view-tabs{display:flex;border-bottom:1px solid var(--border);background:var(--surf);flex-shrink:0}
    .lib-view-tab{background:none;border:none;border-bottom:2px solid transparent;padding:7px 16px;font-size:11.5px;font-weight:500;color:var(--dim);cursor:pointer;transition:color .15s}
    .lib-view-tab.active{color:var(--purple);border-bottom-color:var(--purple);font-weight:600}
    .xl-split{display:flex;flex:1;overflow:hidden}
    .xl-side{flex:1;display:flex;flex-direction:column;overflow:hidden;min-width:0}
    .xl-side+.xl-side{border-left:3px solid var(--border)}
    .xl-side-hdr{padding:6px 12px;font-size:10.5px;font-weight:700;letter-spacing:.5px;text-transform:uppercase;background:var(--surf2);border-bottom:1px solid var(--border);flex-shrink:0}
    .xl-side-hdr.truth{border-top:2px solid var(--green);color:var(--green)}
    .xl-side-hdr.pipeline{border-top:2px solid var(--purple);color:var(--purple)}
    .xl-scroll{overflow:auto;flex:1}
    .xl-tbl{border-collapse:collapse;font-size:11px}
    .xl-tbl td{border:1px solid #e8e8e8;padding:2px 8px;white-space:nowrap;max-width:240px;overflow:hidden;text-overflow:ellipsis;font-family:monospace;font-size:10.5px;color:var(--text)}
    .xl-tbl td.xl-rn{background:var(--surf2);color:var(--dimmer);font-family:inherit;font-size:10px;font-weight:600;position:sticky;left:0;z-index:1;min-width:30px;text-align:center}
    .xl-tbl tr.xl-hdr td{background:var(--surf2);font-weight:700;font-family:inherit;position:sticky;top:0;z-index:2;text-align:center}
    .xl-tbl tr.xl-hdr td.xl-rn{z-index:3}
    .xl-diff{background:rgba(207,34,46,.12)!important;color:var(--red)!important}
    .xl-target{background:rgba(111,66,193,.25)!important;outline:2px solid var(--purple);outline-offset:-2px;font-weight:700!important}
    .xl-missing{color:var(--dimmer);font-style:italic;text-align:center;padding:40px 20px;font-size:13px}
    /* ── All-flags sidebar ── */
    .lib-allflags-btn{display:flex;align-items:center;gap:8px;width:100%;padding:10px 14px;cursor:pointer;border:none;border-bottom:1px solid var(--border);background:var(--surf);font-size:12px;font-weight:600;color:var(--dim);text-align:left;transition:background .1s}
    .lib-allflags-btn:hover{background:var(--surf2)}.lib-allflags-btn.active{background:rgba(111,66,193,.1);color:var(--purple)}
  </style>
</head>
<body>
<div class="topbar">
  <span class="topbar-title">Louisiana Parish Audit Pipeline</span>
  <nav class="tabs">
    <button class="tab-btn active" onclick="switchTab('run',this)">Run Pipeline</button>
    <button class="tab-btn" onclick="switchTab('library',this)" id="tab-library-btn">Training <span id="library-badge" style="display:none;background:rgba(255,255,255,.25);color:#fff;font-size:10px;font-weight:700;padding:1px 6px;border-radius:10px;margin-left:4px">0</span></button>
    <button class="tab-btn" onclick="switchTab('about',this)">How It Works</button>
  </nav>
  <button onclick="openSettings()" id="settings-btn" title="Settings" style="margin-left:12px;background:none;border:none;cursor:pointer;color:rgba(255,255,255,.75);font-size:18px;padding:4px 6px;border-radius:4px;line-height:1" onmouseover="this.style.color='#fff'" onmouseout="this.style.color='rgba(255,255,255,.75)'">&#9881;</button>
  <span id="key-dot" title="No API key configured" style="width:8px;height:8px;border-radius:50%;background:#f85149;display:inline-block;margin-left:4px;margin-right:4px;vertical-align:middle"></span>
</div>

<!-- Settings modal -->
<div id="settings-modal" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.45);z-index:9999;display:none;align-items:center;justify-content:center">
  <div style="background:#fff;border-radius:10px;padding:28px 32px;width:480px;max-width:92vw;box-shadow:0 8px 40px rgba(0,0,0,.25)">
    <h3 style="margin:0 0 6px;font-size:16px;color:#1f2328">Settings</h3>
    <p style="margin:0 0 18px;font-size:13px;color:#57606a">Your API key is saved locally in <code>config.json</code> and never sent anywhere except Anthropic.</p>
    <label style="font-size:13px;font-weight:600;color:#1f2328;display:block;margin-bottom:6px">Anthropic API Key</label>
    <input id="settings-key-input" type="password" placeholder="sk-ant-api03-..." autocomplete="off"
      style="width:100%;box-sizing:border-box;padding:8px 10px;border:1px solid #d0d7de;border-radius:6px;font-size:13px;font-family:monospace;margin-bottom:4px">
    <div id="settings-key-current" style="font-size:11.5px;color:#57606a;margin-bottom:14px"></div>
    <div style="display:flex;gap:8px;justify-content:flex-end">
      <button onclick="closeSettings()" style="padding:6px 16px;border:1px solid #d0d7de;background:#fff;border-radius:6px;cursor:pointer;font-size:13px">Cancel</button>
      <button onclick="saveSettings()" style="padding:6px 16px;background:#6f42c1;color:#fff;border:none;border-radius:6px;cursor:pointer;font-size:13px;font-weight:600">Save Key</button>
    </div>
    <div id="settings-msg" style="margin-top:10px;font-size:12px;min-height:16px"></div>
    <hr style="margin:22px 0 18px;border:none;border-top:1px solid #e0e0e0">
    <h4 style="margin:0 0 5px;font-size:14px;font-weight:700;color:#1f2328">Import Run</h4>
    <p style="margin:0 0 14px;font-size:12px;color:#57606a">Select all files from a local run folder — PDFs, output Excels, and JSON cache files together. The server sorts them automatically.</p>
    <div id="import-drop" style="border:1px dashed #d0d7de;border-radius:7px;padding:20px 16px;text-align:center;cursor:pointer;background:#f6f8fa;transition:border-color .18s"
      onclick="document.getElementById('import-file-input').click()"
      ondragover="event.preventDefault();this.style.borderColor='#6f42c1';this.style.background='rgba(111,66,193,.04)'"
      ondragleave="this.style.borderColor='#d0d7de';this.style.background='#f6f8fa'"
      ondrop="event.preventDefault();this.style.borderColor='#d0d7de';this.style.background='#f6f8fa';_importFilesChosen(event.dataTransfer.files)">
      <div style="font-size:13px;color:#57606a">Drop <b>all files from a run folder</b> here, or <span style="color:#6f42c1;font-weight:600;text-decoration:underline">browse</span></div>
      <div style="font-size:11px;color:#8c959f;margin-top:4px">PDFs + Excels + JSON cache — mixed is fine</div>
      <div id="import-preview" style="margin-top:10px;font-size:11.5px;color:#57606a;display:none"></div>
    </div>
    <input type="file" id="import-file-input" multiple accept=".pdf,.xlsx,.json" style="display:none" onchange="_importFilesChosen(this.files)"/>
    <button onclick="importRunFiles()" style="margin-top:10px;width:100%;padding:7px 16px;background:#6f42c1;color:#fff;border:none;border-radius:6px;cursor:pointer;font-size:13px;font-weight:600">Import to Training</button>
    <div id="import-msg" style="margin-top:8px;font-size:12px;min-height:16px"></div>
  </div>
</div>

<div id="panel-run" class="panel active"><div class="wrap">
  <div id="no-key-banner" style="display:none;margin-bottom:16px;padding:12px 16px;background:#fff3cd;border:1px solid #ffc107;border-radius:6px;font-size:13px;color:#856404">
    <b>&#9888; No API key configured.</b> Click the <b>&#9881; gear icon</b> in the top-right to add your Anthropic key before running the pipeline.
  </div>
  <div class="page-title">
    <h2>Run the Pipeline</h2>
    <p>Upload parish audit PDFs — the pipeline detects statements, calls Claude, and returns clean Excel files.</p>
  </div>
  <div class="upload-zone" id="drop-zone">
    <svg class="uz-icon" width="32" height="32" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round">
      <path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="17 8 12 3 7 8"/><line x1="12" y1="3" x2="12" y2="15"/>
    </svg>
    <p class="uz-main">Drop PDF files here, or <span class="uz-browse" id="browse-link">browse</span></p>
    <p class="uz-hint">One PDF per parish — filename becomes the parish name (e.g. Orleans.pdf)</p>
    <span class="file-badge" id="file-badge" style="display:none"></span>
  </div>
  <input type="file" id="file-input" multiple accept=".pdf" style="display:none"/>
  <div id="filename-warnings"></div>
  <div class="controls">
    <label class="year-field unset" id="year-field-wrap">&#128197;
      <select id="year-input" class="placeholder" onchange="onYearChange(this)">
        <option value="" disabled selected>Select fiscal year</option>
        <option value="2010">2010</option>
        <option value="2011">2011</option>
        <option value="2012">2012</option>
        <option value="2013">2013</option>
        <option value="2014">2014</option>
        <option value="2015">2015</option>
        <option value="2016">2016</option>
        <option value="2017">2017</option>
        <option value="2018">2018</option>
        <option value="2019">2019</option>
        <option value="2020">2020</option>
        <option value="2021">2021</option>
        <option value="2022">2022</option>
        <option value="2023">2023</option>
        <option value="2024">2024</option>
        <option value="2025">2025</option>
      </select>
    </label>
    <button class="run-btn" id="run-btn" onclick="runPipeline()">Run Pipeline</button>
  </div>
  <div id="reconnect-banner" style="display:none;margin-bottom:12px;padding:10px 14px;background:#fff8e1;border:1px solid #f0c040;border-radius:6px;font-size:13px;color:#7a5700;align-items:center;gap:10px">
    <span>&#9889;</span>
    <span id="reconnect-msg">A pipeline run is still in progress.</span>
    <button onclick="reconnectJob()" style="margin-left:auto;padding:4px 12px;background:#6f42c1;color:#fff;border:none;border-radius:4px;cursor:pointer;font-size:12px">Reconnect</button>
    <button onclick="document.getElementById('reconnect-banner').style.display='none'" style="padding:4px 8px;background:none;border:1px solid #c0a000;border-radius:4px;cursor:pointer;font-size:12px">Dismiss</button>
  </div>
  <div class="card" id="log-card" style="display:none">
    <div class="card-hdr" style="display:flex;align-items:center;gap:7px">
      <div class="status-dot" id="status-dot"></div>
      <h3 style="margin:0">Progress</h3>
      <span id="status-label"></span>
      <span id="status-detail" style="font-size:10.5px;color:var(--dimmer);font-weight:400;margin-left:2px"></span>
    </div>
    <div class="log" id="log"></div>
  </div>
  <div class="card" id="downloads-card" style="display:none">
    <div class="card-hdr"><h3>Downloads</h3></div>
    <div id="dl-body"></div>
  </div>
</div></div>


<!-- Page image modal -->
<div class="modal-overlay" id="page-modal">
  <div class="modal-box" style="width:860px;max-width:96vw">
    <div class="modal-hdr">
      <h3 id="modal-title">PDF Page</h3>
      <div id="pdf-nav" style="display:none;align-items:center;gap:6px">
        <button class="rv-btn btn-page" style="padding:3px 10px;font-size:13px" onclick="pdfNavStep(-1)">&#8592;</button>
        <span id="pdf-nav-label" style="font-size:12px;color:var(--dim);min-width:70px;text-align:center"></span>
        <button class="rv-btn btn-page" style="padding:3px 10px;font-size:13px" onclick="pdfNavStep(1)">&#8594;</button>
      </div>
      <button class="modal-close" onclick="closeModal()">&#x2715;</button>
    </div>
    <div class="modal-body" id="modal-body">
      <p class="modal-loading">Loading...</p>
    </div>
  </div>
</div>

<!-- Chat modal -->
<div class="modal-overlay" id="chat-modal">
  <div class="modal-box" style="width:620px;max-width:92vw">
    <div class="modal-hdr">
      <h3 id="chat-modal-title">Ask about this flag</h3>
      <button class="modal-close" onclick="closeChatModal()">&#x2715;</button>
    </div>
    <div class="chat-box">
      <div class="chat-ctx" id="chat-ctx"></div>
      <div class="chat-msgs" id="chat-msgs"></div>
      <div class="chat-input-row">
        <textarea class="chat-input" id="chat-input" placeholder="Ask anything about this flag…" rows="1"
          onkeydown="if(event.key==='Enter'&&!event.shiftKey){event.preventDefault();sendChat();}"></textarea>
        <button class="chat-send" id="chat-send" onclick="sendChat()">Send</button>
      </div>
    </div>
  </div>
</div>


<input type="file" id="train-all-files" multiple accept=".pdf,.xlsx" style="display:none" onchange="libTrainFilesChosen(this.files)"/>
<div id="panel-library" class="panel">
  <div class="lib-layout">
    <!-- Sidebar: year accordion -->
    <div class="lib-sidebar">
      <input class="lib-search-input" id="lib-search" placeholder="Search parishes…" oninput="filterLibrary(this.value)">
      <button class="lib-allflags-btn" id="lib-allflags-btn" onclick="libShowAllFlags()">&#9873; All Flags <span id="allflags-count" style="margin-left:auto;font-size:10.5px;color:var(--dimmer)"></span></button>
      <div id="lib-year-list"><div style="padding:20px 14px;font-size:12px;color:var(--dimmer)">Loading…</div></div>
    </div>
    <!-- Main area -->
    <div class="lib-main">
      <!-- Empty state -->
      <div class="lib-empty" id="lib-empty">
        <svg width="40" height="40" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.2" stroke-linecap="round" stroke-linejoin="round" style="color:var(--border)"><path d="M4 19.5A2.5 2.5 0 0 1 6.5 17H20"/><path d="M6.5 2H20v20H6.5A2.5 2.5 0 0 1 4 19.5v-15A2.5 2.5 0 0 1 6.5 2z"/></svg>
        <span>Select a year or parish from the sidebar</span>
      </div>
      <!-- Dashboard view (year selected) -->
      <div id="lib-dashboard" style="display:none;flex:1;flex-direction:column;overflow:hidden">
        <div class="lib-toolbar">
          <h3 id="lib-dash-title">—</h3>
        </div>
        <div class="dash-scroll" id="lib-dash-body"></div>
      </div>
      <!-- Parish viewer -->
      <div id="lib-viewer" style="display:none;flex:1;flex-direction:column;overflow:hidden">
        <div class="lib-toolbar">
          <button class="lib-back-btn" onclick="libBackToDash()">&#8592; Year</button>
          <h3 id="lib-viewer-title">—</h3>
          <div class="lib-page-ctrl">
            <button class="lib-page-btn" id="lib-prev" onclick="libPageStep(-1)" disabled>&#8592;</button>
            <span class="lib-page-label" id="lib-page-label">— / —</span>
            <button class="lib-page-btn" id="lib-next" onclick="libPageStep(1)" disabled>&#8594;</button>
          </div>
        </div>
        <!-- view mode tabs -->
        <div class="lib-view-tabs">
          <button class="lib-view-tab active" id="vt-data" onclick="libSwitchView('data')">Data</button>
          <button class="lib-view-tab" id="vt-sheet" onclick="libSwitchView('sheet')">Spreadsheet</button>
        </div>
        <!-- data view -->
        <div id="lib-view-data" class="lib-content">
          <div class="lib-pdf-pane" id="lib-pdf-pane"></div>
          <div class="lib-data-pane" id="lib-data-pane"></div>
        </div>
        <!-- spreadsheet view -->
        <div id="lib-view-sheet" style="display:none;flex:1;overflow:hidden;flex-direction:column">
          <div style="padding:6px 12px;background:var(--surf2);border-bottom:1px solid var(--border);font-size:11.5px;color:var(--dim);flex-shrink:0">
            Statement: <select id="xl-stmt-sel" style="border:none;background:transparent;font-size:11.5px;font-weight:600;color:var(--text);cursor:pointer;outline:none" onchange="libLoadSpreadsheet()">
              <option value="cbs">CBS — Balance Sheet</option>
              <option value="soa">SOA — Activities</option>
              <option value="sona">SONA — Net Position</option>
              <option value="ca">CA — Capital Assets</option>
            </select>
            <span id="xl-status" style="margin-left:12px;color:var(--dimmer)"></span>
          </div>
          <div class="xl-split" style="flex:1;overflow:hidden">
            <div class="xl-side">
              <div class="xl-side-hdr truth">Truth (Coded Excel)</div>
              <div class="xl-scroll" id="xl-truth-scroll"><div id="xl-truth"><p class="xl-missing">Loading…</p></div></div>
            </div>
            <div class="xl-side">
              <div class="xl-side-hdr pipeline">Pipeline Output</div>
              <div class="xl-scroll" id="xl-pipe-scroll"><div id="xl-pipe"><p class="xl-missing">Loading…</p></div></div>
            </div>
          </div>
        </div>
      </div>
      <!-- All flags view -->
      <div id="lib-allflags-view" style="display:none;flex:1;flex-direction:column;overflow:hidden">
        <div class="lib-toolbar"><h3>All Flags</h3>
          <select id="af-year-sel" style="margin-left:8px;border:1px solid var(--border);border-radius:5px;padding:3px 8px;font-size:12px;background:var(--surf)" onchange="libLoadAllFlags()">
            <option value="">All years</option>
          </select>
          <select id="af-stmt-sel" style="margin-left:6px;border:1px solid var(--border);border-radius:5px;padding:3px 8px;font-size:12px;background:var(--surf)" onchange="libLoadAllFlags()">
            <option value="">All statements</option>
            <option value="cbs">CBS</option><option value="soa">SOA</option>
            <option value="sona">SONA</option><option value="ca">CA</option>
          </select>
          <span id="af-count" style="margin-left:auto;font-size:11.5px;color:var(--dimmer)"></span>
        </div>
        <div style="flex:1;overflow-y:auto;padding:16px" id="af-list"></div>
      </div>
    </div>
  </div>
</div>

<div id="panel-about" class="panel"><div class="wrap-wide">
  <div class="section">
    <div class="sec-kicker">Architecture</div><div class="sec-title">Five-stage pipeline</div>
    <p class="sec-body">Each parish PDF passes through five sequential stages. Extraction and detection run locally with no API calls. Claude is invoked once per parish, and results are cached to disk so re-runs skip the API entirely.</p>
    <div class="flow">
      <div class="flow-step s1"><div class="step-n">01 / Upload</div><div class="step-title">PDF Ingest</div><div class="step-lib">FastAPI · multipart</div><ul class="step-pts"><li>Browser POST to <code>/run</code></li><li>Saved to per-job temp dir</li><li>Filename becomes parish name</li></ul></div>
      <div class="flow-step s2"><div class="step-n">02 / Detect</div><div class="step-title">Page Detection</div><div class="step-lib">pdfplumber</div><ul class="step-pts"><li>Every page text extracted</li><li>50+ keyword patterns per type</li><li>±3-page confirmation window</li></ul></div>
      <div class="flow-step s3"><div class="step-n">03 / Assemble</div><div class="step-title">Text Assembly</div><div class="step-lib">extractor</div><ul class="step-pts"><li>Anchor + continuation pages</li><li>In-thousands flag detected</li><li>Manual overrides applied</li></ul></div>
      <div class="flow-step s4"><div class="step-n">04 / Parse</div><div class="step-title">LLM Extraction</div><div class="step-lib">Anthropic Claude</div><ul class="step-pts"><li>Structured JSON per statement</li><li>All line items and totals</li><li>Results cached to disk</li></ul></div>
      <div class="flow-step s5"><div class="step-n">05 / Output</div><div class="step-title">Excel Build</div><div class="step-lib">openpyxl</div><ul class="step-pts"><li>4 workbooks, 1 sheet/parish</li><li>Cross-sectional aggregation</li><li>Streamed to browser via SSE</li></ul></div>
    </div>
  </div>
  <div class="section">
    <div class="sec-kicker">Output</div><div class="sec-title">The four financial statements</div>
    <p class="sec-body">Every Louisiana parish audit contains these four statements. The pipeline produces a separate Excel workbook for each, plus a combined per-parish file.</p>
    <div class="stmt-grid">
      <div class="stmt-card cbs"><div class="stmt-abbr">CBS</div><div class="stmt-name">Combined Balance Sheet</div><div class="stmt-desc">Governmental funds view. Each major fund is a column. Shows assets, liabilities, deferred inflows, and fund balances by spendability.</div><div class="stmt-fields"><span class="sf">Cash &amp; deposits</span><span class="sf">Taxes receivable</span><span class="sf">Accounts payable</span><span class="sf">Fund balances by type</span></div></div>
      <div class="stmt-card soa"><div class="stmt-abbr">SOA</div><div class="stmt-name">Statement of Activities</div><div class="stmt-desc">Government-wide income statement. Each function shows expenses, program revenues, and net cost covered by general revenues.</div><div class="stmt-fields"><span class="sf">Expenses by function</span><span class="sf">Charges for services</span><span class="sf">Property taxes</span><span class="sf">Net position change</span></div></div>
      <div class="stmt-card sona"><div class="stmt-abbr">SONA</div><div class="stmt-name">Statement of Net Position</div><div class="stmt-desc">Government-wide balance sheet. Includes capital assets and long-term liabilities unlike the CBS.</div><div class="stmt-fields"><span class="sf">Capital assets net</span><span class="sf">Long-term liabilities</span><span class="sf">Deferred outflows</span><span class="sf">Net position components</span></div></div>
      <div class="stmt-card ca"><div class="stmt-abbr">CA</div><div class="stmt-name">Capital Assets Schedule</div><div class="stmt-desc">Tracks each asset category from beginning to ending balance: additions, disposals, and annual depreciation.</div><div class="stmt-fields"><span class="sf">Land</span><span class="sf">Infrastructure</span><span class="sf">Buildings</span><span class="sf">Accumulated depreciation</span></div></div>
    </div>
  </div>
  <div style="height:40px"></div>
</div></div>

<script>
  function switchTab(name,btn){
    document.querySelectorAll(".panel").forEach(p=>p.classList.remove("active"));
    document.querySelectorAll(".tab-btn").forEach(b=>b.classList.remove("active"));
    document.getElementById("panel-"+name).classList.add("active");
    btn.classList.add("active");
    if(name==='run') checkActiveJobs();
    if(name==='library') loadLibrary();
  }
  let selectedFiles=[],currentJobId=null,_pendingReconnectId=null;
  // Check for in-progress jobs on page load and whenever Run Pipeline tab is activated
  async function checkActiveJobs(){
    try{
      const d=await fetch('/jobs/active').then(r=>r.json());
      const running=d.jobs.filter(j=>j.status==='running');
      if(running.length){
        const j=running[0];
        _pendingReconnectId=j.job_id;
        const pct=j.total?Math.round(j.done/j.total*100):0;
        const label=j.label||j.job_id;
        const when=j.started_fmt?` · started ${j.started_fmt}`:'';
        document.getElementById('reconnect-msg').innerHTML=
          `<strong>${escHtml(label)}</strong>${escHtml(when)} &mdash; ${j.done}/${j.total||'?'} parishes (${pct}%). Reconnect to see live log.`;
        document.getElementById('reconnect-banner').style.display='flex';
      }
    }catch(_){}
  }
  async function reconnectJob(){
    if(!_pendingReconnectId) return;
    const jobId=_pendingReconnectId;
    document.getElementById('reconnect-banner').style.display='none';
    // Replay the log from disk first
    try{
      const d=await fetch('/job-log/'+jobId).then(r=>r.json());
      document.getElementById('log-card').style.display='block';
      document.getElementById('log').innerHTML='';
      appendLog('─── Reconnected to job '+jobId+' ───','');
      (d.lines||[]).forEach(l=>{if(l.trim())appendLog(l);});
      appendLog('─── Live progress below ───','');
      if(d.status==='done'){appendLog('Run already finished.','line-ok');setStatus('done');showDownloads(jobId);return;}
    }catch(_){}
    // Attach live SSE stream
    currentJobId=jobId;
    setStatus('running');
    const runBtn=document.getElementById('run-btn');
    runBtn.disabled=true;
    const source=new EventSource('/progress/'+jobId);
    source.onmessage=e=>{
      try{
        const msg=JSON.parse(e.data);
        if(msg==="__DONE__"){source.close();runBtn.disabled=false;if(_statusPollTimer){clearInterval(_statusPollTimer);_statusPollTimer=null;}setStatus("done","All parishes complete");showDownloads(jobId);}
        else if(typeof msg==="string"&&msg.trim())appendLog(msg);
      }catch(_){}
    };
    source.onerror=()=>{appendLog("⚠ Browser connection lost — job may still be running on the server. Auto-polling for status…","error");source.close();runBtn.disabled=false;setStatus("lost");_startPolling(jobId);};
  }
  const dropZone=document.getElementById("drop-zone"),fileInput=document.getElementById("file-input"),badge=document.getElementById("file-badge");
  document.getElementById("browse-link").addEventListener("click",e=>{e.stopPropagation();fileInput.click();});
  dropZone.addEventListener("click",()=>fileInput.click());
  dropZone.addEventListener("dragover",e=>{e.preventDefault();dropZone.classList.add("drag-over");});
  dropZone.addEventListener("dragleave",()=>dropZone.classList.remove("drag-over"));
  dropZone.addEventListener("drop",e=>{e.preventDefault();dropZone.classList.remove("drag-over");selectedFiles=[...e.dataTransfer.files].filter(f=>f.name.toLowerCase().endsWith(".pdf"));updateBadge();});
  fileInput.addEventListener("change",()=>{selectedFiles=[...fileInput.files];updateBadge();});

  // ── Canonical parish list for filename validation ──────────────────────────
  const _PARISHES=[
    'Acadia','Allen','Ascension','Assumption','Avoyelles','Beauregard','Bienville',
    'Bossier','Caddo','Calcasieu','Caldwell','Cameron','Catahoula','Claiborne',
    'Concordia','DeSoto','East Baton Rouge','East Carroll','East Feliciana','Evangeline',
    'Franklin','Grant','Iberia','Iberville','Jackson','Jefferson','Jefferson Davis',
    'Lafayette','Lafourche','LaSalle','Lincoln','Livingston','Madison','Morehouse',
    'Natchitoches','Orleans','Ouachita','Plaquemines','Pointe Coupee','Rapides',
    'Red River','Richland','Sabine','St. Bernard','St. Charles','St. Helena',
    'St. James','St. John the Baptist','St. Landry','St. Martin','St. Mary',
    'St. Tammany','Tangipahoa','Tensas','Terrebonne','Union','Vermilion','Vernon',
    'Washington','Webster','West Baton Rouge','West Carroll','West Feliciana','Winn'
  ];
  const _normP=s=>s.toLowerCase().replace(/[^a-z]/g,'');
  const _NORM_MAP=Object.fromEntries(_PARISHES.map(p=>[_normP(p),p]));

  function _bestMatch(name){
    const n=_normP(name);
    if(_NORM_MAP[n]) return {match:_NORM_MAP[n],exact:true};
    // Levenshtein distance for fuzzy suggestion
    let best=null,bestD=99;
    for(const [key,canon] of Object.entries(_NORM_MAP)){
      const d=_lev(n,key);
      if(d<bestD){bestD=d;best=canon;}
    }
    return bestD<=4?{match:best,exact:false,dist:bestD}:null;
  }
  function _lev(a,b){
    const m=a.length,n=b.length;
    const dp=Array.from({length:m+1},(_,i)=>Array.from({length:n+1},(_,j)=>i?j?0:i:j));
    for(let i=1;i<=m;i++) for(let j=1;j<=n;j++)
      dp[i][j]=a[i-1]===b[j-1]?dp[i-1][j-1]:1+Math.min(dp[i-1][j],dp[i][j-1],dp[i-1][j-1]);
    return dp[m][n];
  }

  let _fileWarnings=[];
  function _checkFilenames(files){
    _fileWarnings=[];
    for(const f of files){
      const name=f.name.replace(/\.pdf$/i,'');
      const result=_bestMatch(name);
      if(result&&result.exact) continue; // perfect match — all good
      if(result&&!result.exact){
        _fileWarnings.push({file:f.name,type:'warn',
          msg:`<strong>${escHtml(f.name)}</strong> — not an exact parish match. Did you mean <span class="fw-suggest">${escHtml(result.match)}.pdf</span>?`});
      } else {
        _fileWarnings.push({file:f.name,type:'bad',
          msg:`<strong>${escHtml(f.name)}</strong> — not a recognized Louisiana parish name. Rename to match exactly (e.g. <em>Jefferson Davis.pdf</em>).`});
      }
    }
    return _fileWarnings;
  }

  function updateBadge(){
    const wrap=document.getElementById('filename-warnings');
    if(!selectedFiles.length){
      badge.style.display='none'; wrap.style.display='none'; wrap.innerHTML=''; _fileWarnings=[];
      return;
    }
    badge.textContent=selectedFiles.length+' PDF'+(selectedFiles.length!==1?'s':'')+' selected';
    badge.style.display='inline-block';
    const warns=_checkFilenames(selectedFiles);
    if(!warns.length){wrap.style.display='none';wrap.innerHTML='';return;}
    wrap.style.display='block';
    wrap.innerHTML=warns.map(w=>`
      <div class="fw-row ${w.type}">
        <span class="fw-icon">${w.type==='bad'?'&#10060;':'&#9888;&#65039;'}</span>
        <span>${w.msg}</span>
      </div>`).join('');
  }
  function onYearChange(sel){
    const wrap=document.getElementById('year-field-wrap');
    if(sel.value){
      sel.classList.remove('placeholder');
      wrap.classList.remove('unset');
    } else {
      sel.classList.add('placeholder');
      wrap.classList.add('unset');
    }
  }
  async function runPipeline(){
    if(!selectedFiles.length){alert("Please select at least one PDF file.");return;}
    if(_fileWarnings.length){
      const badCount=_fileWarnings.filter(w=>w.type==='bad').length;
      const warnCount=_fileWarnings.filter(w=>w.type==='warn').length;
      let msg='';
      if(badCount) msg+=`${badCount} file${badCount>1?'s are':' is'} not a recognized parish name.\n`;
      if(warnCount) msg+=`${warnCount} file${warnCount>1?'s have':' has'} a possible typo.\n`;
      msg+='\nThe parish name comes directly from the filename. Mis-named files will be processed but results may not match correctly.\n\nRun anyway?';
      if(!confirm(msg)) return;
    }
    const yearVal=document.getElementById("year-input").value;
    if(!yearVal){
      const wrap=document.getElementById('year-field-wrap');
      wrap.classList.add('unset');
      wrap.style.animation='shake .35s ease';
      setTimeout(()=>wrap.style.animation='',400);
      alert("Please select a fiscal year before running.");
      return;
    }
    const year=parseInt(yearVal);
    const runBtn=document.getElementById("run-btn");
    runBtn.disabled=true;
    document.getElementById("log").innerHTML="";
    document.getElementById("downloads-card").style.display="none";
    document.getElementById("log-card").style.display="block";
    const _nParishes=selectedFiles.length;
    const _runLabel=`${year} — ${_nParishes===1?selectedFiles[0].name.replace(/\.pdf$/i,''):_nParishes+' parishes'}`;
    const _now=new Date();
    const _startFmt=_now.toLocaleString('en-US',{month:'short',day:'numeric',year:'numeric',hour:'numeric',minute:'2-digit'});
    setStatus("running",`${_runLabel} · started ${_startFmt}`);
    appendLog("Uploading "+_nParishes+" PDF(s)...","");
    const form=new FormData();form.append("year",year);
    for(const f of selectedFiles)form.append("files",f);
    let jobId;
    try{
      const resp=await fetch("/run",{method:"POST",body:form});
      if(!resp.ok){let d="Server error "+resp.status;try{const j=await resp.json();d=j.detail||d;}catch(_){}throw new Error(d);}
      ({job_id:jobId}=await resp.json());
    }catch(err){appendLog("[ERROR] "+err.message,"error");runBtn.disabled=false;setStatus("error");return;}
    currentJobId=jobId;
    appendLog(`Run: ${_runLabel}  ·  ${_startFmt}`,"line-header");
    appendLog("─".repeat(54),"");
    const source=new EventSource("/progress/"+jobId);
    source.onmessage=e=>{
      try{
        const msg=JSON.parse(e.data);
        if(msg==="__DONE__"){source.close();runBtn.disabled=false;if(_statusPollTimer){clearInterval(_statusPollTimer);_statusPollTimer=null;}setStatus("done","All parishes complete");showDownloads(jobId);}
        else if(typeof msg==="string"&&msg.trim())appendLog(msg);
      }catch(_){}
    };
    source.onerror=()=>{appendLog("⚠ Browser connection lost — job may still be running on the server. Auto-polling for status…","error");source.close();runBtn.disabled=false;setStatus("lost");_startPolling(currentJobId);};
  }
  function appendLog(msg,forceClass){const el=document.getElementById("log");const div=document.createElement("div");const cls=forceClass!==undefined?forceClass:classifyLine(msg);if(cls)div.className=cls;div.textContent=msg;el.appendChild(div);el.scrollTop=el.scrollHeight;}
  function classifyLine(msg){const m=msg.toLowerCase();if(m.includes("[error]"))return "line-error";if(m.includes("[skip]"))return "line-skip";if(m.includes("[manual]"))return "line-manual";if(m.includes("done.")||m.includes("saved:"))return "line-ok";if(/^\[\s*\d+\/\d+\]/.test(msg))return "line-header";return "";}
  const _STATUS_LABELS={running:'Running',done:'Complete',error:'Failed',lost:'Connection lost — polling…'};
  const _STATUS_COLORS={running:'var(--orange)',done:'var(--green)',error:'var(--red)',lost:'#d4a800'};
  let _statusPollTimer=null;
  function setStatus(s,detail){
    const dot=document.getElementById('status-dot');
    const lbl=document.getElementById('status-label');
    const det=document.getElementById('status-detail');
    dot.className='status-dot '+s;
    if(lbl){lbl.textContent=_STATUS_LABELS[s]||s;lbl.style.color=_STATUS_COLORS[s]||'var(--dimmer)';lbl.className=s;}
    if(det) det.textContent=detail||'';
  }
  function _startPolling(jobId){
    if(_statusPollTimer) clearInterval(_statusPollTimer);
    _statusPollTimer=setInterval(async()=>{
      try{
        const d=await fetch('/job-status/'+jobId).then(r=>r.json());
        const pct=d.done_count&&d.total_count?Math.round(d.done_count/d.total_count*100)+'%':'';
        const lbl=d.label||jobId;
        const when=d.started_fmt?` · ${d.started_fmt}`:'';
        if(d.status==='done'){
          clearInterval(_statusPollTimer);_statusPollTimer=null;
          setStatus('done',`${lbl}${when} — complete`);
          document.getElementById('run-btn').disabled=false;
          showDownloads(jobId);
        } else if(d.status==='error'){
          clearInterval(_statusPollTimer);_statusPollTimer=null;
          setStatus('error',`${lbl} — failed`);
          document.getElementById('run-btn').disabled=false;
        } else {
          setStatus('lost', pct ? `${lbl} — ${d.done_count}/${d.total_count} parishes · reconnect for live log` : `${lbl} — reconnect for live log`);
        }
      }catch(_){}
    },4000);
  }
  async function showDownloads(jobId){
    const data=await(await fetch("/files/"+jobId)).json();
    const body=document.getElementById("dl-body");body.innerHTML="";
    const byParish=data.by_parish||[],byStatement=data.by_statement||[];
    if(!byParish.length&&!byStatement.length){body.innerHTML="<p style='padding:14px;color:var(--dimmer);font-size:12px'>No Excel files produced.</p>";document.getElementById("downloads-card").style.display="block";return;}
    function makeSection(label,files,iconClass){
      if(!files.length)return;
      const sec=document.createElement("div");sec.className="dl-section";
      const lbl=document.createElement("div");lbl.className="dl-section-label";lbl.textContent=label;sec.appendChild(lbl);
      const grid=document.createElement("div");grid.className="dl-grid";
      for(const fname of files){
        const a=document.createElement("a");a.href="/download/"+jobId+"/"+encodeURIComponent(fname);a.download=fname;a.className="dl-btn";
        const stem=fname.replace(/\.xlsx$/i,"");const tag=iconClass==="parish"?"XLSX":stem.split(" ").pop();
        a.innerHTML='<span class="dl-icon '+iconClass+'">'+tag+'</span><span class="dl-name">'+fname+'</span>';
        grid.appendChild(a);
      }
      sec.appendChild(grid);body.appendChild(sec);
    }
    makeSection("By Parish — all statements in one file",byParish,"parish");
    makeSection("By Statement — all parishes in one file",byStatement,"statement");
    document.getElementById("downloads-card").style.display="block";
  }

  // ── Utility ───────────────────────────────────────────────────────────────
  function escHtml(s){return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');}
  function escAttr(s){return String(s||'').replace(/'/g,"\\'").replace(/"/g,'&quot;');}

  // ── Run import (Settings modal) ───────────────────────────────────────────
  let _importFiles=[];
  function _importFilesChosen(files){
    _importFiles=Array.from(files);
    const pdfs=_importFiles.filter(f=>f.name.toLowerCase().endsWith('.pdf'));
    const jsons=_importFiles.filter(f=>f.name.toLowerCase().endsWith('.json'));
    const xlsxs=_importFiles.filter(f=>f.name.toLowerCase().endsWith('.xlsx'));
    const prev=document.getElementById('import-preview');
    prev.style.display='block';
    prev.innerHTML=`<span style="color:var(--purple)">&#128196; ${pdfs.length} PDFs</span> &nbsp; <span style="color:var(--green)">&#128196; ${xlsxs.length} Excels</span> &nbsp; <span style="color:var(--dim)">&#128196; ${jsons.length} JSON cache</span> &nbsp; <b>${_importFiles.length} total</b>`;
  }
  async function importRunFiles(){
    const msg=document.getElementById('import-msg');
    if(!_importFiles.length){msg.textContent='No files selected.';msg.style.color='var(--orange)';return;}
    msg.textContent='Uploading…';msg.style.color='var(--dim)';
    const form=new FormData();
    _importFiles.forEach(f=>form.append('files',f));
    try{
      const r=await fetch('/admin/import-run',{method:'POST',body:form});
      const d=await r.json();
      if(!r.ok) throw new Error(d.detail||'Upload failed');
      msg.innerHTML=`&#10003; Imported: ${d.pdfs} PDFs, ${d.excels} Excels, ${d.cache} cache files into run <code>${d.run_id}</code>`;
      msg.style.color='var(--green)';
      _importFiles=[];document.getElementById('import-preview').style.display='none';
      loadLibrary();
    }catch(e){msg.textContent='Error: '+e.message;msg.style.color='var(--red)';}
  }

  // ── Spreadsheet view ──────────────────────────────────────────────────────
  let _xlYear=null, _xlParish=null, _xlTargetVal=null;
  function libSwitchView(mode){
    document.getElementById('lib-view-data').style.display=mode==='data'?'flex':'none';
    document.getElementById('lib-view-sheet').style.display=mode==='sheet'?'flex':'none';
    document.getElementById('vt-data').classList.toggle('active',mode==='data');
    document.getElementById('vt-sheet').classList.toggle('active',mode==='sheet');
    if(mode==='sheet'&&_xlYear&&_xlParish) libLoadSpreadsheet();
  }
  async function libLoadSpreadsheet(targetVal){
    if(targetVal!==undefined) _xlTargetVal=targetVal;
    const stmt=document.getElementById('xl-stmt-sel').value;
    const statusEl=document.getElementById('xl-status');
    statusEl.textContent='Loading…';
    document.getElementById('xl-truth').innerHTML='<p class="xl-missing">Loading…</p>';
    document.getElementById('xl-pipe').innerHTML='<p class="xl-missing">Loading…</p>';
    try{
      const r=await fetch(`/excel-view/${_xlYear}/${encodeURIComponent(_xlParish)}/${stmt}`);
      const d=await r.json();
      _xlRenderSide('xl-truth','xl-truth-scroll',d.truth,_xlTargetVal,'truth');
      _xlRenderSide('xl-pipe','xl-pipe-scroll',d.pipeline,_xlTargetVal,'pipe');
      statusEl.textContent=d.truth?'':'No coded file found for this year.';
    }catch(e){statusEl.textContent='Error: '+e.message;}
  }
  function _goToRunPipeline(){
    const btn=document.querySelector('.tab-btn');
    const allBtns=document.querySelectorAll('.tab-btn');
    const runBtn=Array.from(allBtns).find(b=>b.textContent.trim().startsWith('Run Pipeline'));
    if(runBtn) switchTab('run',runBtn);
    if(_xlYear){
      const sel=document.getElementById('year-input');
      if(sel){sel.value=String(_xlYear);sel.classList.remove('placeholder');onYearChange(sel);}
    }
  }
  function _xlRenderSide(elId,scrollId,rows,targetVal,side){
    const el=document.getElementById(elId);
    const scroll=document.getElementById(scrollId);
    if(!rows||!rows.length){
      if(side==='pipe'){
        el.innerHTML=`<div style="display:flex;flex-direction:column;align-items:center;justify-content:center;height:100%;gap:12px;padding:32px;text-align:center">
          <p class="xl-missing" style="margin:0">No pipeline output found for ${_xlYear||'this year'}.</p>
          <p style="font-size:11px;color:var(--dimmer);margin:0">Run the pipeline for ${_xlYear||'this year'} to generate the comparison spreadsheet.</p>
          <button onclick="_goToRunPipeline()" style="padding:8px 18px;background:var(--purple);color:#fff;border:none;border-radius:6px;cursor:pointer;font-size:12px;font-weight:600">
            &#9654; Run Pipeline for ${_xlYear||'this year'}
          </button>
        </div>`;
      } else {
        el.innerHTML='<p class="xl-missing">Coded truth file not found for this year.</p>';
      }
      return;
    }
    // Build HTML table
    const ncols=Math.max(...rows.map(r=>r.length));
    let html='<table class="xl-tbl"><tr class="xl-hdr"><td class="xl-rn">#</td>';
    for(let c=0;c<ncols;c++) html+=`<td>${String.fromCharCode(65+c)}</td>`;
    html+='</tr>';
    let targetRow=-1,targetCol=-1;
    rows.forEach((row,ri)=>{
      html+=`<tr><td class="xl-rn">${ri+1}</td>`;
      for(let ci=0;ci<ncols;ci++){
        const v=row[ci]!=null?row[ci]:'';
        const vs=String(v);
        let cls='';
        const isTarget=targetVal!=null&&vs!==''&&(vs===String(targetVal)||vs.replace(/,/g,'')===String(targetVal).replace(/,/g,''));
        if(isTarget){cls='xl-target';if(targetRow<0){targetRow=ri;targetCol=ci;}}
        html+=`<td class="${cls}" title="${escHtml(vs)}">${escHtml(vs)}</td>`;
      }
      html+='</tr>';
    });
    html+='</table>';
    el.innerHTML=html;
    // Scroll target into view
    if(targetRow>=0){
      requestAnimationFrame(()=>{
        const tbl=el.querySelector('.xl-tbl');
        if(!tbl) return;
        const trs=tbl.querySelectorAll('tr');
        const tr=trs[targetRow+1]; // +1 for header row
        if(tr){
          const td=tr.querySelectorAll('td')[targetCol+1]; // +1 for row-num col
          if(td) td.scrollIntoView({block:'center',inline:'center'});
        }
      });
    }
  }

  // ── All Flags view ────────────────────────────────────────────────────────
  let _allFlagsData=[];
  function _libShowViewMain(which){
    document.getElementById('lib-empty').style.display='none';
    document.getElementById('lib-dashboard').style.display='none';
    document.getElementById('lib-viewer').style.display='none';
    document.getElementById('lib-allflags-view').style.display='none';
    if(which==='dash') document.getElementById('lib-dashboard').style.display='flex';
    else if(which==='viewer') document.getElementById('lib-viewer').style.display='flex';
    else if(which==='allflags') document.getElementById('lib-allflags-view').style.display='flex';
    else document.getElementById('lib-empty').style.display='flex';
  }
  async function libShowAllFlags(){
    document.querySelectorAll('.lib-parish-row').forEach(r=>r.classList.remove('active'));
    document.querySelectorAll('.lib-year-btn').forEach(b=>b.classList.remove('active'));
    document.getElementById('lib-allflags-btn').classList.add('active');
    _libShowViewMain('allflags');
    // populate year select
    const yrSel=document.getElementById('af-year-sel');
    const years=Object.keys(_libByYear).map(Number).sort((a,b)=>b-a);
    yrSel.innerHTML='<option value="">All years</option>'+years.map(y=>`<option value="${y}">${y}</option>`).join('');
    await libLoadAllFlags();
  }
  async function libLoadAllFlags(){
    const year=document.getElementById('af-year-sel').value;
    const stmt=document.getElementById('af-stmt-sel').value;
    const listEl=document.getElementById('af-list');
    listEl.innerHTML='<p style="color:var(--dimmer);font-size:12px">Loading…</p>';
    try{
      // Load all flag data for selected year(s)
      const years=year?[parseInt(year)]:Object.keys(_libByYear).map(Number);
      let allFlags=[];
      for(const yr of years){
        const d=await fetch('/compare-flags/'+yr).then(r=>r.json()).catch(()=>({items:[]}));
        allFlags=allFlags.concat((d.items||[]).map(f=>({...f,year:yr})));
      }
      if(stmt) allFlags=allFlags.filter(f=>f.statement_type===stmt);
      allFlags.sort((a,b)=>(['high','medium','low'].indexOf(a.severity)-['high','medium','low'].indexOf(b.severity))||(a.parish||'').localeCompare(b.parish||''));
      document.getElementById('af-count').textContent=allFlags.length+' flags';
      document.getElementById('allflags-count').textContent=allFlags.length;
      if(!allFlags.length){listEl.innerHTML='<p style="color:var(--dimmer);font-size:13px;text-align:center;padding:40px">No flags found.</p>';return;}
      listEl.innerHTML=allFlags.map(f=>`
        <div style="background:#fff;border:1px solid var(--border);border-radius:7px;padding:12px 14px;margin-bottom:10px;cursor:pointer" onclick="openLibEntry(${f.year},'${escAttr(f.parish||'')}')">
          <div style="display:flex;align-items:center;gap:7px;margin-bottom:5px;flex-wrap:wrap">
            <span style="font-weight:700;font-size:13px">${escHtml(f.parish||'')}</span>
            <span style="font-size:11px;color:var(--dimmer)">${f.year}</span>
            <span class="badge-stmt">${(f.statement_type||'').toUpperCase()}</span>
            <span class="badge-sev sev-${f.severity||'low'}">${f.severity||''}</span>
          </div>
          <div style="font-size:11.5px;font-weight:600;color:var(--purple)">${escHtml(f.field_path||'')}${f.fund?' — '+escHtml(f.fund):''}</div>
          <div style="display:flex;gap:16px;margin-top:4px;font-size:11.5px">
            ${f.extracted_value!=null?`<span>Pipeline: <b style="font-family:monospace">${Number(f.extracted_value).toLocaleString()}</b></span>`:''}
            ${f.alternative_value!=null?`<span>Coded: <b style="font-family:monospace;color:var(--orange)">${Number(f.alternative_value).toLocaleString()}</b></span>`:''}
          </div>
        </div>`).join('');
    }catch(e){listEl.innerHTML='<p style="color:var(--red)">Error loading flags: '+e.message+'</p>';}
  }
  // ── PDF page viewer (review tab — knows job_id & page number) ────────────
  function viewPage(jobId,parish,page,value,snippet){
    _pdfBrowse=null;
    document.getElementById('pdf-nav').style.display='none';
    document.getElementById('modal-title').textContent=parish+' — Page '+page+(value?' (highlighted)':'');
    document.getElementById('modal-body').innerHTML='<p class="modal-loading">Rendering page'+(value?' and locating value...':'...')+'</p>';
    document.getElementById('page-modal').classList.add('open');
    const params=new URLSearchParams();
    if(value&&value.trim()) params.set('value',value.trim());
    if(snippet&&snippet.trim()) params.set('snippet',snippet.trim().substring(0,120));
    const qs=params.toString();
    const url='/page-image/'+jobId+'/'+encodeURIComponent(parish)+'/'+page+(qs?'?'+qs:'');
    _loadPageImg(url,value,false);
  }

  // ── PDF browser (training tab — no job_id, navigable) ────────────────────
  let _pdfBrowse=null; // {year,parish,page,totalPages,value,field}
  async function browseParishPdf(year,parish,value,field){
    document.getElementById('page-modal').classList.add('open');
    document.getElementById('pdf-nav').style.display='none';
    document.getElementById('modal-title').textContent=parish+' — Loading PDF…';
    document.getElementById('modal-body').innerHTML='<p class="modal-loading">'+(value?'Searching for value in PDF…':'Finding PDF…')+'</p>';
    try{
      const qs=value?'?value='+encodeURIComponent(value):'';
      const info=await(await fetch('/pdf-info/'+year+'/'+encodeURIComponent(parish)+qs)).json();
      if(info.error||!info.total_pages){
        document.getElementById('modal-body').innerHTML='<p class="modal-loading">'+escHtml(info.error||'No PDF found for '+parish+' '+year+'. Run the pipeline for this parish first.')+'</p>';
        return;
      }
      const startPage=info.best_page||1;
      _pdfBrowse={year,parish,page:startPage,totalPages:info.total_pages,value,field};
      document.getElementById('pdf-nav').style.display='flex';
      _renderBrowsePage();
    }catch(e){
      document.getElementById('modal-body').innerHTML='<p class="modal-loading">Error: '+escHtml(e.message)+'</p>';
    }
  }
  function _renderBrowsePage(){
    if(!_pdfBrowse)return;
    const {year,parish,page,totalPages,value,field}=_pdfBrowse;
    document.getElementById('modal-title').textContent=parish+' '+year+(field?' — '+field:'');
    document.getElementById('pdf-nav-label').textContent='Page '+page+' / '+totalPages;
    document.getElementById('modal-body').innerHTML='<p class="modal-loading">Rendering page '+page+'…</p>';
    const params=new URLSearchParams();
    if(value&&value.trim()) params.set('value',value.trim());
    const url='/pdf-browse/'+year+'/'+encodeURIComponent(parish)+'/'+page+(params.toString()?'?'+params.toString():'');
    _loadPageImg(url,value,true);
  }
  function pdfNavStep(delta){
    if(!_pdfBrowse)return;
    const next=_pdfBrowse.page+delta;
    if(next<1||next>_pdfBrowse.totalPages)return;
    _pdfBrowse.page=next;
    _renderBrowsePage();
  }
  function _loadPageImg(url,value,showNav){
    const img=new Image();
    img.onload=()=>{
      const wrap=document.createElement('div');
      wrap.style.cssText='position:relative';
      img.style.cssText='max-width:100%;display:block;border-radius:4px';
      wrap.appendChild(img);
      if(value){
        const note=document.createElement('p');
        note.style.cssText='font-size:11px;color:var(--orange);padding:6px 0 2px;text-align:center';
        note.textContent='Orange boxes = located value on page';
        wrap.appendChild(note);
      }
      if(showNav&&_pdfBrowse){
        const hint=document.createElement('p');
        hint.style.cssText='font-size:11px;color:var(--dim);padding:2px 0 4px;text-align:center';
        hint.textContent='Use ← → arrows to navigate pages';
        wrap.appendChild(hint);
      }
      document.getElementById('modal-body').innerHTML='';
      document.getElementById('modal-body').appendChild(wrap);
    };
    img.onerror=()=>{document.getElementById('modal-body').innerHTML='<p class="modal-loading">Could not render page — PDF may no longer be available on the server.</p>';};
    img.src=url;
  }
  function closeModal(){document.getElementById('page-modal').classList.remove('open');_pdfBrowse=null;document.getElementById('pdf-nav').style.display='none';}
  document.getElementById('page-modal').addEventListener('click',e=>{if(e.target===e.currentTarget)closeModal();});
  // Keyboard navigation for PDF browser
  document.addEventListener('keydown',e=>{
    if(!document.getElementById('page-modal').classList.contains('open')||!_pdfBrowse)return;
    if(e.key==='ArrowLeft')pdfNavStep(-1);
    else if(e.key==='ArrowRight')pdfNavStep(1);
    else if(e.key==='Escape')closeModal();
  });

  // ── Flag chat assistant ───────────────────────────────────────────────────
  let _chatContext=null, _chatHistory=[], _chatHistories={};
  function openChat(itemId){
    const item=_reviewItems[itemId];
    if(!item)return;
    _chatContext=item;
    // Restore existing history for this item, or start fresh
    _chatHistory=_chatHistories[itemId]||[];
    // Build context summary
    const stmtNames={cbs:'Combined Balance Sheet',soa:'Statement of Activities',
                     sona:'Statement of Net Position',ca:'Capital Assets'};
    document.getElementById('chat-ctx').innerHTML=
      `<strong>${item.parish} ${item.year} — ${stmtNames[item.statement_type]||item.statement_type.toUpperCase()}</strong><br>`+
      `<strong>Field:</strong> ${item.field_path||''}${item.fund?' — '+item.fund:''} &nbsp;·&nbsp; `+
      `<strong>Severity:</strong> ${item.severity} &nbsp;·&nbsp; `+
      `<strong>Source:</strong> ${item.source}<br>`+
      `<strong>Reason:</strong> ${item.reason||'—'}`;
    document.getElementById('chat-modal-title').textContent='Ask AI — '+item.parish+' '+item.statement_type.toUpperCase();
    const msgsEl=document.getElementById('chat-msgs');
    msgsEl.innerHTML='';
    if(_chatHistory.length===0){
      // Auto-ask the opening question
      const opener="Can you explain what this flag means and what I should do about it?";
      _chatHistory.push({role:'user',content:opener});
      appendChatBubble('user',opener);
      document.getElementById('chat-modal').classList.add('open');
      doAsk();
    } else {
      _chatHistory.forEach(m=>appendChatBubble(m.role,m.content));
      document.getElementById('chat-modal').classList.add('open');
      msgsEl.scrollTop=msgsEl.scrollHeight;
    }
  }
  function closeChatModal(){
    if(_chatContext) _chatHistories[_chatContext.id]=_chatHistory;
    document.getElementById('chat-modal').classList.remove('open');
  }
  function _classifyCompare(msg){
    const m=msg.toLowerCase();
    if(m.includes('[error]')) return 'line-error';
    if(m.includes('done.')) return 'line-ok';
    if(m.includes('flags')) return 'line-ok';
    if(m.includes('comparing')) return 'line-header';
    return '';
  }

  // ── Settings / API key ────────────────────────────────────────────────────
  async function refreshKeyDot(){
    try{
      const d=await fetch('/api/key-status').then(r=>r.json());
      const dot=document.getElementById('key-dot');
      const banner=document.getElementById('no-key-banner');
      if(d.configured){
        dot.style.background='#1a7f37';dot.title='API key configured';
        if(banner)banner.style.display='none';
      }else{
        dot.style.background='#f85149';dot.title='No API key — click gear to add one';
        if(banner)banner.style.display='block';
      }
      const cur=document.getElementById('settings-key-current');
      if(cur)cur.textContent=d.configured?'Current key: '+d.masked:'No key saved yet.';
    }catch(_){}
  }
  function openSettings(){
    document.getElementById('settings-modal').style.display='flex';
    document.getElementById('settings-key-input').value='';
    document.getElementById('settings-msg').textContent='';
    refreshKeyDot();
  }
  function closeSettings(){document.getElementById('settings-modal').style.display='none';}
  async function saveSettings(){
    const key=document.getElementById('settings-key-input').value.trim();
    const msg=document.getElementById('settings-msg');
    if(!key){msg.style.color='var(--red)';msg.textContent='Please enter a key.';return;}
    try{
      const resp=await fetch('/api/key',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({key})});
      const d=await resp.json();
      if(!resp.ok){msg.style.color='var(--red)';msg.textContent=d.detail||'Error saving key.';return;}
      msg.style.color='var(--green)';msg.textContent='Saved! Key: '+d.masked;
      refreshKeyDot();
      setTimeout(closeSettings,1200);
    }catch(e){msg.style.color='var(--red)';msg.textContent='Network error: '+e.message;}
  }
  document.getElementById('settings-modal').addEventListener('click',e=>{if(e.target===e.currentTarget)closeSettings();});

  // ── Library ───────────────────────────────────────────────────────────────
  let _libData=[], _libByYear={}, _libCurrent=null, _libPage=1, _libTotal=1, _libActiveYear=null;

  async function loadLibrary(){
    try{
      const d=await fetch('/library-data').then(r=>r.json());
      _libData=d.entries||[];
      _libByYear={};
      _libData.forEach(e=>{(_libByYear[e.year]=_libByYear[e.year]||[]).push(e);});
      renderLibSidebar();
      const yrs=Object.keys(_libByYear).length;
      const badge=document.getElementById('library-badge');
      badge.textContent=yrs; badge.style.display='inline-block';
    }catch(e){console.error('library',e);}
  }

  function renderLibSidebar(filter=''){
    const el=document.getElementById('lib-year-list');
    const years=Object.keys(_libByYear).map(Number).sort((a,b)=>b-a);
    if(!years.length){el.innerHTML='<div style="padding:20px 14px;font-size:12px;color:var(--dimmer)">No processed data yet.</div>';return;}
    const fl=filter.toLowerCase();
    let html='';
    years.forEach(yr=>{
      const parishes=(_libByYear[yr]||[]).filter(e=>!fl||e.parish.toLowerCase().includes(fl))
                                         .sort((a,b)=>a.parish.localeCompare(b.parish));
      const isOpen=(yr===_libActiveYear); // only active year expands; filter just filters rows
      const caret=isOpen?'▾':'▸';
      html+=`<div class="lib-year-item">
        <button class="lib-year-btn ${yr===_libActiveYear&&!_libCurrent?'active':''}" data-year="${yr}" onclick="openLibYear(${yr})">
          <span>${yr}</span><span class="lib-year-count">${parishes.length} parishes</span>
          <span class="lib-year-caret" id="lib-caret-${yr}">${caret}</span>
        </button>
        <div class="lib-parishes ${isOpen?'open':''}" id="lib-parishes-${yr}">`;
      parishes.forEach(e=>{
        const dots=['cbs','sona','ca','soa'].map(s=>
          `<span class="dash-stmt-dot ${e.statements.includes(s)?'ok':''}" title="${s.toUpperCase()}"></span>`).join('');
        const pdfB=e.has_pdf?'<span class="lib-badge pdf" style="margin-left:2px">PDF</span>':'';
        html+=`<div class="lib-parish-row ${_libCurrent&&_libCurrent.year===yr&&_libCurrent.parish===e.parish?'active':''}"
          id="librow-${yr}-${encodeURIComponent(e.parish)}" onclick="openLibEntry(${yr},'${escHtml(e.parish).replace(/'/g,"\\'")}')">
          <span style="flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">${escHtml(e.parish)}</span>
          <div class="lib-badges">${dots}${pdfB}</div>
        </div>`;
      });
      html+=`</div></div>`;
    });
    el.innerHTML=html;
  }

  function filterLibrary(q){renderLibSidebar(q);}

  function _libShowOnly(which){_libShowViewMain(which);}

  // ── Training comparison state ─────────────────────────────────────────────
  let _trainFiles=[], _trainYear=null;
  function libTrainFilesChosen(files){
    _trainFiles=Array.from(files);
    const pdfs=_trainFiles.filter(f=>f.name.toLowerCase().endsWith('.pdf')).length;
    const xls=_trainFiles.filter(f=>f.name.toLowerCase().endsWith('.xlsx')).length;
    const btn=document.getElementById('lib-compare-btn');
    if(btn){ btn.disabled=false; btn.textContent=`Upload & Compare (${pdfs} PDFs, ${xls} Excels)`; }
  }
  function libPickFiles(year){
    _trainYear=year;
    document.getElementById('train-all-files').value='';
    document.getElementById('train-all-files').click();
  }
  async function libRunCompare(year){
    _trainYear=year;
    const logEl=document.getElementById('lib-compare-log');
    const statusEl=document.getElementById('lib-compare-status');
    if(logEl) logEl.innerHTML='';
    if(statusEl) statusEl.textContent='Starting…';

    let jobId;
    if(_trainFiles.length){
      // Split into PDFs and Excels
      const pdfs=_trainFiles.filter(f=>f.name.toLowerCase().endsWith('.pdf'));
      const xlsxs=_trainFiles.filter(f=>f.name.toLowerCase().endsWith('.xlsx'));
      const form=new FormData();
      form.append('year',year);
      xlsxs.forEach(f=>form.append('files',f));
      pdfs.forEach(f=>form.append('pdfs',f));
      try{
        const r=await fetch('/upload-coded',{method:'POST',body:form});
        if(!r.ok) throw new Error('Upload failed: '+r.status);
        const d=await r.json(); jobId=d.job_id;
      }catch(e){if(statusEl)statusEl.textContent='Error: '+e.message; return;}
    } else {
      // Re-run with existing coded files
      try{
        const r=await fetch('/rerun-compare/'+year,{method:'POST'});
        if(!r.ok) throw new Error('Server error '+r.status);
        const d=await r.json(); jobId=d.job_id;
      }catch(e){if(statusEl)statusEl.textContent='Error: '+e.message; return;}
    }
    // Stream progress
    const wrap=document.getElementById('lib-compare-log-wrap');
    if(wrap) wrap.style.display='block';
    const dotEl=document.getElementById('lib-compare-dot');
    if(dotEl) dotEl.className='status-dot running';
    const src=new EventSource('/compare-progress/'+jobId);
    src.onmessage=e=>{
      try{
        const msg=JSON.parse(e.data);
        if(msg==='__DONE__'){
          src.close();
          if(dotEl) dotEl.className='status-dot done';
          fetch('/compare-status/'+jobId).then(r=>r.json()).then(d=>{
            const s=d.summary||{};
            if(statusEl) statusEl.innerHTML='<span style="color:var(--green)">&#10003; '+
              (s.total_flags||0)+' flags across '+(s.parishes||[]).length+' parishes.</span>';
            openLibYear(year); // refresh dashboard
          });
        } else if(typeof msg==='string'&&msg.trim()&&logEl){
          const div=document.createElement('div');
          div.className=_classifyCompare(msg); div.textContent=msg;
          logEl.appendChild(div); logEl.scrollTop=logEl.scrollHeight;
        }
      }catch(_){}
    };
    src.onerror=()=>{src.close();if(dotEl)dotEl.className='status-dot error';};
  }

  async function openLibYear(year){
    // Toggle: clicking the already-open year closes it and clears the panel
    if(_libActiveYear===year && !_libCurrent){
      _libActiveYear=null; _libCurrent=null; _trainFiles=[];
      document.querySelectorAll('.lib-year-btn').forEach(b=>b.classList.remove('active'));
      document.querySelectorAll('.lib-parishes').forEach(pl=>pl.classList.remove('open'));
      document.querySelectorAll('.lib-year-caret').forEach(c=>c.textContent='▸');
      _libShowOnly('empty');
      return;
    }
    _libActiveYear=year; _libCurrent=null; _trainFiles=[];
    // Accordion: close all years, open only the selected one
    document.querySelectorAll('.lib-year-btn').forEach(b=>b.classList.remove('active'));
    document.querySelectorAll('.lib-parishes').forEach(pl=>pl.classList.remove('open'));
    document.querySelectorAll('.lib-year-caret').forEach(c=>c.textContent='▸');
    const yb=document.querySelector(`.lib-year-btn[data-year="${year}"]`);
    if(yb){
      yb.classList.add('active');
      const pl=document.getElementById('lib-parishes-'+year);
      if(pl) pl.classList.add('open');
      const caret=document.getElementById('lib-caret-'+year);
      if(caret) caret.textContent='▾';
    }
    _libShowOnly('dash');
    document.getElementById('lib-dash-title').textContent=year+' — Training Dashboard';
    const body=document.getElementById('lib-dash-body');
    body.innerHTML='<div style="color:var(--dimmer);font-size:13px;padding:20px">Loading…</div>';

    const CANONICAL_PARISHES=[
      'Acadia','Allen','Ascension','Assumption','Avoyelles','Beauregard','Bienville',
      'Bossier','Caddo','Calcasieu','Caldwell','Cameron','Catahoula','Claiborne',
      'Concordia','DeSoto','East Baton Rouge','East Carroll','East Feliciana','Evangeline',
      'Franklin','Grant','Iberia','Iberville','Jackson','Jefferson','Jefferson Davis',
      'Lafayette','Lafourche','LaSalle','Lincoln','Livingston','Madison','Morehouse',
      'Natchitoches','Orleans','Ouachita','Plaquemines','Pointe Coupee','Rapides',
      'Red River','Richland','Sabine','St. Bernard','St. Charles','St. Helena',
      'St. James','St. John the Baptist','St. Landry','St. Martin','St. Mary',
      'St. Tammany','Tangipahoa','Tensas','Terrebonne','Union','Vermilion','Vernon',
      'Washington','Webster','West Baton Rouge','West Carroll','West Feliciana','Winn'
    ];
    const TOTAL_LA_PARISHES=64;

    // Strip variant suffixes to get canonical base name
    function _baseName(p){return p.replace(/\s+\(\d+\)$/,'').replace(/\s+-\s*REISSUE.*$/i,'').trim();}
    function _frac(n,d){return `${n}<span style="font-size:55%;color:var(--dimmer);font-weight:400"> / ${d}</span>`;}

    const entries=(_libByYear[year]||[]).sort((a,b)=>a.parish.localeCompare(b.parish));

    // Group by base name to detect duplicates
    const baseGroups={};
    entries.forEach(e=>{const bn=_baseName(e.parish);(baseGroups[bn]=baseGroups[bn]||[]).push(e);});
    const outlierGroups=Object.entries(baseGroups).filter(([,g])=>g.length>1);
    const uniqueBaseNames=new Set(Object.keys(baseGroups));

    // Use one representative per base name for counting (prefer exact match, else first)
    const dedupedEntries=Object.values(baseGroups).map(g=>g.find(e=>_baseName(e.parish)===e.parish)||g[0]);

    const withPdf=dedupedEntries.filter(e=>e.has_pdf).length;
    const stmtCounts={cbs:0,sona:0,ca:0,soa:0};
    dedupedEntries.forEach(e=>e.statements.forEach(s=>{if(stmtCounts[s]!=null)stmtCounts[s]++;}));

    // Fetch only comparison flags — no bulk parish data
    const flagData=await fetch('/compare-flags/'+year).then(r=>r.json()).catch(()=>({items:[]}));
    const flags=flagData.items||[];
    const compFlags=flags.filter(f=>f.source==='comparison').length;
    const codeFlags=flags.filter(f=>f.source==='coding_check').length;
    const resolved=flags.filter(f=>f.resolved).length;
    const flagsByParish={};
    flags.forEach(f=>{(flagsByParish[f.parish]=flagsByParish[f.parish]||[]).push(f);});

    const totalStmts=stmtCounts.cbs+stmtCounts.sona+stmtCounts.ca+stmtCounts.soa;
    const missingParishes=CANONICAL_PARISHES.filter(p=>!uniqueBaseNames.has(p));

    // KPI row
    let html=`<div class="dash-kpi-row">
      <div class="dash-kpi"><div class="dash-kpi-val">${_frac(uniqueBaseNames.size,TOTAL_LA_PARISHES)}</div><div class="dash-kpi-label">Parishes cached</div></div>
      <div class="dash-kpi"><div class="dash-kpi-val">${_frac(withPdf,TOTAL_LA_PARISHES)}</div><div class="dash-kpi-label">With PDF</div></div>
      <div class="dash-kpi"><div class="dash-kpi-val">${_frac(totalStmts,TOTAL_LA_PARISHES*4)}</div><div class="dash-kpi-label">Statements extracted</div></div>
      <div class="dash-kpi"><div class="dash-kpi-val" style="color:${compFlags?'var(--orange)':'var(--green)'}">${compFlags}</div><div class="dash-kpi-label">Discrepancy flags</div></div>
      <div class="dash-kpi"><div class="dash-kpi-val" style="color:${codeFlags?'var(--orange)':'var(--green)'}">${codeFlags}</div><div class="dash-kpi-label">Coding checks</div></div>
      <div class="dash-kpi"><div class="dash-kpi-val">${resolved}</div><div class="dash-kpi-label">Resolved</div></div>
    </div>`;

    // Outlier warning — duplicate base-name entries
    if(outlierGroups.length){
      const msgs=outlierGroups.map(([bn,g])=>`<b>${escHtml(bn)}</b>: ${g.map(e=>escHtml(e.parish)).join(', ')}`).join(' &nbsp;·&nbsp; ');
      html+=`<div style="background:rgba(255,170,0,.12);border:1px solid rgba(255,170,0,.35);border-radius:6px;padding:8px 12px;margin-bottom:12px;font-size:12px;color:var(--text)">
        <span style="color:var(--orange);font-weight:700">⚠ Duplicate entries detected</span> — only the canonical file is counted in totals above. Review and remove extras if they are re-runs of the same parish.<br>
        <span style="color:var(--dim)">${msgs}</span>
      </div>`;}


    // Comparison upload / re-run section
    html+=`<div class="dash-section-hdr" style="display:flex;align-items:center;gap:10px">
      Coded Excel Comparison
      <span style="margin-left:auto;display:flex;gap:8px">
        <button class="rv-btn" onclick="libPickFiles(${year})">&#128196; Upload files</button>
        <button class="rv-btn" id="lib-compare-btn" onclick="libRunCompare(${year})">${flags.length?'&#8635; Re-run':'Run Comparison'}</button>
      </span>
    </div>
    <div id="lib-compare-status" style="font-size:12px;color:var(--dim);margin-bottom:6px">${flags.length?'Last run: '+flags.length+' flags found.':'Upload coded Excels (.xlsx) and source PDFs to run comparison.'}</div>
    <div id="lib-compare-log-wrap" style="display:none;margin-bottom:16px">
      <div style="display:flex;align-items:center;gap:8px;margin-bottom:6px">
        <div class="status-dot" id="lib-compare-dot"></div>
        <span style="font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.6px;color:var(--dim)">Comparison Progress</span>
      </div>
      <div class="log" id="lib-compare-log" style="height:180px"></div>
    </div>`;

    // Parish list — names + statement indicators + flag count only (no financial values)
    html+=`<div class="dash-section-hdr">Parishes — ${year}</div>
    <table class="dash-table"><thead><tr>
      <th>Parish</th><th style="text-align:center">CBS</th><th style="text-align:center">SONA</th>
      <th style="text-align:center">CA</th><th style="text-align:center">SOA</th>
      <th style="text-align:center">PDF</th><th style="text-align:center">Flags</th>
    </tr></thead><tbody>`;
    entries.forEach(e=>{
      const ok='<span style="color:var(--green);font-weight:700">&#10003;</span>';
      const no='<span style="color:var(--border)">—</span>';
      const pFlags=(flagsByParish[e.parish]||[]).length;
      const flagCell=pFlags?`<span style="color:var(--orange);font-weight:700">${pFlags}</span>`:no;
      const isDup=_baseName(e.parish)!==e.parish; // has a variant suffix → duplicate
      const dupBadge=isDup?` <span style="font-size:10px;padding:1px 5px;border-radius:3px;background:rgba(255,170,0,.2);color:var(--orange);font-weight:700">DUP</span>`:'';
      html+=`<tr class="clickable${isDup?' dup-row':''}" onclick="openLibEntry(${year},'${escHtml(e.parish).replace(/'/g,"\\'")}')">
        <td>${escHtml(e.parish)}${dupBadge}</td>
        <td style="text-align:center">${e.statements.includes('cbs')?ok:no}</td>
        <td style="text-align:center">${e.statements.includes('sona')?ok:no}</td>
        <td style="text-align:center">${e.statements.includes('ca')?ok:no}</td>
        <td style="text-align:center">${e.statements.includes('soa')?ok:no}</td>
        <td style="text-align:center">${e.has_pdf?ok:no}</td>
        <td style="text-align:center">${flagCell}</td>
      </tr>`;
    });
    html+='</tbody></table>';

    // Missing parishes collapsible
    if(missingParishes.length){
      html+=`<details style="margin-top:16px;border:1px solid var(--border);border-radius:6px;overflow:hidden">
        <summary style="padding:10px 14px;cursor:pointer;font-size:12.5px;font-weight:600;background:var(--surf);display:flex;align-items:center;gap:8px;list-style:none;user-select:none">
          <span style="color:var(--dimmer)">&#9660;</span>
          Missing parishes
          <span style="margin-left:6px;background:var(--surf2);border:1px solid var(--border);border-radius:10px;padding:1px 8px;font-size:11px;font-weight:700;color:var(--dim)">${missingParishes.length} of ${TOTAL_LA_PARISHES}</span>
        </summary>
        <div style="padding:10px 14px;display:flex;flex-wrap:wrap;gap:6px">
          ${missingParishes.map(p=>`<span style="font-size:11.5px;padding:3px 9px;border-radius:4px;background:var(--surf2);border:1px solid var(--border);color:var(--dimmer)">${escHtml(p)}</span>`).join('')}
        </div>
      </details>`;
    }
    body.innerHTML=html;
  }

  async function openLibEntry(year, parish){
    _libCurrent={year,parish}; _libActiveYear=year;
    _xlYear=year; _xlParish=parish; _xlTargetVal=null;
    document.getElementById('lib-allflags-btn').classList.remove('active');
    renderLibSidebar(document.getElementById('lib-search').value);
    _libShowOnly('viewer');
    // Reset to data view when opening new parish
    libSwitchView('data');
    _libPage=1;
    document.getElementById('lib-viewer-title').textContent=parish+' ('+year+')';
    const info=await fetch('/pdf-info/'+year+'/'+encodeURIComponent(parish)).then(r=>r.json()).catch(()=>({total_pages:0,best_page:1}));
    _libTotal=info.total_pages||0;
    _libPage=info.best_page||1;
    _libUpdatePageCtrl();
    _libRenderPage();
    _libRenderData(year,parish);
  }

  function libBackToDash(){if(_libActiveYear) openLibYear(_libActiveYear);}

  function _libUpdatePageCtrl(){
    document.getElementById('lib-page-label').textContent=_libTotal?`${_libPage} / ${_libTotal}`:'No PDF';
    document.getElementById('lib-prev').disabled=_libPage<=1||!_libTotal;
    document.getElementById('lib-next').disabled=_libPage>=_libTotal||!_libTotal;
  }
  function libPageStep(d){
    _libPage=Math.max(1,Math.min(_libTotal,_libPage+d));
    _libUpdatePageCtrl(); _libRenderPage();
  }
  function _libRenderPage(){
    const pane=document.getElementById('lib-pdf-pane');
    if(!_libTotal){pane.innerHTML='<div style="color:var(--dimmer);font-size:13px;padding:40px">No PDF for this parish.</div>';return;}
    pane.innerHTML='<div style="color:var(--dimmer);font-size:13px">Loading…</div>';
    const img=new Image();
    img.onload=()=>{pane.innerHTML='';pane.appendChild(img);};
    img.onerror=()=>{pane.innerHTML='<div style="color:var(--red);padding:20px">Could not load page.</div>';};
    img.style.cssText='max-width:100%;border-radius:4px;box-shadow:0 2px 12px rgba(0,0,0,.2)';
    img.src='/pdf-browse/'+_libCurrent.year+'/'+encodeURIComponent(_libCurrent.parish)+'/'+_libPage;
  }
  async function _libRenderData(year,parish){
    const pane=document.getElementById('lib-data-pane');
    pane.innerHTML='<div style="padding:16px 14px;font-size:12px;color:var(--dimmer)">Loading…</div>';
    try{
      const [d, flagData]=await Promise.all([
        fetch('/library-entry/'+year+'/'+encodeURIComponent(parish)).then(r=>r.json()).catch(()=>({})),
        fetch('/compare-flags/'+year).then(r=>r.json()).catch(()=>({items:[]}))
      ]);
      const parishFlags=(flagData.items||[]).filter(f=>f.parish===parish);
      const sections=[
        {src:d.cbs,label:'Balance Sheet',fields:[['Total Fund Balances','fund_balances.total_fund_balances']]},
        {src:d.sona,label:'Net Assets',fields:[
          ['Total Assets','governmental_activities.total_assets'],
          ['Total Liabilities','governmental_activities.total_liabilities'],
          ['Net Position',null,()=>getNestedVal(d,'net_position.total_net_position')],
        ]},
        {src:d.ca,label:'Capital Assets',fields:[
          ['Net (beg)','governmental_activities.total_capital_assets_net.beginning'],
          ['Net (end)','governmental_activities.total_capital_assets_net.ending'],
        ]},
        {src:d.soa,label:'Activities',fields:[
          ['Change in Net','governmental_activities.change_in_net_assets'],
          ['Net Position End','governmental_activities.net_position_ending'],
        ]},
      ];
      let html='';
      sections.forEach(sec=>{
        if(!sec.src) return;
        html+=`<div class="lib-data-section"><h4>${sec.label}</h4>`;
        sec.fields.forEach(([label,path,fn])=>{
          const val=fn?fn():getNestedVal(sec.src,path);
          html+=`<div class="lib-data-row"><span class="lib-data-label">${label}</span><span class="lib-data-val">${escHtml(fmtNum(val))}</span></div>`;
        });
        html+='</div>';
      });
      // Comparison flags — grouped by statement type
      const sevColor={high:'var(--red)',medium:'var(--orange)',low:'var(--dim)'};
      const stmtOrder=['cbs','sona','ca','soa'];
      const stmtLabels={cbs:'Balance Sheet (CBS)',sona:'Statement of Net Position',ca:'Capital Assets',soa:'Statement of Activities'};
      if(parishFlags.length){
        // Group flags by statement type
        const byStmt={};
        parishFlags.forEach(f=>{const s=f.statement_type||'unknown';if(!byStmt[s])byStmt[s]=[];byStmt[s].push(f);});
        const orderedKeys=[...stmtOrder.filter(s=>byStmt[s]),...Object.keys(byStmt).filter(s=>!stmtOrder.includes(s))];
        orderedKeys.forEach(stmtType=>{
          const flags=byStmt[stmtType];
          const unresolved=flags.filter(f=>!f.resolved).length;
          const stmtLabel=stmtLabels[stmtType]||stmtType.toUpperCase();
          const noteAreaId=`stmtnote-${parish.replace(/[^a-z0-9]/gi,'_')}-${year}-${stmtType}`;
          html+=`<div class="lib-data-section">
            <h4 style="color:var(--orange)">&#9888; ${escHtml(stmtLabel)} &mdash; ${flags.length} Flag${flags.length>1?'s':''} <span style="font-size:11px;font-weight:400;color:var(--dimmer)">(${unresolved} unresolved)</span></h4>
            <div style="margin-bottom:10px;padding:8px 10px;background:rgba(88,166,255,.05);border:1px solid rgba(88,166,255,.18);border-radius:5px">
              <div style="font-size:10.5px;color:var(--blue);font-weight:700;margin-bottom:5px">&#128172; Note to Claude — ${escHtml(stmtLabel)}</div>
              <textarea id="${noteAreaId}" placeholder="Explain the issue in plain language, e.g. &#34;This is a comparative statement — 2013 values are in the LEFT column, 2012 in the right. Always use the left column.&#34;" style="width:100%;padding:6px 8px;background:var(--surf);border:1px solid var(--border);border-radius:4px;color:var(--text);font-size:11px;resize:vertical;min-height:56px;box-sizing:border-box;font-family:inherit"></textarea>
              <div style="display:flex;align-items:center;gap:8px;margin-top:5px">
                <button onclick="saveStmtNote(${JSON.stringify(parish)},${year},${JSON.stringify(stmtType)},${JSON.stringify(noteAreaId)},this)"
                  style="padding:4px 12px;background:rgba(88,166,255,.12);border:1px solid rgba(88,166,255,.3);color:var(--blue);border-radius:4px;cursor:pointer;font-size:10.5px;font-weight:700">
                  Save note
                </button>
                <span id="${noteAreaId}-saved" style="display:none;font-size:10.5px;color:var(--green);font-weight:700">&#10003; Saved — will be injected into next re-run</span>
              </div>
            </div>`;
          flags.forEach(f=>{
            const col=sevColor[f.severity]||'var(--dim)';
            const field=(f.field_path||'').split('.').pop().replace(/_/g,' ');
            const extVal=f.extracted_value||'—';
            const altVal=f.alternative_value||'—';
            const resolved=f.resolved;
            html+=`<div id="libflag-${f.id}" style="border:1px solid var(--border);border-radius:6px;padding:8px 10px;margin-bottom:8px;font-size:11.5px;${resolved?'opacity:.55':''}">
              <div style="display:flex;align-items:center;gap:6px;margin-bottom:5px">
                <span style="width:6px;height:6px;border-radius:50%;background:${resolved?'var(--green)':col};flex-shrink:0"></span>
                <span style="font-weight:600;color:var(--text)">${escHtml(field)}</span>
                ${resolved?`<span style="margin-left:auto;font-size:10px;color:var(--green);font-weight:700">&#10003; RESOLVED</span>`:''}
              </div>
              <div style="display:flex;justify-content:space-between;gap:4px;margin-bottom:2px">
                <span style="color:var(--dimmer)">Pipeline:</span><span style="font-family:monospace;color:var(--text)">${escHtml(extVal)}</span>
              </div>
              <div style="display:flex;justify-content:space-between;gap:4px;margin-bottom:6px">
                <span style="color:var(--dimmer)">Coded Excel:</span><span style="font-family:monospace;color:var(--orange)">${escHtml(altVal)}</span>
              </div>
              <button onclick="libHighlightFlag(this)"
                data-value="${escAttr(extVal)}" data-page="${f.page_number||''}"
                style="width:100%;padding:5px 8px;background:rgba(111,66,193,.1);border:1px solid rgba(111,66,193,.3);color:var(--purple);border-radius:4px;cursor:pointer;font-size:11px;font-weight:600;margin-bottom:4px;transition:background .1s,transform .08s,box-shadow .1s"
                onmousedown="this.style.transform='scale(.97)'" onmouseup="this.style.transform=''"
                onmouseleave="this.style.transform=''" onmouseenter="this.style.background='rgba(111,66,193,.2)'">
                &#128269; Find in PDF
              </button>
              <button data-xlstmt="${escAttr(f.statement_type||'cbs')}" data-xlval="${escAttr(extVal)}"
                onclick="document.getElementById('xl-stmt-sel').value=this.dataset.xlstmt;_xlTargetVal=this.dataset.xlval;libSwitchView('sheet')"
                style="width:100%;padding:5px 8px;background:rgba(26,127,55,.08);border:1px solid rgba(26,127,55,.3);color:var(--green);border-radius:4px;cursor:pointer;font-size:11px;font-weight:600;margin-bottom:4px">
                &#128196; View in Spreadsheet
              </button>
              <button onclick="libDiagnoseFlag(${f.id},this)"
                style="width:100%;padding:5px 8px;background:rgba(139,92,246,.08);border:1px solid rgba(139,92,246,.25);color:#a78bfa;border-radius:4px;cursor:pointer;font-size:11px;font-weight:600;margin-bottom:6px;transition:background .1s,transform .08s"
                onmousedown="this.style.transform='scale(.97)'" onmouseup="this.style.transform=''"
                onmouseleave="this.style.transform=''" onmouseenter="this.style.background='rgba(139,92,246,.18)'">
                &#128300; Diagnose — ask Claude to explain this discrepancy
              </button>
              <div id="libdiag-${f.id}" style="display:none;margin-bottom:6px"></div>
              ${!resolved?`
              <div style="display:flex;gap:5px;margin-bottom:0">
                <button onclick="libResolveFlag(${f.id},'confirmed',this)"
                  style="flex:1;padding:5px 4px;background:rgba(26,127,55,.1);border:1px solid rgba(26,127,55,.35);color:var(--green);border-radius:4px;cursor:pointer;font-size:10.5px;font-weight:700">
                  &#10003; Pipeline correct
                </button>
                <button onclick="libResolveFlag(${f.id},'coded',this)"
                  style="flex:1;padding:5px 4px;background:rgba(210,100,0,.1);border:1px solid rgba(210,100,0,.35);color:var(--orange);border-radius:4px;cursor:pointer;font-size:10.5px;font-weight:700"
                  data-coded="${escAttr(altVal)}">
                  &#8635; Use coded value
                </button>
                <button onclick="libToggleCorrectForm(${f.id})"
                  style="flex:1;padding:5px 4px;background:var(--surf2);border:1px solid var(--border);color:var(--dim);border-radius:4px;cursor:pointer;font-size:10.5px;font-weight:700">
                  &#9998; Enter value
                </button>
              </div>
              <div id="libcorrect-${f.id}" style="display:none;flex-direction:column;gap:5px;margin-top:6px">
                <input id="libcorrval-${f.id}" type="text" placeholder="Correct value (number)" style="padding:5px 8px;border:1px solid var(--border);border-radius:4px;font-size:11.5px;background:var(--surf);color:var(--text)"/>
                <textarea id="libcorrnote-${f.id}" placeholder="Tell Claude why — explain in plain language (e.g. &#34;comparative statement, 2013 is the left column&#34;)" style="padding:6px 8px;border:1px solid rgba(88,166,255,.3);border-radius:4px;font-size:11px;background:var(--surf);color:var(--text);resize:vertical;min-height:52px;font-family:inherit"></textarea>
                <div style="display:flex;gap:5px">
                  <button onclick="libSubmitCorrection(${f.id})" style="flex:1;padding:5px;background:rgba(26,127,55,.1);border:1px solid rgba(26,127,55,.35);color:var(--green);border-radius:4px;cursor:pointer;font-size:11px;font-weight:700">Save correction</button>
                  <button onclick="libToggleCorrectForm(${f.id})" style="padding:5px 10px;background:var(--surf2);border:1px solid var(--border);color:var(--dim);border-radius:4px;cursor:pointer;font-size:11px">Cancel</button>
                </div>
              </div>`:''}
            </div>`;
          });
          html+='</div>';
        });
      } else {
        html+=`<div class="lib-data-section"><h4 style="color:var(--green)">&#10003; No flags</h4><div style="font-size:11.5px;color:var(--dimmer)">No comparison discrepancies for this parish.</div></div>`;
      }
      pane.innerHTML=html||'<div style="padding:14px;font-size:12px;color:var(--dimmer)">No fields.</div>';
    }catch(e){pane.innerHTML='<div style="padding:14px;font-size:12px;color:var(--red)">Error.</div>';}
  }

  function libToggleCorrectForm(id){
    const el=document.getElementById('libcorrect-'+id);
    el.style.display=el.style.display==='flex'?'none':'flex';
  }
  async function libResolveFlag(id, action, btn){
    let correctedValue=null;
    if(action==='coded') correctedValue=btn.dataset.coded||null;
    btn.disabled=true; btn.textContent='Saving…';
    try{
      await fetch('/feedback/'+id,{method:'POST',
        headers:{'Content-Type':'application/json'},
        body:JSON.stringify({action:action==='coded'?'corrected':'confirmed',
          corrected_value:correctedValue, note:action==='coded'?'Coded Excel value accepted':null})
      });
      _markLibFlagResolved(id);
    }catch(e){btn.disabled=false; btn.textContent='Error';}
  }
  async function libSubmitCorrection(id){
    const val=document.getElementById('libcorrval-'+id).value.trim();
    const note=document.getElementById('libcorrnote-'+id).value.trim();
    if(!val){alert('Enter the correct value first.');return;}
    const btn=document.querySelector(`#libcorrect-${id} button`);
    if(btn){btn.disabled=true;btn.textContent='Saving…';}
    try{
      await fetch('/feedback/'+id,{method:'POST',
        headers:{'Content-Type':'application/json'},
        body:JSON.stringify({action:'corrected',corrected_value:val,note:note||null})
      });
      _markLibFlagResolved(id);
    }catch(e){if(btn){btn.disabled=false;btn.textContent='Save correction';}}
  }
  function _markLibFlagResolved(id){
    const card=document.getElementById('libflag-'+id);
    if(!card) return;
    card.style.opacity='0.55';
    // Swap severity dot to green, add RESOLVED badge, remove action buttons
    const dot=card.querySelector('span[style*="border-radius:50%"]');
    if(dot) dot.style.background='var(--green)';
    const label=card.querySelector('span[style*="font-weight:600"]');
    if(label&&!card.querySelector('.lib-resolved-badge')){
      const badge=document.createElement('span');
      badge.className='lib-resolved-badge';
      badge.style.cssText='margin-left:auto;font-size:10px;color:var(--green);font-weight:700';
      badge.textContent='✓ RESOLVED';
      label.parentElement.appendChild(badge);
    }
    // Remove the action row
    const actionRow=card.querySelector('div[style*="display:flex;gap:5px"]');
    if(actionRow) actionRow.remove();
    const correctForm=card.querySelector(`#libcorrect-${id}`);
    if(correctForm) correctForm.remove();
  }

  async function libHighlightFlag(btn){
    const value=btn.dataset.value||'';
    const knownPage=parseInt(btn.dataset.page)||0;
    if(!_libCurrent) return;
    const pane=document.getElementById('lib-pdf-pane');

    // Guard: pipeline didn't extract a value for this field
    const isNull=(value===''||value==='—'||value==='null');
    if(isNull){
      pane.innerHTML=`<div style="padding:20px;font-size:12px;color:var(--orange);line-height:1.6">
        <strong>&#128683; Pipeline did not extract a value for this field.</strong><br>
        The PDF may use a layout Claude couldn't parse, or this line item doesn't exist in this parish's statements.<br>
        <span style="color:var(--dimmer)">Use "Note to Claude" above to explain the layout and re-run the pipeline to attempt extraction.</span>
      </div>`;
      return;
    }

    if(!_libTotal){
      pane.innerHTML='<div style="padding:20px;font-size:12px;color:var(--red)">No PDF available for this parish. Upload it to use Find in PDF.</div>';
      return;
    }

    // Spinner keyframes (inject once)
    if(!document.getElementById('spin-style')){
      const st=document.createElement('style');st.id='spin-style';
      st.textContent='@keyframes spin{to{transform:rotate(360deg)}}';
      document.head.appendChild(st);
    }

    const origHTML=btn.innerHTML;
    btn.disabled=true; btn.style.opacity='0.75'; btn.style.cursor='default';
    btn.innerHTML='<span style="display:inline-block;animation:spin .7s linear infinite">&#8635;</span> Searching…';
    pane.innerHTML='<div style="color:var(--dimmer);font-size:12px;padding:40px;text-align:center">&#128269; Locating value in PDF…</div>';

    try{
      let found=false;
      // Ask server to locate the value unless we already have the page
      if(knownPage>0){
        _libPage=knownPage; found=true;
      } else {
        const info=await fetch(
          '/pdf-info/'+_libCurrent.year+'/'+encodeURIComponent(_libCurrent.parish)
          +'?value='+encodeURIComponent(value)
        ).then(r=>r.json()).catch(()=>null);
        if(info&&info.found){_libPage=info.best_page; found=true;}
        else if(info&&info.best_page) _libPage=info.best_page;
      }
      _libUpdatePageCtrl();

      // Show not-found banner if search failed (still render best-guess page)
      let notFoundBanner='';
      if(!found){
        notFoundBanner=`<div style="padding:6px 10px;background:rgba(230,100,0,.12);border-radius:5px;margin-bottom:8px;font-size:11px;color:var(--orange)">
          &#9888; Value <strong>${escHtml(value)}</strong> not found in PDF text — showing statement page. The PDF may be scanned/image-only or use a different number format.
        </div>`;
      }

      const imgUrl='/pdf-browse/'+_libCurrent.year+'/'+encodeURIComponent(_libCurrent.parish)
        +'/'+_libPage+'?value='+encodeURIComponent(value)+'&trace=1&spread=1';

      await new Promise((resolve,reject)=>{
        const img=new Image();
        img.onload=()=>{
          pane.innerHTML=notFoundBanner;
          pane.appendChild(img);
          resolve();
        };
        img.onerror=()=>{
          pane.innerHTML=notFoundBanner+
            '<div style="color:var(--red);padding:20px;font-size:12px">&#9888; Could not render page.</div>';
          reject();
        };
        img.style.cssText='max-width:100%;border-radius:4px;box-shadow:0 2px 12px rgba(0,0,0,.2)';
        img.src=imgUrl;
      });
    } catch(e){
      // already handled in img.onerror; don't overwrite the message
    } finally {
      btn.disabled=false; btn.style.opacity=''; btn.style.cursor='pointer';
      btn.innerHTML=origHTML;
    }
  }

  async function saveStmtNote(parish, year, stmtType, textareaId, btn){
    const note=(document.getElementById(textareaId)||{}).value||'';
    if(!note.trim()){alert('Please type a note before saving.');return;}
    const origText=btn.textContent;
    btn.disabled=true; btn.textContent='Saving…';
    try{
      const r=await fetch('/statement-note',{method:'POST',
        headers:{'Content-Type':'application/json'},
        body:JSON.stringify({parish,year,stmt_type:stmtType,note:note.trim()})
      });
      if(!r.ok) throw new Error('Server error');
      const savedEl=document.getElementById(textareaId+'-saved');
      if(savedEl) savedEl.style.display='inline';
      btn.textContent=origText;
      btn.disabled=false;
    }catch(e){
      btn.disabled=false;
      btn.textContent='Error — retry';
    }
  }

  async function libDiagnoseFlag(id, btn){
    const panel=document.getElementById('libdiag-'+id);
    // Toggle: if already showing a result, collapse it
    if(panel.style.display!=='none'&&panel.dataset.loaded){
      panel.style.display='none'; panel.dataset.loaded='';
      btn.innerHTML='&#128300; Diagnose &mdash; ask Claude to explain this discrepancy';
      return;
    }
    if(!document.getElementById('spin-style')){
      const st=document.createElement('style');st.id='spin-style';
      st.textContent='@keyframes spin{to{transform:rotate(360deg)}}';
      document.head.appendChild(st);
    }
    const origHTML=btn.innerHTML;
    btn.disabled=true;
    btn.innerHTML='<span style="display:inline-block;animation:spin .7s linear infinite">&#8635;</span> Diagnosing…';
    panel.style.display='block';
    panel.dataset.loaded='';
    panel.innerHTML=`<div style="padding:10px 12px;border:1px solid rgba(139,92,246,.2);border-radius:6px;font-size:11px;color:var(--dimmer)">
      <span style="display:inline-block;animation:spin .7s linear infinite">&#8635;</span> Reading PDF and comparing values — this takes a few seconds…
    </div>`;
    try{
      const d=await fetch('/diagnose-flag/'+id,{method:'POST'}).then(r=>{
        if(!r.ok) throw new Error('Server error '+r.status);
        return r.json();
      });
      // Verdict styling
      const vColor=d.verdict==='pipeline_correct'?'var(--green)':d.verdict==='coded_correct'?'var(--orange)':'var(--dimmer)';
      const vIcon =d.verdict==='pipeline_correct'?'&#10003;':d.verdict==='coded_correct'?'&#9888;':'&#63;';
      const vLabel=d.verdict==='pipeline_correct'?'Pipeline value appears correct'
                  :d.verdict==='coded_correct'    ?'Coded Excel value appears correct'
                  :'Uncertain — manual review needed';
      const cColor=d.confidence==='high'?'var(--green)':d.confidence==='medium'?'var(--orange)':'var(--red)';
      const cLabel=(d.confidence||'?').toUpperCase();

      // Source quote helper
      const srcBox=(label,text,color)=>text?`
        <div style="margin-top:6px">
          <span style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.4px;color:var(--dimmer)">${label}</span>
          <div style="margin-top:2px;padding:5px 8px;background:rgba(0,0,0,.18);border-left:2px solid ${color};border-radius:0 4px 4px 0;font-size:10.5px;font-style:italic;color:var(--text);line-height:1.45">${escHtml(text)}</div>
        </div>`:'';

      const secondGuess=d.second_guess!=null?`
        <div style="margin-top:8px;padding:6px 10px;background:rgba(88,166,255,.08);border:1px solid rgba(88,166,255,.2);border-radius:5px">
          <span style="font-size:10px;font-weight:700;color:var(--blue);text-transform:uppercase;letter-spacing:.4px">Second guess</span>
          <span style="margin-left:8px;font-family:monospace;font-size:12px;color:var(--text);font-weight:700">${escHtml(String(d.second_guess))}</span>
          ${d.second_guess_source?`<div style="margin-top:3px;font-size:10.5px;color:var(--dimmer);font-style:italic">${escHtml(d.second_guess_source)}</div>`:''}
        </div>`:'' ;

      // Auto-fill correction form if verdict is coded_correct
      if(d.verdict==='coded_correct'){
        const altInput=document.getElementById('libcorrval-'+id);
        if(altInput&&!altInput.value){
          // peek at the coded value from the Use coded button
          const codedBtn=document.querySelector(`#libflag-${id} button[data-coded]`);
          if(codedBtn) altInput.value=codedBtn.dataset.coded||'';
        }
      }

      panel.innerHTML=`
        <div style="background:rgba(139,92,246,.06);border:1px solid rgba(139,92,246,.22);border-radius:6px;padding:10px 12px;font-size:11.5px">
          <div style="display:flex;align-items:center;gap:8px;margin-bottom:8px;flex-wrap:wrap">
            <span style="font-weight:700;color:${vColor};font-size:12px">${vIcon} ${vLabel}</span>
            <span style="margin-left:auto;font-size:10px;font-weight:700;color:${cColor};letter-spacing:.3px">${cLabel} CONFIDENCE</span>
          </div>
          <p style="margin:0 0 6px;color:var(--text);line-height:1.55;font-size:11.5px">${escHtml(d.explanation||'')}</p>
          ${d.verdict_reason?`<div style="padding:5px 9px;background:rgba(0,0,0,.2);border-radius:4px;font-size:11px;color:var(--dim);margin-bottom:6px">&#128161; ${escHtml(d.verdict_reason)}</div>`:''}
          ${srcBox('Pipeline read this as',d.pipeline_source,'var(--text)')}
          ${srcBox('Coded Excel used',d.coded_source,'var(--orange)')}
          ${secondGuess}
          <div style="margin-top:10px;display:flex;gap:6px;flex-wrap:wrap">
            ${d.verdict==='pipeline_correct'?`<button onclick="libResolveFlag(${id},'confirmed',this)" style="padding:4px 10px;background:rgba(26,127,55,.1);border:1px solid rgba(26,127,55,.35);color:var(--green);border-radius:4px;cursor:pointer;font-size:10.5px;font-weight:700">&#10003; Accept — pipeline correct</button>`:''}
            ${d.verdict==='coded_correct'?`<button onclick="libResolveFlag(${id},'coded',this)" data-coded="${escAttr(document.querySelector('#libflag-'+id+' button[data-coded]')?.dataset?.coded||'')}" style="padding:4px 10px;background:rgba(210,100,0,.1);border:1px solid rgba(210,100,0,.35);color:var(--orange);border-radius:4px;cursor:pointer;font-size:10.5px;font-weight:700">&#8635; Accept — use coded value</button>`:''}
            <button onclick="libToggleCorrectForm(${id})" style="padding:4px 10px;background:var(--surf2);border:1px solid var(--border);color:var(--dim);border-radius:4px;cursor:pointer;font-size:10.5px">&#9998; Enter value manually</button>
          </div>
        </div>`;
      panel.dataset.loaded='1';
      btn.innerHTML='&#128300; Diagnosis &#9650; (click to collapse)';
      btn.disabled=false;
    }catch(e){
      panel.innerHTML=`<div style="padding:8px 10px;border:1px solid rgba(255,60,60,.3);border-radius:6px;font-size:11px;color:var(--red)">&#9888; Diagnosis failed: ${escHtml(e.message||'unknown error')}. Check that the server is running and try again.</div>`;
      panel.dataset.loaded='1';
      btn.disabled=false; btn.innerHTML=origHTML;
    }
  }

  function getNestedVal(obj,path){
    if(!obj||!path) return null;
    return path.split('.').reduce((o,k)=>o&&o[k]!=null?o[k]:null, obj);
  }
  function fmtNum(v){
    if(v==null) return '—';
    const n=typeof v==='string'?parseFloat(v.replace(/[,$]/g,'')):v;
    if(isNaN(n)) return String(v);
    return (n<0?'-':'')+'$'+Math.abs(n).toLocaleString('en-US',{maximumFractionDigits:0});
  }

  // Check for active jobs on page load
  checkActiveJobs();
  refreshKeyDot();

  document.getElementById('chat-modal').addEventListener('click',e=>{if(e.target===e.currentTarget)closeChatModal();});
  function appendChatBubble(role,text){
    const el=document.getElementById('chat-msgs');
    const div=document.createElement('div');
    div.className='chat-bubble '+role;
    div.textContent=text;
    el.appendChild(div);
    el.scrollTop=el.scrollHeight;
    return div;
  }
  function sendChat(){
    const input=document.getElementById('chat-input');
    const text=input.value.trim();
    if(!text||!_chatContext)return;
    input.value='';
    _chatHistory.push({role:'user',content:text});
    appendChatBubble('user',text);
    doAsk();
  }
  async function doAsk(){
    const sendBtn=document.getElementById('chat-send');
    sendBtn.disabled=true;
    const thinking=appendChatBubble('assistant thinking','Thinking…');
    try{
      const resp=await fetch('/ask',{
        method:'POST',
        headers:{'Content-Type':'application/json'},
        body:JSON.stringify({context:_chatContext, messages:_chatHistory})
      });
      if(!resp.ok) throw new Error('Server error '+resp.status);
      const {response}=await resp.json();
      thinking.className='chat-bubble assistant';
      thinking.textContent=response;
      _chatHistory.push({role:'assistant',content:response});
    }catch(err){
      thinking.className='chat-bubble assistant';
      thinking.textContent='[Error: '+err.message+']';
    }finally{
      sendBtn.disabled=false;
      document.getElementById('chat-msgs').scrollTop=9999;
    }
  }
</script>
</body></html>"""

# ══════════════════════════════════════════════════════════════════════════════
# FASTAPI SERVER
# ══════════════════════════════════════════════════════════════════════════════

from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile, Body as _Body
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse

_seed_data()  # pre-populate Training with seeded run data on first start
app = FastAPI()
JOBS: dict[str, dict] = {}

_STMT_NAMES = {
    "cbs":  "Combined Balance Sheet of Governmental Funds",
    "soa":  "Statement of Activities",
    "sona": "Statement of Net Position / Net Assets",
    "ca":   "Capital Assets Schedule",
}

@app.post("/ask")
async def ask_about_flag(payload: dict = _Body(default={})):
    """
    Chat endpoint — explain a review flag in plain English.
    Body: { "context": <uncertainty dict>, "messages": [{"role":..,"content":..}] }
    """
    ctx  = payload.get("context", {})
    msgs = payload.get("messages", [])

    stmt_full = _STMT_NAMES.get(str(ctx.get("statement_type","")).lower(),
                                str(ctx.get("statement_type","")).upper())
    src_note = ("The AI model flagged this field because it was uncertain during extraction."
                if ctx.get("source") == "claude"
                else "The accounting validation engine flagged this because an identity check failed "
                     "(e.g. Total Assets ≠ Total Liabilities + Fund Balances).")

    system = f"""You are an expert Louisiana government finance auditor assistant. \
A user is reviewing an automated flag raised on a parish audit statement and needs help understanding it.

FLAGGED ITEM:
  Parish       : {ctx.get('parish','?')}
  Fiscal Year  : {ctx.get('year','?')}
  Statement    : {stmt_full}
  Field        : {ctx.get('field_path','?')}{(' — ' + ctx.get('fund')) if ctx.get('fund') else ''}
  Extracted    : {ctx.get('extracted_value','not available')}
  Alternative  : {ctx.get('alternative_value','none')}
  Severity     : {ctx.get('severity','?')}
  Flag source  : {ctx.get('source','?').upper()} — {src_note}
  Reason given : {ctx.get('reason','not specified')}
  PDF snippet  : {ctx.get('text_snippet','not available')}

YOUR ROLE:
1. Explain in plain English what this flag means and why it matters.
2. Tell the user exactly what to look for in the PDF to verify whether the extracted value is right.
3. Explain the accounting concept briefly if it helps (e.g. why total assets must equal total liabilities + fund balances).
4. Give a clear recommendation: confirm the extraction, correct it, or flag for a human auditor.
Keep answers concise, use bullet points, and avoid jargon where possible."""

    try:
        resp = _get_client().messages.create(
            model="claude-sonnet-4-5",
            max_tokens=1024,
            system=system,
            messages=msgs if msgs else [{"role":"user","content":"Can you explain this flag and what I should do about it?"}],
        )
        return {"response": resp.content[0].text}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/api/key-status")
def api_key_status():
    key = _get_api_key()
    if key:
        masked = key[:8] + "..." + key[-4:]
        return {"configured": True, "masked": masked}
    return {"configured": False, "masked": ""}

@app.post("/api/key")
async def api_key_save(request: Request):
    body = await request.json()
    key = (body.get("key") or "").strip()
    if not key:
        raise HTTPException(400, "Key cannot be empty")
    if not key.startswith("sk-ant-"):
        raise HTTPException(400, "That doesn't look like an Anthropic key (should start with sk-ant-)")
    _save_api_key(key)
    masked = key[:8] + "..." + key[-4:]
    return {"ok": True, "masked": masked}

@app.get("/", response_class=HTMLResponse)
def index():
    return HTMLResponse(_HTML)

@app.post("/run")
async def run(files: list[UploadFile] = File(...), year: int = Form(...)):
    if not _get_api_key():
        raise HTTPException(400, "No API key configured. Open Settings (gear icon) and paste your Anthropic key.")
    job_id  = str(uuid.uuid4())[:8]
    # Use a permanent runs/ folder next to server.py (resolve() ensures absolute path)
    runs_root = _app_dir() / "runs"
    job_dir   = runs_root / job_id
    pdf_dir   = job_dir / "pdfs"
    out_dir   = job_dir / "output"
    cache_dir = out_dir / "cache"
    pdf_dir.mkdir(parents=True, exist_ok=True)
    out_dir.mkdir(parents=True, exist_ok=True)
    cache_dir.mkdir(parents=True, exist_ok=True)
    print(f"[{job_id}] Output dir: {out_dir}", flush=True)
    parish_names = []
    for f in files:
        data = await f.read()
        print(f"[{job_id}] Received: {f.filename} ({len(data):,} bytes)", flush=True)
        (pdf_dir/f.filename).write_bytes(data)
        parish_names.append(os.path.splitext(f.filename)[0])
    # Write human-readable metadata for this run
    from datetime import datetime as _dt
    _started = _dt.now()
    _label = f"{year} — {len(parish_names)} parish{'es' if len(parish_names)!=1 else ''}"
    if len(parish_names) == 1:
        _label = f"{year} — {parish_names[0]}"
    _meta = {
        "job_id":    job_id,
        "label":     _label,
        "year":      year,
        "started_at": _started.isoformat(),
        "started_fmt": _started.strftime("%b %d %Y, %I:%M %p").replace(" 0", " "),
        "parish_count": len(parish_names),
        "parishes":  parish_names,
    }
    (job_dir / "meta.json").write_text(json.dumps(_meta, indent=2))
    q: queue.Queue[str] = queue.Queue()
    JOBS[job_id] = {"queue": q, "status": "running", "output_dir": str(out_dir),
                    "pdf_dir": str(pdf_dir), "label": _label,
                    "year": year, "started_fmt": _meta["started_fmt"]}

    _log_path = out_dir / "server.log"

    def _run():
        import traceback
        try:
            os.environ["ANTHROPIC_API_KEY"] = _get_api_key()
            def _cb(m):
                q.put(m)
                try:
                    with open(_log_path, "a", encoding="utf-8") as _lf:
                        _lf.write(m + "\n")
                except Exception:
                    pass
            run_pipeline(str(pdf_dir), year, str(out_dir), str(cache_dir),
                         skip_cache=True, progress_callback=_cb, job_id=job_id)
            JOBS[job_id]["status"] = "done"
        except Exception as exc:
            tb = traceback.format_exc()
            print(f"[{job_id}] CRASH:\n{tb}", flush=True)
            try:
                with open(_log_path, "a", encoding="utf-8") as _lf:
                    _lf.write(f"[CRASH]\n{tb}\n")
            except Exception:
                pass
            q.put(f"[ERROR] {exc}"); JOBS[job_id]["status"] = "error"
        finally:
            q.put("__DONE__")
    threading.Thread(target=_run, daemon=True).start()
    return {"job_id": job_id}

@app.get("/library-data")
def library_data():
    """Return all parishes that have cached extraction data, with PDF availability."""
    entries = []
    seen = set()
    runs_root = _app_dir() / "runs"
    # Also check flat legacy output dir
    search_dirs = []
    if runs_root.exists():
        search_dirs += [d / "output" / "cache" for d in runs_root.iterdir()]
    flat_out = _app_dir() / "output"
    if flat_out.exists():
        for yr_dir in flat_out.iterdir():
            search_dirs.append(yr_dir / "cache")
    for cache_dir in search_dirs:
        if not cache_dir.exists():
            continue
        for jf in cache_dir.glob("*.json"):
            # Filename pattern: ParishName_YYYY.json
            parts = jf.stem.rsplit("_", 1)
            if len(parts) != 2 or not parts[1].isdigit():
                continue
            parish, year = parts[0], int(parts[1])
            key = (parish, year)
            if key in seen:
                continue
            seen.add(key)
            try:
                data = json.loads(jf.read_text())
            except Exception:
                continue
            stmts = [s for s in ["cbs", "sona", "ca", "soa"] if data.get(s)]
            has_pdf = _find_parish_pdf(year, parish) is not None
            entries.append({"parish": parish, "year": year,
                            "statements": stmts, "has_pdf": has_pdf})
    entries.sort(key=lambda e: (e["year"], e["parish"]))
    return {"entries": entries}


@app.get("/library-entry/{year}/{parish}")
def library_entry(year: int, parish: str):
    """Return the cached extraction JSON for a specific parish+year."""
    runs_root = _app_dir() / "runs"
    search_dirs = []
    if runs_root.exists():
        search_dirs += sorted(
            [d / "output" / "cache" for d in runs_root.iterdir()],
            key=lambda p: p.stat().st_mtime if p.exists() else 0, reverse=True
        )
    flat = _app_dir() / "output" / str(year) / "cache"
    search_dirs.append(flat)
    for cache_dir in search_dirs:
        jf = cache_dir / f"{parish}_{year}.json"
        if jf.exists():
            try:
                return json.loads(jf.read_text())
            except Exception as e:
                return {"error": str(e)}
    return {"error": "not found"}


@app.get("/library-year/{year}")
def library_year(year: int):
    """Return all parish data for a year in one request — avoids N parallel fetches from browser."""
    cache_dir = _find_best_cache(year)
    result = {}
    if cache_dir and cache_dir.exists():
        for jf in cache_dir.glob(f"*_{year}.json"):
            parish = jf.stem.rsplit("_", 1)[0]
            try:
                result[parish] = json.loads(jf.read_text())
            except Exception:
                pass
    return result


def _read_run_meta(job_dir: Path) -> dict:
    """Read meta.json for a run, falling back to parsing the log for legacy runs."""
    meta_path = job_dir / "meta.json"
    if meta_path.exists():
        try:
            return json.loads(meta_path.read_text())
        except Exception:
            pass
    # Legacy: derive from log
    meta = {}
    log_path = job_dir / "output" / "server.log"
    if log_path.exists():
        for line in log_path.read_text(errors="replace").splitlines():
            m = re.search(r"Processing \d+ parishes for year (\d+)", line)
            if m:
                meta["year"] = int(m.group(1))
                break
        # Use log mtime as started_at proxy
        from datetime import datetime as _dt
        mtime = _dt.fromtimestamp(log_path.stat().st_mtime)
        meta.setdefault("started_fmt", mtime.strftime("%b %d %Y, %I:%M %p").replace(" 0", " "))
    if meta.get("year"):
        meta.setdefault("label", f"{meta['year']} run")
    return meta


@app.get("/jobs/active")
def jobs_active():
    """Return any currently running or recently completed pipeline jobs."""
    result = []
    for jid, info in JOBS.items():
        out_dir = Path(info.get("output_dir", ""))
        cache_dir = out_dir / "cache"
        done_count = len(list(cache_dir.glob("*.json"))) if cache_dir.exists() else 0
        pdf_dir = Path(info.get("pdf_dir", ""))
        total = len(list(pdf_dir.glob("*.pdf"))) if pdf_dir.exists() else 0
        job_dir = out_dir.parent
        meta = _read_run_meta(job_dir)
        result.append({
            "job_id":      jid,
            "status":      info.get("status", "running"),
            "done":        done_count,
            "total":       total,
            "year":        info.get("year") or meta.get("year"),
            "label":       info.get("label") or meta.get("label", jid),
            "started_fmt": info.get("started_fmt") or meta.get("started_fmt", ""),
        })
    result.sort(key=lambda x: x["done"], reverse=True)
    return {"jobs": result}

@app.get("/job-log/{job_id}")
def job_log(job_id: str):
    """Return the contents of the server log for a job (for reconnect replay)."""
    if job_id not in JOBS: raise HTTPException(404)
    out_dir = Path(JOBS[job_id]["output_dir"])
    log_path = out_dir / "server.log"
    if not log_path.exists():
        return {"lines": []}
    lines = log_path.read_text(errors="replace").splitlines()
    return {"lines": lines, "status": JOBS[job_id].get("status","running")}


@app.get("/job-status/{job_id}")
def job_status(job_id: str):
    """
    Lightweight status poll used when the SSE stream is disconnected.
    Works even after a server restart by reading the log file from disk.
    Returns: {status, done_count, total_count, last_line}
    """
    # --- Try in-memory first (server hasn't restarted) ---
    if job_id in JOBS:
        info = JOBS[job_id]
        out_dir = Path(info["output_dir"])
        cache_dir = out_dir / "cache"
        done = len(list(cache_dir.glob("*.json"))) if cache_dir.exists() else 0
        log_path = out_dir / "server.log"
        last = ""
        total = 0
        if log_path.exists():
            lines = [l for l in log_path.read_text(errors="replace").splitlines() if l.strip()]
            last = lines[-1] if lines else ""
            import re as _re
            for l in reversed(lines):
                m = _re.search(r'\[(\d+)/(\d+)\]', l)
                if m: total = int(m.group(2)); break
        return {"status": info.get("status","running"), "done_count": done,
                "total_count": total or done, "last_line": last,
                "label": info.get("label", job_id),
                "started_fmt": info.get("started_fmt", "")}

    # --- Server restarted — derive status from disk ---
    runs_root = _app_dir() / "runs"
    out_dir = runs_root / job_id / "output"
    if not out_dir.exists():
        raise HTTPException(404, "Job not found")
    log_path = out_dir / "server.log"
    cache_dir = out_dir / "cache"
    done = len(list(cache_dir.glob("*.json"))) if cache_dir.exists() else 0
    if not log_path.exists():
        return {"status": "unknown", "done_count": done, "total_count": 0, "last_line": ""}
    lines = [l for l in log_path.read_text(errors="replace").splitlines() if l.strip()]
    last = lines[-1] if lines else ""
    import re as _re
    total = 0
    for l in reversed(lines):
        m = _re.search(r'\[(\d+)/(\d+)\]', l)
        if m: total = int(m.group(2)); break
    # Infer status from log content
    if last.strip() == "Done.":
        status = "done"
    elif "[ERROR]" in last or "Traceback" in last:
        status = "error"
    else:
        # Check if log file was written to recently (within 60s → likely still running)
        import time as _time
        age = _time.time() - log_path.stat().st_mtime
        status = "running" if age < 60 else "stopped"
    return {"status": status, "done_count": done, "total_count": total or done, "last_line": last}

@app.get("/progress/{job_id}")
def progress(job_id: str):
    if job_id not in JOBS: raise HTTPException(404)
    def stream():
        q = JOBS[job_id]["queue"]
        while True:
            try:   msg = q.get(timeout=25)
            except queue.Empty: yield ": keepalive\n\n"; continue
            yield f"data: {json.dumps(msg)}\n\n"
            if msg == "__DONE__": break
    return StreamingResponse(stream(), media_type="text/event-stream",
                             headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no"})

@app.get("/files/{job_id}")
def list_files(job_id: str):
    if job_id not in JOBS: raise HTTPException(404)
    out_dir = Path(JOBS[job_id]["output_dir"]).resolve()
    all_files = sorted(f.name for f in out_dir.glob("*.xlsx")) if out_dir.exists() else []
    print(f"[{job_id}] /files → {out_dir} | found: {all_files}", flush=True)
    stmt_suffixes = (" CBS.xlsx", " SOA.xlsx", " SONA.xlsx", " Capital Assets.xlsx")
    return {"by_statement": [f for f in all_files if any(f.endswith(s) for s in stmt_suffixes)],
            "by_parish":    [f for f in all_files if not any(f.endswith(s) for s in stmt_suffixes)],
            "status": JOBS[job_id]["status"]}

@app.get("/download/{job_id}/{filename}")
def download(job_id: str, filename: str):
    if job_id not in JOBS: raise HTTPException(404)
    path = Path(JOBS[job_id]["output_dir"])/filename
    if not path.exists(): raise HTTPException(404)
    return FileResponse(str(path), filename=filename)

@app.get("/uncertainties/{job_id}")
def uncertainties(job_id: str):
    if job_id not in JOBS: raise HTTPException(404)
    try:
        _pipeline_dir = str(_resource_dir() / "pipeline")
        if _pipeline_dir not in sys.path:
            sys.path.insert(0, _pipeline_dir)
        from feedback import get_uncertainties_for_job
        db_path = str(_app_dir() / "feedback.db")
        return {"items": get_uncertainties_for_job(db_path, job_id)}
    except Exception as e:
        return {"items": [], "error": str(e)}

@app.post("/feedback/{uncertainty_id}")
def submit_feedback(uncertainty_id: int, payload: dict = _Body(...)):
    try:
        _pipeline_dir = str(_resource_dir() / "pipeline")
        if _pipeline_dir not in sys.path:
            sys.path.insert(0, _pipeline_dir)
        from feedback import save_correction, get_uncertainties_for_job
        import sqlite3 as _sq
        db_path = str(_app_dir() / "feedback.db")
        # Get uncertainty details for denorm columns
        c = _sq.connect(db_path); c.row_factory = _sq.Row
        row = c.execute("SELECT * FROM uncertainties WHERE id=?", (uncertainty_id,)).fetchone()
        c.close()
        if not row: raise HTTPException(404, "Uncertainty not found")
        save_correction(
            db_path, uncertainty_id,
            row["parish"], row["year"], row["statement_type"], row["field_path"],
            action=payload.get("action", "confirmed"),
            corrected_value=payload.get("corrected_value"),
            note=payload.get("note"),
        )
        return {"ok": True}
    except HTTPException: raise
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/statement-note")
def save_statement_note_endpoint(payload: dict = _Body(...)):
    """Save a free-text statement-level note for a parish/year/stmt_type to inject into future prompts."""
    try:
        _pipeline_dir = str(_resource_dir() / "pipeline")
        if _pipeline_dir not in sys.path:
            sys.path.insert(0, _pipeline_dir)
        from feedback import save_statement_note
        parish = payload.get("parish", "")
        year = int(payload.get("year", 0))
        stmt_type = payload.get("stmt_type", "")
        note = (payload.get("note") or "").strip()
        if not (parish and year and stmt_type and note):
            raise HTTPException(400, "parish, year, stmt_type, and note are required")
        db_path = str(_app_dir() / "feedback.db")
        save_statement_note(db_path, parish, year, stmt_type, note)
        return {"ok": True}
    except HTTPException: raise
    except Exception as e:
        raise HTTPException(500, str(e))

from fastapi.responses import Response as _Resp
from typing import Optional as _Opt
@app.get("/page-image/{job_id}/{parish}/{page}")
def page_image(job_id: str, parish: str, page: int,
               value: _Opt[str] = None, snippet: _Opt[str] = None):
    if job_id not in JOBS: raise HTTPException(404)
    pdf_dir = Path(JOBS[job_id].get("pdf_dir", ""))
    if not pdf_dir.exists(): raise HTTPException(404, "PDF dir not found")
    pdf_path = next((str(f) for f in pdf_dir.glob("*.pdf")
                     if get_parish_name(str(f)) == parish), None)
    if not pdf_path: raise HTTPException(404, f"No PDF for {parish}")
    page_0 = page - 1  # convert 1-indexed → 0-indexed
    # Find highlight boxes if caller passed a value or snippet
    boxes = []
    if value or snippet:
        try:
            boxes = find_text_boxes(pdf_path, page_0, value=value, snippet=snippet)
        except Exception as e:
            print(f"[page-image] highlight search failed: {e}", flush=True)
    img = render_pdf_page(pdf_path, page_0, highlight_boxes=boxes or None)
    if not img: raise HTTPException(500, "Page render failed")
    return _Resp(content=img, media_type="image/png")

# ══════════════════════════════════════════════════════════════════════════════
# TRAINING / COMPARISON ENDPOINTS
# ══════════════════════════════════════════════════════════════════════════════

def _find_best_cache(year: int) -> Path:
    """Return the cache directory with the most JSON files for the given year.
    Considers all locations — flat legacy dir AND every run dir — and picks
    whichever has the highest file count (not just the first non-empty one)."""
    candidates: list[tuple[int, Path]] = []

    flat = _app_dir() / "output" / str(year) / "cache"
    if flat.exists():
        count = sum(1 for _ in flat.glob(f"*_{year}.json"))
        if count:
            candidates.append((count, flat))

    runs_root = _app_dir() / "runs"
    if runs_root.exists():
        for job_dir in runs_root.iterdir():
            candidate = job_dir / "output" / "cache"
            if not candidate.exists():
                continue
            count = sum(1 for _ in candidate.glob(f"*_{year}.json"))
            if count:
                candidates.append((count, candidate))

    if candidates:
        return max(candidates, key=lambda x: x[0])[1]
    return flat  # fall back even if empty

_COMPARE_JOBS: dict[str, dict] = {}  # job_id -> {status, summary, error}
_CODED_STORE: dict[int, str] = {}    # year -> path to coded excel dir


@app.post("/admin/import-run")
async def import_run(files: list[UploadFile] = File(...)):
    """Accept a mixed batch of files from a local run folder and sort them into the right places."""
    import datetime as _dt
    ts = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    run_id = f"imported_{ts}"
    run_dir = _app_dir() / "runs" / run_id
    pdf_dir   = run_dir / "pdfs"
    out_dir   = run_dir / "output"
    cache_dir = out_dir / "cache"
    for d in (pdf_dir, out_dir, cache_dir):
        d.mkdir(parents=True, exist_ok=True)

    counts = {"pdfs": 0, "excels": 0, "cache": 0, "other": 0}
    for f in files:
        data = await f.read()
        name = f.filename or ""
        lo = name.lower()
        if lo.endswith(".pdf"):
            (pdf_dir / name).write_bytes(data)
            counts["pdfs"] += 1
        elif lo == "meta.json":
            (run_dir / name).write_bytes(data)
        elif lo.endswith(".json"):
            (cache_dir / name).write_bytes(data)
            counts["cache"] += 1
        elif lo.endswith(".xlsx"):
            (out_dir / name).write_bytes(data)
            counts["excels"] += 1
        else:
            counts["other"] += 1

    # Write a minimal meta.json if one wasn't included
    meta_path = run_dir / "meta.json"
    if not meta_path.exists():
        import datetime as _dt2
        now = _dt2.datetime.now()
        meta_path.write_text(json.dumps({
            "job_id": run_id,
            "label": f"Imported {now.strftime('%b %d %Y')}",
            "started_fmt": now.strftime("%b %d %Y, %I:%M %p").replace(" 0", " "),
        }))

    return {**counts, "run_id": run_id}


@app.get("/excel-view/{year}/{parish}/{stmt_type}")
def excel_view(year: int, parish: str, stmt_type: str):
    """Return both truth and pipeline Excel rows for side-by-side display."""
    import openpyxl as _xl
    stmt_map = {"cbs": "CBS", "soa": "SOA", "sona": "SONA", "ca": "Capital Assets"}
    suffix = stmt_map.get(stmt_type.lower())
    if not suffix:
        raise HTTPException(400, "Unknown statement type")
    filename = f"Louisiana {year} {suffix}.xlsx"

    def _read_parish_sheet(path: Path):
        if not path or not path.exists():
            return None
        try:
            wb = _xl.load_workbook(path, data_only=True)
            # Find sheet matching parish name (case-insensitive, space-insensitive)
            norm = lambda s: s.lower().replace(" ", "").replace(".", "")
            target = norm(parish)
            sheet = next((wb[n] for n in wb.sheetnames if norm(n) == target), None)
            if not sheet:
                # Try partial match
                sheet = next((wb[n] for n in wb.sheetnames if target in norm(n)), None)
            if not sheet:
                return None
            rows = []
            for row in sheet.iter_rows():
                rows.append([cell.value for cell in row])
            # Trim trailing empty rows
            while rows and all(v is None for v in rows[-1]):
                rows.pop()
            return rows
        except Exception as e:
            print(f"[excel-view] error reading {path}: {e}", flush=True)
            return None

    # Truth file
    truth_path = _app_dir() / "coded" / str(year) / filename
    truth_rows = _read_parish_sheet(truth_path)

    # Pipeline output — search runs dirs most recent first
    pipeline_rows = None
    runs_root = _app_dir() / "runs"
    if runs_root.exists():
        for job_dir in sorted(runs_root.iterdir(), reverse=True):
            p = job_dir / "output" / filename
            if p.exists():
                pipeline_rows = _read_parish_sheet(p)
                if pipeline_rows is not None:
                    break

    # Serialize (convert non-string cell values)
    def _ser(rows):
        if rows is None:
            return None
        out = []
        for row in rows:
            out.append([str(v) if v is not None else None for v in row])
        return out

    return {"truth": _ser(truth_rows), "pipeline": _ser(pipeline_rows), "filename": filename}


def _ensure_pipeline_dir():
    d = str(_resource_dir() / "pipeline")
    if d not in sys.path:
        sys.path.insert(0, d)


@app.post("/upload-coded")
async def upload_coded(year: int = Form(...), files: list[UploadFile] = File(...)):
    """Accept coded Excel files, store them, kick off comparison job."""
    _ensure_pipeline_dir()
    job_id = f"compare_{year}_{uuid.uuid4().hex[:8]}"
    coded_dir = _app_dir() / "coded" / str(year)
    coded_dir.mkdir(parents=True, exist_ok=True)

    for f in files:
        data = await f.read()
        (coded_dir / f.filename).write_bytes(data)
        print(f"[compare] Stored coded: {f.filename}", flush=True)

    _CODED_STORE[year] = str(coded_dir)
    q: queue.Queue[str] = queue.Queue()
    _COMPARE_JOBS[job_id] = {"status": "running", "summary": None, "error": None, "queue": q}

    def _cb(m):
        print(m, flush=True)
        q.put(m)

    def _run():
        try:
            from coded_compare import run_comparison
            cache_dir = _find_best_cache(year)
            db_path = str(_app_dir() / "feedback.db")
            summary = run_comparison(
                coded_excel_dir=str(coded_dir),
                cache_dir=str(cache_dir),
                db_path=db_path,
                year=year,
                job_id=job_id,
                progress_callback=_cb,
            )
            _COMPARE_JOBS[job_id]["status"] = "done"
            _COMPARE_JOBS[job_id]["summary"] = summary
        except Exception as exc:
            import traceback
            _cb(f"[ERROR] {exc}")
            print(f"[compare] ERROR: {traceback.format_exc()}", flush=True)
            _COMPARE_JOBS[job_id]["status"] = "error"
            _COMPARE_JOBS[job_id]["error"] = str(exc)
        finally:
            q.put("__DONE__")

    threading.Thread(target=_run, daemon=True).start()
    return {"job_id": job_id, "year": year}


@app.get("/compare-status/{job_id}")
def compare_status(job_id: str):
    if job_id not in _COMPARE_JOBS:
        raise HTTPException(404)
    return _COMPARE_JOBS[job_id]


@app.get("/compare-progress/{job_id}")
def compare_progress(job_id: str):
    if job_id not in _COMPARE_JOBS:
        raise HTTPException(404)
    def stream():
        q = _COMPARE_JOBS[job_id]["queue"]
        while True:
            try:
                msg = q.get(timeout=25)
            except queue.Empty:
                yield ": keepalive\n\n"
                continue
            yield f"data: {json.dumps(msg)}\n\n"
            if msg == "__DONE__":
                break
    return StreamingResponse(stream(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.post("/rerun-compare/{year}")
def rerun_compare(year: int):
    """Re-run comparison for a year using previously uploaded coded files."""
    _ensure_pipeline_dir()
    coded_dir = _app_dir() / "coded" / str(year)
    if not coded_dir.exists():
        raise HTTPException(404, f"No coded files found for {year}. Please upload them first.")

    job_id = f"compare_{year}_{uuid.uuid4().hex[:8]}"
    q: queue.Queue[str] = queue.Queue()
    _COMPARE_JOBS[job_id] = {"status": "running", "summary": None, "error": None, "queue": q}

    def _cb(m):
        print(m, flush=True)
        q.put(m)

    def _run():
        try:
            from coded_compare import run_comparison
            cache_dir = _find_best_cache(year)
            db_path = str(_app_dir() / "feedback.db")
            summary = run_comparison(
                coded_excel_dir=str(coded_dir),
                cache_dir=str(cache_dir),
                db_path=db_path,
                year=year,
                job_id=job_id,
                progress_callback=_cb,
            )
            _COMPARE_JOBS[job_id]["status"] = "done"
            _COMPARE_JOBS[job_id]["summary"] = summary
        except Exception as exc:
            import traceback
            _cb(f"[ERROR] {exc}")
            print(f"[compare] ERROR: {traceback.format_exc()}", flush=True)
            _COMPARE_JOBS[job_id]["status"] = "error"
            _COMPARE_JOBS[job_id]["error"] = str(exc)
        finally:
            q.put("__DONE__")

    threading.Thread(target=_run, daemon=True).start()
    return {"job_id": job_id}


@app.get("/training-stats")
def training_stats():
    _ensure_pipeline_dir()
    try:
        from feedback import get_training_stats, init_db
        db_path = str(_app_dir() / "feedback.db")
        init_db(db_path)
        return {"years": get_training_stats(db_path)}
    except Exception as e:
        return {"years": [], "error": str(e)}


@app.get("/compare-flags/{year}")
def compare_flags(year: int):
    _ensure_pipeline_dir()
    try:
        from feedback import init_db
        import sqlite3 as _sq
        db_path = str(_app_dir() / "feedback.db")
        init_db(db_path)
        c = _sq.connect(db_path); c.row_factory = _sq.Row
        # Only show flags from the most recent comparison job for this year
        latest = c.execute("""
            SELECT job_id FROM uncertainties
            WHERE year=? AND source IN ('comparison','coding_check')
            ORDER BY id DESC LIMIT 1
        """, (year,)).fetchone()
        if not latest:
            c.close()
            return {"items": []}
        job_id = latest["job_id"]
        rows = c.execute("""
            SELECT u.*,
                   cor.action AS cor_action,
                   cor.corrected_value AS cor_value,
                   cor.user_note AS cor_note
            FROM uncertainties u
            LEFT JOIN corrections cor ON cor.uncertainty_id = u.id
            WHERE u.job_id=? AND u.source IN ('comparison','coding_check')
            ORDER BY u.parish, u.source, u.severity DESC, u.id
        """, (job_id,)).fetchall()
        c.close()
        return {"items": [dict(r) for r in rows]}
    except Exception as e:
        return {"items": [], "error": str(e)}


def _find_parish_pdf(year: int, parish: str):
    """
    Search for a PDF matching the given parish name. Checks in order:
    1. pdfs/<year>/  — explicitly uploaded training PDFs
    2. runs/*/pdfs/  — PDFs stored during web UI pipeline runs
    3. Source folders next to the app
    Matching is fuzzy: normalise both sides by lowercasing and stripping
    spaces, underscores, hyphens, digits and common suffixes so that
    "Bossier 2013.pdf", "bossier_parish_2013_cafr.pdf" all match parish "Bossier".
    """
    import re as _re
    def _norm(s: str) -> str:
        s = s.lower()
        # remove year digits (4-digit sequences)
        s = _re.sub(r'\b\d{4}\b', '', s)
        # remove common non-name words
        for word in ('parish', 'cafr', 'audit', 'report', 'annual', 'financial',
                     'statements', 'statement', 'reissue', 'revised', 'final'):
            s = _re.sub(rf'\b{word}\b', '', s)
        # collapse punctuation/whitespace
        s = _re.sub(r'[\s_\-\.]+', '', s)
        return s.strip()

    target = _norm(parish)

    def _pdf_matches(path: Path) -> bool:
        stem = _norm(get_parish_name(str(path)))
        return stem == target or target in stem or stem in target

    # 1. Training PDF store (most specific — year-matched)
    training_pdf_dir = _app_dir() / "pdfs" / str(year)
    if training_pdf_dir.exists():
        match = next((str(f) for f in training_pdf_dir.glob("*.pdf") if _pdf_matches(f)), None)
        if match:
            return match

    # 2. Web UI run directories (any year, most recent first)
    runs_root = _app_dir() / "runs"
    if runs_root.exists():
        for job_dir in sorted(runs_root.iterdir(), reverse=True):
            pdf_dir = job_dir / "pdfs"
            if not pdf_dir.exists():
                continue
            match = next((str(f) for f in pdf_dir.glob("*.pdf") if _pdf_matches(f)), None)
            if match:
                return match

    # 3. Source folders next to the app: <year>_source/<year>/ and <year>/
    extra_dirs = [
        _app_dir() / f"{year}_source" / str(year),
        _app_dir() / f"{year}_source",
        _app_dir() / str(year),
    ]
    for src_dir in extra_dirs:
        if not src_dir.exists():
            continue
        match = next((str(f) for f in src_dir.glob("*.pdf") if _pdf_matches(f)), None)
        if match:
            return match

    return None


@app.post("/upload-source-pdfs")
async def upload_source_pdfs(year: int = Form(...), files: list[UploadFile] = File(...)):
    """Store source audit PDFs for a year so the training PDF browser can access them."""
    pdf_dir = _app_dir() / "pdfs" / str(year)
    pdf_dir.mkdir(parents=True, exist_ok=True)
    parishes = []
    for f in files:
        data = await f.read()
        dest = pdf_dir / f.filename
        dest.write_bytes(data)
        parishes.append(get_parish_name(f.filename))
        print(f"[training-pdfs] Stored {f.filename} ({len(data):,} bytes) for {year}", flush=True)
    return {"count": len(files), "parishes": parishes, "year": year}


@app.get("/pdf-info/{year}/{parish}")
def pdf_info(year: int, parish: str, value: _Opt[str] = None):
    """Return total page count + best matching page for a value (used by training PDF browser)."""
    pdf_path = _find_parish_pdf(year, parish)
    if not pdf_path:
        return {"total_pages": 0, "found": False,
                "error": f"No PDF found for {parish} {year}. Upload it via the pipeline first."}
    try:
        import pdfplumber as _plb
        with _plb.open(pdf_path) as pdf:
            total = len(pdf.pages)
            if not value:
                return {"total_pages": total, "found": False, "best_page": 1}
            # Scan all pages; return first page that contains the value
            for i in range(total):
                boxes = find_text_boxes(pdf_path, i, value=value)
                if boxes:
                    return {"total_pages": total, "found": True,
                            "best_page": i + 1, "value_box": boxes[0]}
            return {"total_pages": total, "found": False, "best_page": 1}
    except Exception as e:
        return {"total_pages": 0, "found": False, "error": str(e)}


@app.get("/pdf-browse/{year}/{parish}/{page}")
def pdf_browse_page(year: int, parish: str, page: int,
                    value: _Opt[str] = None,
                    trace: int = 0,
                    spread: int = 0):
    """
    Render a PDF page for the training tab PDF browser.
    trace=1  → draw blue row-label and green column-header connector lines
    spread=1 → also render the next page stitched below (for tables that continue)
    """
    pdf_path = _find_parish_pdf(year, parish)
    if not pdf_path:
        raise HTTPException(404, f"No PDF found for {parish} {year}")
    page_0 = page - 1
    boxes, trace_row, trace_col, second_page = [], [], [], None

    if value:
        try:
            boxes = find_text_boxes(pdf_path, page_0, value=value)
        except Exception as e:
            print(f"[pdf-browse] highlight failed: {e}", flush=True)
        if trace and boxes:
            try:
                labels   = find_row_col_labels(pdf_path, page_0, boxes[0])
                trace_row = labels.get("row_boxes", [])
                trace_col = labels.get("col_boxes", [])
            except Exception as e:
                print(f"[pdf-browse] trace failed: {e}", flush=True)

    if spread:
        try:
            import pdfplumber as _plb
            with _plb.open(pdf_path) as _pdf:
                total_pages = len(_pdf.pages)
                if page_0 + 1 < total_pages:
                    second_page = page_0 + 1
        except Exception:
            pass

    img = render_pdf_page(
        pdf_path, page_0,
        highlight_boxes   = boxes     or None,
        trace_row_boxes   = trace_row or None,
        trace_col_boxes   = trace_col or None,
        second_page_0indexed = second_page,
    )
    if not img:
        raise HTTPException(500, "Page render failed")
    return _Resp(content=img, media_type="image/png")


def _stmt_full_name(stmt_type: str) -> str:
    return {
        "cbs":  "Combined Balance Sheet of Governmental Funds",
        "sona": "Statement of Net Position / Statement of Net Assets",
        "soa":  "Statement of Activities",
        "ca":   "Capital Assets Schedule",
    }.get(stmt_type, stmt_type.upper())


@app.post("/diagnose-flag/{uncertainty_id}")
def diagnose_flag(uncertainty_id: int):
    """
    Ask Claude to diagnose a specific comparison flag:
      - locate both the pipeline value AND the coded-Excel value in the PDF text
      - explain in plain English where each came from
      - give a verdict (which is correct) and a second-guess value if applicable
    Returns JSON: {explanation, pipeline_source, coded_source,
                   verdict, verdict_reason, second_guess, second_guess_source, confidence}
    """
    try:
        import sqlite3 as _sq
        db_path = str(_app_dir() / "feedback.db")
        conn = _sq.connect(db_path); conn.row_factory = _sq.Row
        row = conn.execute("SELECT * FROM uncertainties WHERE id=?",
                           (uncertainty_id,)).fetchone()
        conn.close()
        if not row:
            raise HTTPException(404, "Flag not found")

        flag       = dict(row)
        parish     = flag["parish"]
        year       = flag["year"]
        stmt_type  = flag["statement_type"]
        field_path = flag["field_path"] or ""
        extracted  = flag["extracted_value"]  or "null (not extracted)"
        alternative= flag["alternative_value"] or "null"
        page_num   = flag["page_number"] or 0
        reason     = flag["reason"] or ""

        # ── Pull relevant PDF pages ──────────────────────────────────────────
        pdf_path = _find_parish_pdf(year, parish)
        pdf_text = ""
        if pdf_path:
            try:
                import pdfplumber as _plb
                with _plb.open(pdf_path) as _pdf:
                    total = len(_pdf.pages)
                    # Candidate pages: known page ± 1, or first 6 if unknown
                    if page_num > 0:
                        lo = max(0, page_num - 2)
                        hi = min(total, page_num + 2)
                    else:
                        lo, hi = 0, min(total, 6)
                    for pi in range(lo, hi):
                        t = _pdf.pages[pi].extract_text() or ""
                        if t.strip():
                            pdf_text += f"\n--- PAGE {pi+1} ---\n{t}\n"
            except Exception as _e:
                print(f"[diagnose] PDF read: {_e}")

        # ── Build prompt ─────────────────────────────────────────────────────
        field_readable = field_path.replace(".", " → ").replace("_", " ")
        stmt_full      = _stmt_full_name(stmt_type)
        no_pdf_note    = "" if pdf_text else (
            "\n[NOTE: No PDF text could be extracted — the document may be scanned/image-only. "
            "Base your diagnosis on the field name, the two values, and general knowledge of "
            "Louisiana CAFR formatting.]\n"
        )

        prompt = f"""You are an expert Louisiana government finance auditor diagnosing a specific data discrepancy.

CONTEXT
  Parish       : {parish}
  Year         : {year}
  Statement    : {stmt_full}  [{stmt_type.upper()}]
  Field        : {field_readable}
  Flagged why  : {reason}

DISCREPANCY
  Pipeline extracted : {extracted}
  Coded Excel shows  : {alternative}
{no_pdf_note}
PDF TEXT (pages closest to where this value appears):
{pdf_text[:9000] if pdf_text else '[No extractable text]'}

YOUR TASK
1. Find the number that the pipeline likely extracted and the number the coded Excel likely used.
2. Determine if this is a comparative-statement issue (two fiscal years side-by-side), a column-selection error, a thousands-scaling error, or a genuine audit difference.
3. State which value you believe is correct and why, with a confidence level.
4. Provide a second-best numeric guess if the situation is ambiguous.

Return ONLY valid JSON — no markdown, no preamble:
{{
  "explanation": "<2-5 sentence plain-English explanation of the discrepancy for an auditor>",
  "pipeline_source": "<quote or describe the exact line/column the pipeline likely read>",
  "coded_source": "<quote or describe where the coded Excel value comes from>",
  "verdict": "pipeline_correct" | "coded_correct" | "uncertain",
  "verdict_reason": "<one sentence: why you chose this verdict>",
  "second_guess": <number or null>,
  "second_guess_source": "<where the second guess comes from, or null>",
  "confidence": "high" | "medium" | "low"
}}"""

        resp = _get_client().messages.create(
            model="claude-haiku-4-5",
            max_tokens=1200,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = resp.content[0].text.strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"): raw = raw[4:]
        result = json.loads(raw.strip())
        return result

    except HTTPException:
        raise
    except json.JSONDecodeError as e:
        return {"explanation": "Diagnosis returned malformed JSON — try again.",
                "verdict": "uncertain", "confidence": "low", "pipeline_source": "",
                "coded_source": "", "verdict_reason": "", "second_guess": None,
                "second_guess_source": None}
    except Exception as e:
        raise HTTPException(500, str(e))


# ══════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import uvicorn
    _port = int(os.environ.get("PORT", 8000))
    _host = "0.0.0.0" if os.environ.get("WEBSITE_SITE_NAME") else "127.0.0.1"
    if not _get_api_key() and not os.environ.get("ANTHROPIC_API_KEY"):
        print("No API key found. Open the app and click the gear icon to add your Anthropic key.")
    if _host == "127.0.0.1":
        threading.Thread(target=lambda: (time.sleep(1.5), webbrowser.open(f"http://localhost:{_port}")), daemon=True).start()
    print(f"Parish Audit Pipeline @ http://{'localhost' if _host == '127.0.0.1' else '0.0.0.0'}:{_port}  (Ctrl+C to stop)")
    uvicorn.run(app, host=_host, port=_port, log_level="info")
